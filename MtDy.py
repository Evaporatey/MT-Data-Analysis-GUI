# author: Ye Yang
# MT Dynamics Analysis with HMM
# unicode: utf-8

# 标准库导入
import math
import os
import sys
import datetime  # 添加 datetime 模块导入
from functools import lru_cache
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor

# 第三方库导入
import matplotlib.pyplot as plt
import nptdms as nt
import numpy as np
import openpyxl
import pandas as pd
import scipy
from hmmlearn import hmm
from scipy.signal import find_peaks, peak_prominences

# PySide6导入
from PySide6.QtCore import QRect, QFileInfo, Qt
from PySide6.QtWidgets import (
    QApplication, QButtonGroup, QCheckBox, QComboBox, QDialog, QDialogButtonBox,
    QFileDialog, QGroupBox, QGridLayout, QHBoxLayout, QLabel, QLineEdit, QMainWindow, 
    QMenu, QMenuBar, QMessageBox, QPushButton, QRadioButton, QSizePolicy, 
    QSpinBox, QStatusBar, QVBoxLayout, QWidget, QTabWidget, QFormLayout
)
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg, NavigationToolbar2QT

# 本地模块导入
from tdms_reader import read_tdms_chunk, read_tdms_file


class KineticsAnalysis(QWidget):
    def __init__(self, data_for_figure):
        super().__init__()
        
        # 初始化数据属性
        self._initialize_data_attributes()
        
        # 继承传递的数据
        self._load_inherited_data(data_for_figure)
        
        # 加载TDMS数据
        self._load_tdms_data()
        
        # 创建UI界面（先于图形设置）
        self._setup_ui()
        
        # 创建图形对象
        self._setup_figure()
        
        # 首次绘图
        self.plotfig()
    
    def _initialize_data_attributes(self):
        """初始化所有数据属性"""
        self.bead_data_list = None
        self.time_data = None
        self.time_data_list = None
        self.time_point_end = None
        self.time_point_end_data = None
        self.time_point_start_data = None
        self.time_point_start = None
        self.bead_data = None
        self.base_line = None
        self.bead_array = None
        self.bin_data = None
        self.bin_data_diff = None
        self.extension_start_point = None
        self.extension_end_point = None
        self.time_start = None
        self.time_end = None
        self.above_time_segment = None
        self.above_time_data = []
        self.below_time_segment = None
        self.below_time_data = []
        self.current_data_segment = None
        self.force_regions = []
        self.region_highlight_rectangle = None
        self.min_region_duration = 1.0  # 默认最小区域持续时间（秒）
    
    def _load_inherited_data(self, data_for_figure):
        """加载从父窗口传递的数据"""
        self.Data_Saved_Path = data_for_figure['Data_Saved_Path']
        self.file_name = data_for_figure['file_name']
        self.file_type = data_for_figure['file_type']
        self.file_info = data_for_figure['self.file_info']
        self.base_name = data_for_figure['base_name']
    
    def _load_tdms_data(self):
        """加载并处理TDMS数据"""
        # 获取TDMS数据并转换为DataFrame
        self.tdms_data_frame = read_tdms_file(self.file_name, need_force=True)
        self.tdms_data_store = self.tdms_data_frame
        
        # 获取磁铁移动状态数据
        beads_list = self.tdms_data_frame.columns.values.tolist()
        str_magnet_move = str(beads_list[-1])
        magnet_move_state = self.tdms_data_store[str(str_magnet_move)].values.tolist()
        
        # 处理磁铁移动数据
        new_magnet_move_state = [i for i in magnet_move_state if not math.isnan(i)]
        self.new_int_magnet_move_state = [int(i) for i in new_magnet_move_state]
        self.num_of_state = len(self.new_int_magnet_move_state) / 2
        self.final_slice_magnet_move_state = [
            self.new_int_magnet_move_state[i:i+2] 
            for i in range(0, len(self.new_int_magnet_move_state), 2)
        ]
        self.num_of_sliced_data = len(self.final_slice_magnet_move_state)
        self.beads_list = beads_list
    
    def _create_widget(self, parent, objectName="", sizePolicy=None):
        """创建标准widget的辅助方法"""
        widget = QWidget(parent)
        if objectName:
            widget.setObjectName(objectName)
        if sizePolicy:
            widget.setSizePolicy(*sizePolicy)
        return widget
    
    def _create_layout(self, parent, layout_type=QHBoxLayout, objectName="", margins=None):
        """创建布局的辅助方法"""
        layout = layout_type(parent)
        if objectName:
            layout.setObjectName(objectName)
        if margins is not None:
            layout.setContentsMargins(*margins)
        return layout
    
    def _create_label(self, parent, text="", objectName="", fixed_size=True):
        """创建标签的辅助方法"""
        label = QLabel(parent)
        if objectName:
            label.setObjectName(objectName)
        label.setText(text)
        if fixed_size:
            label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        return label
    
    def _create_combobox(self, parent, items=None, objectName="", fixed_size=True, 
                         enabled=True, connection=None):
        """创建下拉选择框的辅助方法"""
        combobox = QComboBox(parent)
        if objectName:
            combobox.setObjectName(objectName)
        if fixed_size:
            combobox.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        if items:
            combobox.addItems(items)
        combobox.setEnabled(enabled)
        if connection:
            combobox.currentTextChanged.connect(connection)
        return combobox
    
    def _create_checkbox(self, parent, text="", objectName="", checked=False, 
                         fixed_size=True, connection=None):
        """创建复选框的辅助方法"""
        checkbox = QCheckBox(parent)
        if objectName:
            checkbox.setObjectName(objectName)
        checkbox.setText(text)
        checkbox.setChecked(checked)
        if fixed_size:
            checkbox.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        if connection:
            checkbox.stateChanged.connect(connection)
        return checkbox
    
    def _create_button(self, parent, text="", objectName="", fixed_size=True, connection=None):
        """创建按钮的辅助方法"""
        button = QPushButton(parent)
        if objectName:
            button.setObjectName(objectName)
        button.setText(text)
        if fixed_size:
            button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        if connection:
            button.clicked.connect(connection)
        return button
    
    def _setup_ui(self):
        """设置UI界面"""
        # 设置窗口属性
        self.setWindowTitle("MT Dynamics Analysis")
        self.resize(1800, 900)  # 更大的窗口尺寸
        
        # 创建主界面布局为水平布局
        self.main_layout = QHBoxLayout(self)
        self.main_layout.setContentsMargins(5, 5, 5, 5)
        self.main_layout.setSpacing(5)  # 减小间距
        
        # 创建左侧图形面板（可伸缩）
        self.figure_panel = QWidget()
        self.figure_layout = QVBoxLayout(self.figure_panel)
        self.figure_layout.setContentsMargins(0, 0, 0, 0)
        
        # 创建右侧控制面板（宽度固定但更窄）
        self.control_panel = QWidget()
        self.control_panel.setFixedWidth(380)  # 减小控制面板宽度从480px到380px
        self.control_layout = QVBoxLayout(self.control_panel)
        self.control_layout.setSpacing(8)  # 减小垂直间距
        self.control_layout.setContentsMargins(5, 5, 5, 5)  # 减小边距
        
        # 创建标签页控件
        self.tab_widget = QTabWidget()
        
        # 创建合并的分析标签页
        combined_tab = self._create_combined_tab()
        self.tab_widget.addTab(combined_tab, "Data Analysis")
        
        # 添加标签页控件到控制面板
        self.control_layout.addWidget(self.tab_widget)
        
        # 添加数据信息显示区域
        self._create_data_info_panel()
        
        # 将面板添加到主布局
        self.main_layout.addWidget(self.figure_panel, 4)  # 图形面板占据更多空间
        self.main_layout.addWidget(self.control_panel, 1)

        style_sheet = """
        QGroupBox {
            font-weight: bold;
            border: 1px solid #CCCCCC;
            border-radius: 5px;
            margin-top: 8px;
            padding-top: 8px;
        }

        QGroupBox::title {
            subcontrol-origin: margin;
            left: 7px;
            padding: 0 3px 0 3px;
        }

        QPushButton {
            border: 1px solid #8f8f91;
            border-radius: 4px;
            background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #f6f7fa, stop:1 #dadbde);
            min-width: 80px;
            padding: 4px;
        }

        QPushButton:pressed {
            background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #dadbde, stop:1 #f6f7fa);
        }

        QLabel {
            font-size: 11px;
        }

        QComboBox, QSpinBox {
            border: 1px solid #CCCCCC;
            border-radius: 3px;
            padding: 1px;
            min-height: 20px;
        }
        """
        self.setStyleSheet(style_sheet)

    def _apply_filter_to_data(self, y_data, filter_type=None, kernel_size=None, poly_order=3):
        """统一的滤波方法，应用到数据
        
        Args:
            y_data (array): 需要滤波的数据
            filter_type (str, optional): 滤波类型，若为None则使用当前UI选择
            kernel_size (int, optional): 滤波器核大小，若为None则使用当前UI设置
            poly_order (int, optional): Savitzky-Golay多项式阶数
                
        Returns:
            array: 滤波后的数据
        """
        if filter_type is None:
            filter_type = self.filterselection.currentText()
        
        if kernel_size is None:
            kernel_size = self.kernel_size_box.value()
        
        # 确保核大小为奇数
        if kernel_size % 2 == 0:
            kernel_size += 1
        
        if filter_type == 'Median Filter':
            return scipy.signal.medfilt(y_data, kernel_size=kernel_size)
        elif filter_type == 'Moving Average':
            kernel = np.ones(kernel_size) / kernel_size
            return np.convolve(y_data, kernel, mode='same')
        elif filter_type == 'Savitzky-Golay':
            # 确保窗口长度足够且多项式阶数小于窗口长度
            if len(y_data) > kernel_size and kernel_size > poly_order:
                return scipy.signal.savgol_filter(y_data, kernel_size, poly_order)
            elif kernel_size <= poly_order:
                # 如果窗口长度小于或等于多项式阶数，调整多项式阶数
                adjusted_poly_order = kernel_size - 1
                print(f"警告: 多项式阶数({poly_order})不能大于等于窗口长度({kernel_size})，已自动调整为{adjusted_poly_order}")
                return scipy.signal.savgol_filter(y_data, kernel_size, adjusted_poly_order)
            else:
                return y_data
        
        return y_data  # 如果没有选择滤波器，返回原始数据

    def _create_combined_tab(self):
        """创建合并后的数据分析标签页 - 优化布局"""
        combined_tab = QWidget()
        combined_layout = QVBoxLayout(combined_tab)
        combined_layout.setSpacing(8)  # 减小垂直间距
        combined_layout.setContentsMargins(5, 8, 5, 8)  # 减小边距
        
        # 1. 数据选择与滤波组
        data_group = QGroupBox("Data Selection & Filtering")
        data_layout = QVBoxLayout(data_group)
        data_layout.setSpacing(6)  # 减小组内垂直间距
        data_layout.setContentsMargins(5, 8, 5, 8)  # 减小边距
        
        # Y轴选择 - 使用FormLayout代替HBoxLayout
        y_form = QFormLayout()
        y_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        y_form.setLabelAlignment(Qt.AlignLeft)
        y_form.setRowWrapPolicy(QFormLayout.DontWrapRows)
        y_form.setFormAlignment(Qt.AlignLeft | Qt.AlignTop)
        y_form.setHorizontalSpacing(10)  # 减小水平间距
        y_form.setVerticalSpacing(5)  # 减小垂直间距
        
        self.y_axis_box = QComboBox()
        y_items = [str(self.beads_list[i]) for i in range(4, (len(self.beads_list) - 1))]
        self.y_axis_box.addItems(y_items)
        self.y_axis_box.currentTextChanged.connect(self.plotfig)
        self.y_axis_box.setMinimumWidth(220)  # 减小宽度从300到220
        
        y_form.addRow("Y Axis:", self.y_axis_box)
        data_layout.addLayout(y_form)
        
        # 滤波器设置 - 使用FormLayout
        filter_form = QFormLayout()
        filter_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        filter_form.setLabelAlignment(Qt.AlignLeft)
        filter_form.setRowWrapPolicy(QFormLayout.DontWrapRows)
        filter_form.setFormAlignment(Qt.AlignLeft | Qt.AlignTop)
        filter_form.setHorizontalSpacing(10)  # 减小水平间距
        filter_form.setVerticalSpacing(5)  # 减小垂直间距
        
        # 滤波类型
        filter_items = ["Moving Average", "Median Filter", "Savitzky-Golay"]
        self.filterselection = QComboBox()
        self.filterselection.addItems(filter_items)
        self.filterselection.currentTextChanged.connect(self.plotfig)
        self.filterselection.currentTextChanged.connect(self._update_poly_order_visibility)
        self.filterselection.setMinimumWidth(150)  # 减小宽度从200到150
        filter_form.addRow("Filter:", self.filterselection)
        
        # 内核大小
        self.kernel_size_box = QSpinBox()
        self.kernel_size_box.setRange(3, 1001)
        self.kernel_size_box.setValue(51)
        self.kernel_size_box.setSingleStep(2)
        self.kernel_size_box.setMinimumWidth(70)  # 减小宽度从100到70
        self.kernel_size_box.valueChanged.connect(self.plotfig)
        self.kernel_size_box.valueChanged.connect(self._on_kernel_size_changed)
        filter_form.addRow("Size:", self.kernel_size_box)
        
        # 多项式阶数
        self.poly_order_spin = QSpinBox()
        self.poly_order_spin.setRange(2, 5)
        self.poly_order_spin.setValue(3)
        self.poly_order_spin.setToolTip("Polynomial order for Savitzky-Golay filter")
        self.poly_order_spin.valueChanged.connect(self.plotfig)
        self.poly_order_spin.setMinimumWidth(70)  # 减小宽度从100到70
        filter_form.addRow("Poly Order:", self.poly_order_spin)
        
        # 启用滤波
        self.check_fitted_data_box = QCheckBox("Use Filtered Data")
        self.check_fitted_data_box.setChecked(True)
        self.check_fitted_data_box.stateChanged.connect(self.plotfig)
        filter_form.addRow("", self.check_fitted_data_box)
        
        data_layout.addLayout(filter_form)
        
        # 数据分段选择 - 改为更紧凑的布局
        segment_layout = QHBoxLayout()
        segment_layout.setSpacing(5)  # 减小水平间距
        
        self.sliced_data_box = QCheckBox("Segment")
        self.sliced_data_box.stateChanged.connect(self.plotfig)
        
        segment_label = QLabel("Segment:")
        
        self.chose_sliced_data_box = QComboBox()
        slice_items = [str(i) for i in range(1, self.num_of_sliced_data + 1)]
        self.chose_sliced_data_box.addItems(slice_items)
        self.chose_sliced_data_box.currentTextChanged.connect(self.plotfig)
        self.chose_sliced_data_box.setEnabled(False)
        self.chose_sliced_data_box.setMinimumWidth(60)  # 减小宽度从100到60
        
        segment_layout.addWidget(self.sliced_data_box)
        segment_layout.addWidget(segment_label)
        segment_layout.addWidget(self.chose_sliced_data_box)
        segment_layout.addStretch(1)
        
        data_layout.addLayout(segment_layout)
        
        combined_layout.addWidget(data_group)
        
        # 2. 区域设置与选择组
        region_group = QGroupBox("Region Selection")
        region_form = QFormLayout(region_group)
        region_form.setLabelAlignment(Qt.AlignLeft)
        region_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        region_form.setHorizontalSpacing(10)  # 减小水平间距
        region_form.setVerticalSpacing(5)  # 减小垂直间距
        
        self.min_duration_spin = QSpinBox()
        self.min_duration_spin.setRange(0, 60)
        self.min_duration_spin.setValue(1)
        self.min_duration_spin.valueChanged.connect(self.on_min_duration_changed)
        self.min_duration_spin.setMinimumWidth(70)  # 减小宽度从100到70
        region_form.addRow("Min Duration:", self.min_duration_spin)
        
        self.region_selector = QComboBox()
        self.region_selector.currentTextChanged.connect(self.select_force_region)
        self.region_selector.setMinimumWidth(250)  # 减小宽度从350到250
        region_form.addRow("Region:", self.region_selector)
        
        combined_layout.addWidget(region_group)
        
        # 3. 分析与数据操作组
        operation_group = QGroupBox("Analysis & Data Operations")
        operation_layout = QVBoxLayout(operation_group)
        operation_layout.setSpacing(5)  # 减小垂直间距
        operation_layout.setContentsMargins(5, 8, 5, 8)  # 减小边距
        
        # HMM分析按钮 - 全宽按钮
        self.kinetics_analysis_calculation_button = QPushButton("HMM Kinetics Analysis")
        self.kinetics_analysis_calculation_button.clicked.connect(self.kinetics_analysis_calculation)
        self.kinetics_analysis_calculation_button.setMinimumHeight(25)  # 减小按钮高度从30到25
        operation_layout.addWidget(self.kinetics_analysis_calculation_button)
        
        # 数据保存按钮组 - 使用水平布局
        buttons_layout = QHBoxLayout()
        buttons_layout.setSpacing(5)  # 减小间距
        
        # 保存数据按钮
        self.SaveSelectedDataButton = QPushButton("Save Data")  # 缩短文字
        self.SaveSelectedDataButton.clicked.connect(self.save_selected_data)
        
        # 保存时间数据按钮
        self.save_time_data_button = QPushButton("Save Time")  # 缩短文字
        self.save_time_data_button.clicked.connect(self.save_time_data)
        
        buttons_layout.addWidget(self.SaveSelectedDataButton)
        buttons_layout.addWidget(self.save_time_data_button)
        
        operation_layout.addLayout(buttons_layout)
        
        combined_layout.addWidget(operation_group)
        combined_layout.addStretch(1)  # 添加弹性空间
        
        # 设置初始多项式阶数控件状态
        self._update_poly_order_visibility()
        
        return combined_tab

    def _create_basic_tab(self):
        """创建基本控制标签页"""
        basic_tab = QWidget()
        basic_layout = QVBoxLayout(basic_tab)
        
        # 添加数据选择组
        data_group = QGroupBox("Data Selection")
        data_layout = QVBoxLayout(data_group)
        
        # Y轴选择
        y_layout = QHBoxLayout()
        y_label = QLabel("Y Axis (Bead):")
        self.y_axis_box = QComboBox()
        y_items = [str(self.beads_list[i]) for i in range(4, (len(self.beads_list) - 1))]
        self.y_axis_box.addItems(y_items)
        self.y_axis_box.currentTextChanged.connect(self.plotfig)
        y_layout.addWidget(y_label)
        y_layout.addWidget(self.y_axis_box)
        data_layout.addLayout(y_layout)
        
        # 滤波器选择
        filter_layout = QHBoxLayout()
        filter_label = QLabel("Filter Type:")
        filter_items = ["Moving Average", "Median Filter", "Savitzky-Golay"]
        self.filterselection = QComboBox()
        self.filterselection.addItems(filter_items)
        self.filterselection.currentTextChanged.connect(self.plotfig)
        filter_layout.addWidget(filter_label)
        filter_layout.addWidget(self.filterselection)
        data_layout.addLayout(filter_layout)
                
        # 数据分段
        slice_layout = QHBoxLayout()
        self.sliced_data_box = QCheckBox("Segment Data")
        self.sliced_data_box.stateChanged.connect(self.plotfig)
        slice_label = QLabel("Select Segment:")
        self.chose_sliced_data_box = QComboBox()
        slice_items = [str(i) for i in range(1, self.num_of_sliced_data + 1)]
        self.chose_sliced_data_box.addItems(slice_items)
        self.chose_sliced_data_box.currentTextChanged.connect(self.plotfig)
        self.chose_sliced_data_box.setEnabled(False)  # 默认禁用
        slice_layout.addWidget(self.sliced_data_box)
        slice_layout.addWidget(slice_label)
        slice_layout.addWidget(self.chose_sliced_data_box)
        data_layout.addLayout(slice_layout)
        
        basic_layout.addWidget(data_group)
        
        # 添加滤波设置组
        filter_group = QGroupBox("Filter Settings")
        filter_layout = QVBoxLayout(filter_group)

        # 在滤波设置组中添加多项式阶数控制
        poly_layout = QHBoxLayout()
        poly_label = QLabel("Polynomial Order:")
        self.poly_order_spin = QSpinBox()
        self.poly_order_spin.setRange(2, 5)
        self.poly_order_spin.setValue(3)
        self.poly_order_spin.setSingleStep(1)
        self.poly_order_spin.setToolTip("Polynomial order for Savitzky-Golay filter")
        self.poly_order_spin.valueChanged.connect(self.plotfig)
        self.poly_order_spin.setEnabled(self.filterselection.currentText() == "Savitzky-Golay")
        poly_layout.addWidget(poly_label)
        poly_layout.addWidget(self.poly_order_spin)
        filter_layout.addLayout(poly_layout)

        # 连接滤波类型变更信号
        self.filterselection.currentTextChanged.connect(self._update_poly_order_visibility)
        
        # 启用滤波
        self.check_fitted_data_box = QCheckBox("Use Filtered Data")
        self.check_fitted_data_box.setChecked(True)
        self.check_fitted_data_box.stateChanged.connect(self.plotfig)
        filter_layout.addWidget(self.check_fitted_data_box)
        
        # 内核大小
        kernel_layout = QHBoxLayout()
        kernel_label = QLabel("Filter Kernel Size:")
        self.kernel_size_box = QSpinBox()
        self.kernel_size_box.setRange(1, 1000)
        self.kernel_size_box.setValue(51)
        self.kernel_size_box.setSingleStep(2)
        self.kernel_size_box.valueChanged.connect(self.plotfig)
        self.kernel_size_box.valueChanged.connect(self._on_kernel_size_changed)
        kernel_layout.addWidget(kernel_label)
        kernel_layout.addWidget(self.kernel_size_box)
        filter_layout.addLayout(kernel_layout)
        
        basic_layout.addWidget(filter_group)
        
        # 添加最小区域持续时间设置
        region_group = QGroupBox("Region Settings")
        region_layout = QVBoxLayout(region_group)
        
        duration_layout = QHBoxLayout()
        duration_label = QLabel("Min Region Duration (sec):")
        self.min_duration_spin = QSpinBox()
        self.min_duration_spin.setRange(0, 60)
        self.min_duration_spin.setValue(1)
        self.min_duration_spin.setSingleStep(1)
        self.min_duration_spin.valueChanged.connect(self.on_min_duration_changed)
        duration_layout.addWidget(duration_label)
        duration_layout.addWidget(self.min_duration_spin)
        region_layout.addLayout(duration_layout)
        
        basic_layout.addWidget(region_group)
        
        # 添加基本数据操作按钮
        button_group = QGroupBox("Data Operations")
        button_layout = QVBoxLayout(button_group)
        
        # 保存选中数据按钮
        self.SaveSelectedDataButton = QPushButton("Save Selected Data")
        self.SaveSelectedDataButton.clicked.connect(self.save_selected_data)
        button_layout.addWidget(self.SaveSelectedDataButton)
        
        basic_layout.addWidget(button_group)
        basic_layout.addStretch(1)  # 添加弹性空间
        
        self.tab_widget.addTab(basic_tab, "Basic")
    
    def _create_analysis_tab(self):
        """创建分析控制标签页"""
        analysis_tab = QWidget()
        analysis_layout = QVBoxLayout(analysis_tab)
        
        # 添加区域选择组
        region_group = QGroupBox("Force Region Selection")
        region_layout = QVBoxLayout(region_group)
        
        region_select_layout = QHBoxLayout()
        region_label = QLabel("Select Region:")
        self.region_selector = QComboBox()
        self.region_selector.currentTextChanged.connect(self.select_force_region)
        region_select_layout.addWidget(region_label)
        region_select_layout.addWidget(self.region_selector)
        region_layout.addLayout(region_select_layout)
        
        analysis_layout.addWidget(region_group)
        
        # 添加动力学分析组
        kinetics_group = QGroupBox("Kinetics Analysis")
        kinetics_layout = QVBoxLayout(kinetics_group)
        
        # 动力学分析按钮
        self.kinetics_analysis_calculation_button = QPushButton("HMM Kinetics Analysis")
        self.kinetics_analysis_calculation_button.clicked.connect(self.kinetics_analysis_calculation)
        kinetics_layout.addWidget(self.kinetics_analysis_calculation_button)
        
        analysis_layout.addWidget(kinetics_group)
        
        # 添加数据保存组
        save_group = QGroupBox("Data Saving")
        save_layout = QVBoxLayout(save_group)
        
        # 保存时间数据按钮
        self.save_time_data_button = QPushButton("Save Time Data")
        self.save_time_data_button.clicked.connect(self.save_time_data)
        save_layout.addWidget(self.save_time_data_button)
        
        analysis_layout.addWidget(save_group)
        
        analysis_layout.addStretch(1)  # 添加弹性空间
        
        self.tab_widget.addTab(analysis_tab, "Analysis")
    
    def _create_data_info_panel(self):
        """创建数据信息显示面板 - 优化布局"""
        info_group = QGroupBox("Data Information")
        info_layout = QFormLayout(info_group)
        info_layout.setLabelAlignment(Qt.AlignLeft)
        info_layout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)
        info_layout.setHorizontalSpacing(10)  # 减小水平间距
        info_layout.setVerticalSpacing(5)  # 减小垂直间距
        info_layout.setContentsMargins(5, 8, 5, 8)  # 减小边距
        
        # T1 (时间点1)
        self.force_data_info = QLabel("0")
        self.force_data_info.setMinimumWidth(70)  # 减小宽度从100到70
        info_layout.addRow("T1:", self.force_data_info)
        
        # T2 (时间点2)
        self.extension_data_info = QLabel("0")
        self.extension_data_info.setMinimumWidth(70)  # 减小宽度从100到70
        info_layout.addRow("T2:", self.extension_data_info)
        
        # 力数据
        self.force_data_display = QLabel("0")
        self.force_data_display.setMinimumWidth(70)  # 减小宽度从100到70
        info_layout.addRow("Force (pN):", self.force_data_display)
        
        # 添加到控制面板
        self.control_layout.addWidget(info_group)

    def _on_kernel_size_changed(self, value):
        """当kernel size改变时确保poly_order合法"""
        if hasattr(self, 'poly_order_spin') and self.filterselection.currentText() == "Savitzky-Golay":
            if self.poly_order_spin.value() >= value:
                self.poly_order_spin.setValue(value - 1)
        self.plotfig()  # 更新图形

    def on_min_duration_changed(self, value):
        """当最小区域持续时间改变时重新绘图"""
        self.min_region_duration = value
        self.plotfig()
    
    def _setup_figure(self):
        """设置matplotlib图形和事件处理"""
        self.fig = plt.figure(figsize=(10, 8))
        self.canvas = FigureCanvasQTAgg(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.toolbar = NavigationToolbar2QT(self.canvas, self)
        
        # 添加到图形面板
        self.figure_layout.addWidget(self.canvas, 1)
        self.figure_layout.addWidget(self.toolbar, 0)
        
        # 设置事件处理
        self._setup_wheel_events()
        self._setup_point_selection_events()
    
    def _setup_wheel_events(self):
        """设置滚轮事件处理"""
        def wheelEvent(event):
            if self.y_axis_box.underMouse():
                if event.angleDelta().y() > 0:
                    self.y_axis_box.setCurrentIndex((self.y_axis_box.currentIndex() - 1) % self.y_axis_box.count())
                else:
                    self.y_axis_box.setCurrentIndex((self.y_axis_box.currentIndex() + 1) % self.y_axis_box.count())
            elif self.chose_sliced_data_box.underMouse():
                if event.angleDelta().y() > 0:
                    self.chose_sliced_data_box.setCurrentIndex(
                        (self.chose_sliced_data_box.currentIndex() - 1) % self.chose_sliced_data_box.count())
                else:
                    self.chose_sliced_data_box.setCurrentIndex(
                        (self.chose_sliced_data_box.currentIndex() + 1) % self.chose_sliced_data_box.count())
            elif self.kernel_size_box.underMouse():
                if event.angleDelta().y() > 0:
                    self.kernel_size_box.setValue(self.kernel_size_box.value() - 2)
                else:
                    self.kernel_size_box.setValue(self.kernel_size_box.value() + 2)
            else:
                super(KineticsAnalysis, self).wheelEvent(event)

        self.y_axis_box.wheelEvent = wheelEvent
        self.chose_sliced_data_box.wheelEvent = wheelEvent
    
    def _setup_point_selection_events(self):
        """设置点选择事件处理"""
        self.points_for_save = []

        def get_point(event):
            if event.inaxes is None or event.button != 2:
                return
                
            # 如果已有两个点，清空重新开始
            if len(self.points_for_save) == 2:
                self.points_for_save.clear()
                self.force_data_info.setText("0")
                self.extension_data_info.setText("0")
                self.force_data_display.setText("0")

            # 清除区域选择 - 添加这行
            if hasattr(self, 'region_highlight_rectangle') and self.region_highlight_rectangle:
                try:
                    self.region_highlight_rectangle.remove()
                    self.region_highlight_rectangle = None
                    self.canvas.draw_idle()
                except:
                    pass

            # 保存第一个点
            if len(self.points_for_save) == 0:
                self.x1 = event.xdata
                self.y1 = event.ydata
                self.points_for_save.append((self.x1, self.y1))
                self.force_data_info.setText(f"{self.x1:.4f}")
            # 保存第二个点并更新数据
            else:
                self.x2 = event.xdata
                self.y2 = event.ydata
                self.points_for_save.append((self.x2, self.y2))
                self.extension_data_info.setText(f"{self.x2:.4f}")
                
                # 更新时间点和力值显示
                self._update_time_points()
                self._update_force_display()
                
                # 在图表上标记选择的区域
                if hasattr(self, 'fig') and hasattr(self.fig, 'axes') and len(self.fig.axes) > 0:
                    ax = self.fig.axes[0]
                    ymin, ymax = ax.get_ylim()
                    # 添加垂直线标记选择的时间点
                    ax.axvline(x=self.x1, color='g', linestyle='--', alpha=0.5)
                    ax.axvline(x=self.x2, color='g', linestyle='--', alpha=0.5)
                    # 更新图表
                    self.canvas.draw_idle()

        # 连接事件处理函数
        self.canvas.mpl_connect('button_press_event', get_point)
        
    def _update_dynamic_force_labels(self):
        """更新图表上的动态力值标签位置以适应缩放"""
        if not hasattr(self, 'fig') or not hasattr(self.fig, 'axes') or len(self.fig.axes) == 0:
            return
            
        ax = self.fig.axes[0]
        
        # 如果有力值区域标签，更新它们的位置
        if hasattr(self, 'force_region_labels') and self.force_region_labels:
            x_min, x_max = ax.get_xlim()
            y_min, y_max = ax.get_ylim()
            
            for region_idx, label in enumerate(self.force_region_labels):
                if region_idx < len(self.force_regions):
                    region = self.force_regions[region_idx]
                    # 计算新的标签位置 - 在视图中央
                    label_x = (region['start_time'] + region['end_time']) / 2
                    
                    # 如果标签在当前视图外，则隐藏
                    if label_x < x_min or label_x > x_max:
                        label.set_visible(False)
                    else:
                        label.set_visible(True)
                        # 重新定位在视图顶部
                        label_y = y_max - (y_max - y_min) * 0.05
                        label.set_position((label_x, label_y))
                        label.set_text(f"{region['mean_force']:.2f} pN")

    def _get_selected_data_slice(self, start_time, end_time):
        """统一的数据选择方法，根据时间范围返回对应的数据切片
        
        Args:
            start_time (float): 起始时间点
            end_time (float): 结束时间点
            
        Returns:
            tuple: 包含时间数据，位移数据，力数据的切片
        """
        # 确保时间点的顺序正确
        if start_time > end_time:
            start_time, end_time = end_time, start_time
            
        # 找到最接近的数据点索引
        start_point = min(self.xx, key=lambda x: abs(x - start_time))
        start_idx = self.xx.tolist().index(start_point)
        end_point = min(self.xx, key=lambda x: abs(x - end_time))
        end_idx = self.xx.tolist().index(end_point)
        
        # 确保起始索引在结束索引之前
        if start_idx > end_idx:
            start_idx, end_idx = end_idx, start_idx
        
        # 提取数据切片
        time_range = slice(start_idx, end_idx)
        time_data = self.tdms_data_store['time s'].values[time_range]
        bead_data = self.tdms_data_store[str(self.y_axis_box.currentText())].values[time_range]
        force_data = self.tdms_data_store['force pN'].values[time_range]
        
        # 更新UI显示
        self.force_data_info.setText(f"{start_point:.4f}")
        self.extension_data_info.setText(f"{end_point:.4f}")
        self.force_calculate = np.mean(force_data)
        self.force_data_display.setText(f"{self.force_calculate:.2f} pN")
        
        # 保存计算结果供其他方法使用
        self.time_point_start = start_point
        self.time_point_start_data = start_idx
        self.time_point_end = end_point
        self.time_point_end_data = end_idx
        self.time_data = time_data
        self.bead_data = bead_data
        self.force_data = force_data
        
        return time_data, bead_data, force_data
    

    def plotfig(self):
        """绘制并更新图形显示"""
        # 准备数据
        self.currentfilterselection = self.filterselection.currentText()
        self.chosen_bead = self.y_axis_box.currentText()
        
        # X轴固定为时间
        self.x_axis = self.tdms_data_store['time s']
        self.y_axis = self.tdms_data_store[str(self.chosen_bead)]
        self.y_axis_2 = self.tdms_data_store['force pN']  # 保留力数据用于分析，但不绘制曲线
        self.xx = self.x_axis.values
        self.yy = self.y_axis.values
        self.yy_2 = self.y_axis_2.values

        # 清除当前图形 - 修复坐标轴重叠问题
        self.fig.clear()
        
        # 初始化动态标签列表
        self.force_region_labels = []
        
        # 根据UI状态更新控件可用性
        self.chose_sliced_data_box.setEnabled(self.sliced_data_box.isChecked())
        self.kernel_size_box.setEnabled(self.check_fitted_data_box.isChecked())
        
        # 获取数据段(原始或切片)
        axis_x, axis_y, axis_y_2 = self._get_plot_data()
        
        # 创建图形和子图
        ax = self.fig.add_subplot(111)
        
        # 绘制原始数据
        ax.plot(axis_x, axis_y, color='darkgrey', label='Raw Data')
        
        # 如果启用了拟合,应用滤波器并绘制拟合数据
        if self.check_fitted_data_box.isChecked():
            filter_type = self.filterselection.currentText()
            kernel_size = self.kernel_size_box.value()
            
            if filter_type == "Savitzky-Golay":
                poly_order = self.poly_order_spin.value()
                y_fitted = self._apply_filter_to_data(axis_y, filter_type, kernel_size, poly_order)
            else:
                y_fitted = self._apply_filter_to_data(axis_y, filter_type, kernel_size)
                
            ax.plot(axis_x, y_fitted, color='red', 
                    label=f'Filtered Data ({filter_type}, Kernel Size={kernel_size})')
        
        # 添加力值区域划分并更新区域选择下拉框
        self._highlight_force_regions(ax, axis_x, axis_y_2)
        self._update_region_selector()
        
        # 添加图例
        ax.legend(loc='upper left')
        
        # 设置标签
        ax.set_xlabel('Time (s)')
        ax.set_ylabel('Extension (nm)')
        
        # 更新画布
        self.canvas.draw_idle()
    
    def _update_poly_order_visibility(self):
        """更新多项式阶数控件的可见性"""
        if hasattr(self, 'poly_order_spin'):
            is_savgol = self.filterselection.currentText() == "Savitzky-Golay"
            self.poly_order_spin.setEnabled(is_savgol)
            
            # 确保多项式阶数小于窗口大小
            if is_savgol and self.poly_order_spin.value() >= self.kernel_size_box.value():
                self.poly_order_spin.setValue(self.kernel_size_box.value() - 1)

    def save_selected_data(self):
        """保存选中的数据到Excel和图片文件"""
        # 检查是否有选择的时间点或区域
        has_time_points = hasattr(self, 'points_for_save') and len(self.points_for_save) == 2
        has_selected_region = hasattr(self, 'region_highlight_rectangle') and self.region_highlight_rectangle
        
        if not (has_time_points or has_selected_region):
            QMessageBox.warning(self, "选择错误", "请先选择两个时间点或一个力值区域")
            return

        # 获取选中的数据范围
        # 修改条件判断逻辑：如果有手动选择的两个点，优先使用手动选择
        if has_time_points:
            # 使用时间点选择 - 优先使用手动选择的点
            self.x1 = float(self.force_data_info.text())
            self.x2 = float(self.extension_data_info.text())
            time_data, bead_data, force_data = self._get_selected_data_slice(self.x1, self.x2)
        elif has_selected_region and self.region_selector.currentIndex() >= 0:
            # 只有在没有手动选择点的情况下，才使用区域选择
            selected_idx = self.region_selector.currentIndex()
            if selected_idx < len(self.force_regions):
                region = self.force_regions[selected_idx]
                time_data, bead_data, force_data = self._get_selected_data_slice(
                    region['start_time'], region['end_time'])
        else:
            QMessageBox.warning(self, "选择错误", "无法获取有效的数据区域")
            return

        # 应用滤波 - 确保滤波参数有效
        filter_type = self.filterselection.currentText()
        kernel_size = self.kernel_size_box.value()

        if filter_type == "Savitzky-Golay":
            poly_order = self.poly_order_spin.value()
            # 确保poly_order < kernel_size
            if poly_order >= kernel_size:
                poly_order = kernel_size - 1
            fit_bead_data = self._apply_filter_to_data(bead_data, filter_type, kernel_size, poly_order)
        else:
            fit_bead_data = self._apply_filter_to_data(bead_data, filter_type, kernel_size)

        # 计算平均力值
        self.force_calculate = np.mean(force_data)

        # Prepare file paths
        self.current_bead_name = self.y_axis_box.currentText()
        base_filename = f"{self.base_name}_{self.current_bead_name}_{self.force_calculate:.2f}pN"
        xlsx_file_path = os.path.join(self.Data_Saved_Path, f"{base_filename}_SelectedData.xlsx")
        image_file_path = os.path.join(self.Data_Saved_Path, f"{base_filename}_SelectedData.png")
        worksheet_name = f"{self.current_bead_name}_{self.force_calculate:.2f}pN"

        # Create and save the plot
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.plot(time_data, bead_data, 'gray', alpha=0.5, label='Original Data')
        ax.plot(time_data, fit_bead_data, 'r-', linewidth=2, 
                label=f'Filtered Data ({self.filterselection.currentText()}, Kernel Size={self.kernel_size_box.value()})')
        
        # 添加力值区域标注
        self._highlight_force_regions_in_save_plot(ax, time_data, force_data)
        
        # 添加手动选择的点的垂直线标记 - 如果是从手动选择保存的
        if has_time_points:
            ax.axvline(x=self.x1, color='g', linestyle='--', alpha=0.5)
            ax.axvline(x=self.x2, color='g', linestyle='--', alpha=0.5)
        
        ax.set_xlabel('Time (s)')
        ax.set_ylabel('Extension (nm)')
        ax.set_title(f'Selected Data - {self.current_bead_name}, Force: {self.force_calculate:.2f} pN')
        ax.legend(loc='upper left')
        fig.savefig(image_file_path, format='png', dpi=300, bbox_inches='tight')
        plt.close(fig)  # Close the figure to free memory

        # Create or load workbook
        try:
            workbook = openpyxl.load_workbook(xlsx_file_path) if os.path.exists(xlsx_file_path) else openpyxl.Workbook()
            
            if worksheet_name not in workbook.sheetnames:
                # Use active sheet for new workbook or create new sheet
                if len(workbook.sheetnames) == 1 and workbook.active.title == 'Sheet':
                    worksheet = workbook.active
                    worksheet.title = worksheet_name
                else:
                    worksheet = workbook.create_sheet(worksheet_name)
                
                # Add headers and data
                worksheet.append(['Time (s)', 'Extension (nm)', 'Filtered Extension (nm)', 'Force (pN)'])
                for i in range(len(time_data)):
                    worksheet.cell(i + 2, 1, time_data[i])
                    worksheet.cell(i + 2, 2, bead_data[i])
                    worksheet.cell(i + 2, 3, fit_bead_data[i])
                    worksheet.cell(i + 2, 4, force_data[i])
                
                # Save workbook
                workbook.save(xlsx_file_path)
                
            # Show confirmation message
            QMessageBox.information(self, "Save Complete", 
                                f"Data saved to:\n- Excel: {xlsx_file_path}\n- Image: {image_file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save data: {str(e)}")

    def _highlight_force_regions(self, ax, time_data, force_data, threshold=0.5):
        """根据力值变化添加背景色块以区分不同区域
        
        Args:
            ax: matplotlib轴对象
            time_data: 时间数据
            force_data: 力数据
            threshold: 力值变化阈值，用于检测力值显著变化
        """
        if len(time_data) < 2 or len(force_data) < 2:
            return
        
        # 清除之前的标签
        self.force_region_labels = []
        
        # 重置区域信息
        self.force_regions = []
        
        # 计算力值的梯度
        force_gradient = np.abs(np.gradient(force_data))
        
        # 获取力值变化显著的点
        # 使用 threshold 乘以力梯度的标准差，使检测更加稳健
        change_points = np.where(force_gradient > threshold * np.std(force_gradient))[0]
        
        # 合并临近的变化点，避免过度分段
        if len(change_points) > 0:
            merged_points = [change_points[0]]
            for point in change_points[1:]:
                if point - merged_points[-1] > 10:  # 至少间隔10个点
                    merged_points.append(point)
            change_points = np.array(merged_points)
        
        # 添加开始和结束点
        all_points = np.concatenate(([0], change_points, [len(time_data) - 1]))
        
        # 使用交替的背景色
        colors = ['lightyellow', 'lightblue', 'lightgreen', 'lightpink', 'lightgrey']
        
        # 为每个区域添加背景色和力值标注
        for i in range(len(all_points) - 1):
            start_idx = all_points[i]
            end_idx = all_points[i+1]
            
            # 确保区段长度足够
            if end_idx - start_idx < 10:
                continue
            
            # 计算区域持续时间
            region_duration = time_data[end_idx] - time_data[start_idx]
            
            # 如果区域持续时间小于最小值，跳过此区域
            if region_duration < self.min_region_duration:
                continue
            
            # 计算区域内的平均力值
            mean_force = np.mean(force_data[start_idx:end_idx])
            
            # 添加背景色块
            color = colors[i % len(colors)]
            ax.axvspan(time_data[start_idx], time_data[end_idx], 
                      alpha=0.3, color=color)
            
            # 添加力值标签 (放在区域顶部)
            if end_idx - start_idx > 20:  # 只有区域足够大才添加标签
                label_x = (time_data[start_idx] + time_data[end_idx]) / 2
                y_lim = ax.get_ylim()
                label_y = y_lim[1] - (y_lim[1] - y_lim[0]) * 0.05  # 放在顶部附近
                label = ax.text(label_x, label_y, f"{mean_force:.2f} pN", 
                       ha='center', va='top', fontsize=9,
                       bbox=dict(facecolor='white', alpha=0.7, edgecolor='none'))
                self.force_region_labels.append(label)
            
            # 记录区域信息，便于后续选择
            self.force_regions.append({
                'start_time': time_data[start_idx],
                'end_time': time_data[end_idx],
                'mean_force': mean_force,
                'start_idx': start_idx,
                'end_idx': end_idx,
                'duration': region_duration
            })

    def _update_region_selector(self):
        """更新区域选择下拉框"""
        # 清空当前项
        self.region_selector.clear()
        
        # 如果没有区域，显示提示并返回
        if not self.force_regions:
            self.region_selector.addItem("No regions detected")
            return
            
        # 添加每个区域到下拉框
        for i, region in enumerate(self.force_regions):
            # 显示区域编号、时间范围和平均力值
            self.region_selector.addItem(
                f"Region {i+1}: {region['start_time']:.2f}-{region['end_time']:.2f}s ({region['mean_force']:.2f} pN)"
            )

    def _get_plot_data(self):
        """根据UI设置获取绘图数据"""
        if self.sliced_data_box.isChecked():
            # 获取所选切片
            selected_sliced_data_num = int(self.chose_sliced_data_box.currentText()) - 1
            section_of_selected_data = self.final_slice_magnet_move_state[selected_sliced_data_num]
            
            # 计算切片起止点
            start_point = section_of_selected_data[0]
            if start_point != 0:
                start_point = start_point - 1
            end_point = section_of_selected_data[1] - 1
            
            # 切片数据
            axis_x = self.xx[start_point:end_point]
            axis_y = self.yy[start_point:end_point]
            axis_y_2 = self.yy_2[start_point:end_point]
        else:
            # 使用全部数据
            axis_x = self.xx
            axis_y = self.yy
            axis_y_2 = self.yy_2
            
        return axis_x, axis_y, axis_y_2

    def _apply_filter(self, x_data, y_data, kernel_size):
        """应用滤波器到数据 - 过渡方法，使用统一滤波接口"""
        if self.currentfilterselection == 'Median Filter':
            y_fitted = self._apply_filter_to_data(y_data, 'Median Filter', kernel_size)
            return x_data, y_fitted
        elif self.currentfilterselection == 'Moving Average':
            y_fitted = self._apply_filter_to_data(y_data, 'Moving Average', kernel_size)
            return x_data, y_fitted
        elif self.currentfilterselection == 'Savitzky-Golay':
            poly_order = self.poly_order_spin.value()
            y_fitted = self._apply_filter_to_data(y_data, 'Savitzky-Golay', kernel_size, poly_order)
            return x_data, y_fitted
        
        return x_data, y_data
    
    def _update_time_points(self):
        """更新选中的时间点数据"""
        try:
            # 直接调用统一的数据选择方法
            self.x1 = float(self.force_data_info.text())
            self.x2 = float(self.extension_data_info.text())
            self._get_selected_data_slice(self.x1, self.x2)
        except Exception as e:
            print(f"更新时间点时出错: {e}")

    def _update_force_display(self):
        """计算并显示选定区域的平均力值"""
        try:
            if hasattr(self, 'force_data') and len(self.force_data) > 0:
                # 计算平均力值
                self.force_calculate = np.mean(self.force_data)
                # 更新UI显示
                self.force_data_display.setText(f"{self.force_calculate:.2f} pN")
        except Exception as e:
            print(f"更新力值显示时出错: {e}")

    def _prepare_time_data(self):
        """Helper method to prepare data for saving time intervals"""
        # 使用统一的数据选择方法
        self.x1 = float(self.force_data_info.text())
        self.x2 = float(self.extension_data_info.text())
        self._get_selected_data_slice(self.x1, self.x2)
        
        # 计算时间间隔
        return self.x2 - self.x1

    def select_force_region(self):
        """选择特定力值区域进行分析"""
        if not hasattr(self, 'force_regions') or not self.force_regions:
            return
            
        selected_idx = self.region_selector.currentIndex()
        if (selected_idx < 0 or selected_idx >= len(self.force_regions)):
            return
            
        region = self.force_regions[selected_idx]
        
        # 使用统一的数据选择方法
        start_time = region['start_time']
        end_time = region['end_time']
        self._get_selected_data_slice(start_time, end_time)
        
        # 清除手动选择的点 - 添加这行
        if hasattr(self, 'points_for_save'):
            self.points_for_save = []
        
        # 更新图表选择
        ax = self.fig.axes[0]
        if hasattr(self, 'region_highlight_rectangle') and self.region_highlight_rectangle:
            try:
                self.region_highlight_rectangle.remove()
            except:
                pass
        
        # 添加高亮矩形
        y_min, y_max = ax.get_ylim()
        self.region_highlight_rectangle = ax.axvspan(region['start_time'], region['end_time'], 
                                                    alpha=0.4, color='yellow', zorder=-10)
        
        # 更新图表
        self.canvas.draw_idle()
        
    def kinetics_analysis_calculation(self):
        # 使用统一的数据选择方法
        self.x1 = float(self.force_data_info.text())
        self.x2 = float(self.extension_data_info.text())
        time_data, bead_data, force_data = self._get_selected_data_slice(self.x1, self.x2)

        # 计算基线和平均力值
        self.base_line = np.mean(bead_data)
        self.force_calculate = np.mean(force_data)

        # Create enhanced HMM analysis dialog
        class StateSelectionDialog(QDialog):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("Enhanced HMM State Selection")
                self.resize(400, 300)
                
                layout = QVBoxLayout(self)
                
                # Create group box for radio buttons
                groupBox = QGroupBox("Select Number of States:")
                radioLayout = QVBoxLayout()
                
                # Create radio buttons
                self.radio2 = QRadioButton("2 States")
                self.radio2.setChecked(True)
                self.radio3 = QRadioButton("3 States")
                
                radioLayout.addWidget(self.radio2)
                radioLayout.addWidget(self.radio3)
                groupBox.setLayout(radioLayout)
                layout.addWidget(groupBox)
                
                # 添加数据预处理选项组
                preprocessing_group = QGroupBox("Data Preprocessing:")
                preprocessing_layout = QVBoxLayout()
                
                # 添加是否使用当前滤波的选项
                self.use_current_filter = QCheckBox("Use Current Filter Settings")
                self.use_current_filter.setChecked(True)
                self.use_current_filter.setToolTip("Apply the filter settings from visualization to HMM analysis")
                preprocessing_layout.addWidget(self.use_current_filter)
                
                # 添加直方图初始化选项
                self.use_histogram_initialization = QCheckBox("Use Histogram-based Initialization")
                self.use_histogram_initialization.setChecked(True)
                self.use_histogram_initialization.setToolTip("Automatically detect initial state means from data histogram")
                preprocessing_layout.addWidget(self.use_histogram_initialization)
                
                preprocessing_group.setLayout(preprocessing_layout)
                layout.addWidget(preprocessing_group)
                
                # Add buttons
                buttonBox = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                buttonBox.accepted.connect(self.accept)
                buttonBox.rejected.connect(self.reject)
                layout.addWidget(buttonBox)
            
            def getSettings(self):
                n_states = 2 if self.radio2.isChecked() else 3
                return {
                    'n_states': n_states,
                    'use_histogram_initialization': self.use_histogram_initialization.isChecked(),
                    'use_current_filter': self.use_current_filter.isChecked()
                }

        # Show dialog to select number of states and settings
        dialog = StateSelectionDialog(self)
        if dialog.exec():
            settings = dialog.getSettings()
            n_states = settings['n_states']
            use_histogram_init = settings['use_histogram_initialization']
            use_current_filter = settings['use_current_filter']
            
            # 根据选择应用滤波
            if use_current_filter:
                # 使用当前滤波设置
                filter_type = self.filterselection.currentText()
                kernel_size = self.kernel_size_box.value()
                
                if filter_type == "Savitzky-Golay":
                    poly_order = self.poly_order_spin.value()
                    filtered_data = self._apply_filter_to_data(bead_data, filter_type, kernel_size, poly_order)
                else:
                    filtered_data = self._apply_filter_to_data(bead_data, filter_type, kernel_size)
            else:
                # 默认使用中值滤波
                filtered_data = scipy.signal.medfilt(bead_data, kernel_size=3)
            
            # Prepare data for HMM (reshape to 2D array)
            extension_data_for_hmm = np.array(filtered_data).reshape(-1, 1)
            
            # Create enhanced HMM model with optimized initialization
            initial_means = None
            if use_histogram_init:
                initial_means = self._estimate_initial_states_from_histogram(filtered_data, n_states)
            
            # Create HMM results window with interactive controls
            self.hmm_window = EnhancedHMMResultWindow(
                self, time_data, bead_data, force_data, extension_data_for_hmm, 
                n_states, self.force_calculate, self.y_axis_box.currentText(),
                initial_means, use_current_filter  # 传递滤波设置给结果窗口
            )
            self.hmm_window.show()

    def _estimate_initial_states_from_histogram(self, data, n_states):
        """通过数据直方图分析估计初始状态均值"""
        try:
            # 计算数据直方图
            hist_counts, bin_edges = np.histogram(data, bins=50, density=True)
            bin_centers = (bin_edges[:-1] + bin_edges[1:]) / 2
            
            # 使用scipy找峰值
            peaks, properties = find_peaks(hist_counts, height=np.max(hist_counts) * 0.1, 
                                         distance=len(hist_counts) // (n_states * 2))
            
            if len(peaks) >= n_states:
                # 选择最高的n_states个峰值
                peak_heights = hist_counts[peaks]
                top_peaks_idx = np.argsort(peak_heights)[-n_states:]
                selected_peaks = peaks[top_peaks_idx]
                initial_means = np.sort(bin_centers[selected_peaks])
            else:
                # 如果峰值不够，使用分位数方法
                quantiles = np.linspace(0.1, 0.9, n_states)
                initial_means = np.quantile(data, quantiles)
            
            return initial_means
            
        except Exception as e:
            print(f"Histogram initialization failed: {e}")
            # 回退到分位数方法
            quantiles = np.linspace(0.1, 0.9, n_states)
            return np.quantile(data, quantiles)
    
    def _highlight_force_regions_in_save_plot(self, ax, time_data, force_data, threshold=0.3):
        """为保存的图片标注不同的力值区域
        
        Args:
            ax: matplotlib轴对象
            time_data: 时间数据
            force_data: 力数据
            threshold: 力值变化阈值，用于检测力值显著变化
        """
        if len(time_data) < 2 or len(force_data) < 2:
            return
        
        # 计算力值的梯度
        force_gradient = np.abs(np.gradient(force_data))
        
        # 获取力值变化显著的点
        # 使用自适应阈值，确保小区域的变化也能检测到
        change_points = np.where(force_gradient > threshold * np.std(force_gradient))[0]
        
        # 合并临近的变化点，避免过度分段
        if len(change_points) > 0:
            merged_points = [change_points[0]]
            for point in change_points[1:]:
                if point - merged_points[-1] > 5:  # 更小的间隔以捕获更多的变化
                    merged_points.append(point)
            change_points = np.array(merged_points)
        
        # 添加开始和结束点
        all_points = np.concatenate(([0], change_points, [len(time_data) - 1]))
        
        # 使用交替的背景色
        colors = ['lightyellow', 'lightblue', 'lightgreen', 'lightpink', 'lightgrey']
        
        # 为每个区域添加背景色和力值标注
        for i in range(len(all_points) - 1):
            start_idx = all_points[i]
            end_idx = all_points[i+1]
            
            # 确保区段长度足够
            if end_idx - start_idx < 5:
                continue
            
            # 计算区域内的平均力值
            mean_force = np.mean(force_data[start_idx:end_idx])
            
            # 添加背景色块
            color = colors[i % len(colors)]
            ax.axvspan(time_data[start_idx], time_data[end_idx], 
                      alpha=0.3, color=color)
            
            # 添加更加明显的力值标签 (放在区域中央)
            if end_idx - start_idx > 10:  # 只有区域足够大才添加标签
                label_x = (time_data[start_idx] + time_data[end_idx]) / 2
                y_lim = ax.get_ylim()
                label_y = y_lim[1] - (y_lim[1] - y_lim[0]) * 0.1  # 放在顶部附近
                
                # 添加带边框的白色背景标签
                ax.text(label_x, label_y, f"{mean_force:.2f} pN", 
                       ha='center', va='center', fontsize=10, fontweight='bold',
                       bbox=dict(facecolor='white', alpha=0.8, edgecolor='black', boxstyle='round,pad=0.5'))

    def _save_time_data(self, is_open_time):
        """Save time data to Excel file
        
        Args:
            is_open_time (bool): True for open time, False for close time
        """
        # 使用统一的数据选择方法
        time_interval = self._prepare_time_data()
        
        # Configure columns based on data type
        if is_open_time:
            time_col, force_col, count_col = 1, 2, 3
            other_count_col = 6
            time_type = "open"
            column_header = 'Time Interval (Open)'
            force_header = 'Force (Open, pN)'
            count_header = 'Count (Open)'
        else:
            time_col, force_col, count_col = 4, 5, 6
            other_count_col = 3
            time_type = "close"
            column_header = 'Time Interval (Closed)'
            force_header = 'Force (Closed, pN)'
            count_header = 'Count (Closed)'
        
        # Prepare file and worksheet
        self.current_bead_name = self.y_axis_box.currentText()
        xlsx_file_path = os.path.join(self.Data_Saved_Path, f"{self.base_name}_Dy.xlsx")
        worksheet_name = f"{self.current_bead_name}_{self.chose_sliced_data_box.currentText()}"
        
        # Create or open workbook
        if os.path.exists(xlsx_file_path):
            workbook = openpyxl.load_workbook(xlsx_file_path)
        else:
            workbook = openpyxl.Workbook()
            if 'Sheet' in workbook.sheetnames:
                # 移除默认的Sheet
                workbook.remove(workbook['Sheet'])
        
        # Create or get worksheet
        if worksheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(worksheet_name)
            
            # 添加标题行
            headers = ['Time Interval (Open)', 'Force (Open, pN)', 'Count (Open)', 
                       'Time Interval (Closed)', 'Force (Closed, pN)', 'Count (Closed)']
            for col, header in enumerate(headers, 1):
                worksheet.cell(1, col, header)
            
            # 初始化计数器
            worksheet.cell(2, 3, 0)  # 打开计数
            worksheet.cell(2, 6, 0)  # 关闭计数
        else:
            worksheet = workbook[worksheet_name]
        
        # 确保标题行存在
        if worksheet.cell(1, time_col).value != column_header:
            headers = ['Time Interval (Open)', 'Force (Open, pN)', 'Count (Open)', 
                       'Time Interval (Closed)', 'Force (Closed, pN)', 'Count (Closed)']
            for col, header in enumerate(headers, 1):
                worksheet.cell(1, col, header)
        
        # Update counter
        n_counts = worksheet.cell(2, count_col).value or 0
        n_counts += 1
        
        # Write data
        worksheet.cell(n_counts + 2, time_col, time_interval)  # 从第3行开始写入，前两行是标题和计数
        worksheet.cell(n_counts + 2, force_col, self.force_calculate)
        worksheet.cell(2, count_col, n_counts)  # 更新计数
        
        # Save workbook
        workbook.save(xlsx_file_path)
        
        # 显示保存成功消息
        QMessageBox.information(self, "Save Complete", 
                               f"{time_type.capitalize()} time data saved to: {xlsx_file_path}")

    def save_time_data(self):
        """合并后的时间数据保存功能"""
        # 检查是否有选择的时间点
        if len(self.points_for_save) != 2:
            QMessageBox.warning(self, "选择错误", "请先选择两个时间点以定义时间间隔")
            return
        
        # 创建选择对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("Select Time Type")
        dialog.resize(300, 150)
        
        layout = QVBoxLayout(dialog)
        
        # 创建单选按钮组
        group_box = QGroupBox("Select Time Type:")
        radio_layout = QVBoxLayout()
        
        open_radio = QRadioButton("Open Time")
        open_radio.setChecked(True)
        close_radio = QRadioButton("Close Time")
        
        radio_layout.addWidget(open_radio)
        radio_layout.addWidget(close_radio)
        group_box.setLayout(radio_layout)
        layout.addWidget(group_box)
        
        # 添加按钮
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)
        
        # 显示对话框并处理结果
        if dialog.exec():
            is_open_time = open_radio.isChecked()
            self._save_time_data(is_open_time)

class EnhancedHMMResultWindow(QMainWindow):
    def __init__(self, parent, time_data, extension_data, force_data, extension_data_for_hmm, 
                n_states, force_calculate, bead_name, initial_means=None, 
                used_current_filter=False):
        super().__init__()
        self.parent = parent
        self.time_data = time_data
        self.extension_data = extension_data
        self.force_data = force_data
        self.extension_data_for_hmm = extension_data_for_hmm
        self.n_states = n_states
        self.force = force_calculate
        self.bead_name = bead_name
        self.initial_means = initial_means
        self.used_current_filter = used_current_filter  # 记录是否使用了当前滤波
        
        # 创建状态映射
        self.state_map = {}
        if n_states == 2:
            self.state_names = ["Closed", "Open"]
        else:
            self.state_names = ["Closed", "Intermediate", "Open"]
        
        # 初始化HMM模型和结果
        self.model = None
        self.original_transmat = None  # 保存原始转移矩阵
        self.hidden_states = None
        self.square_wave = None
        self.dwell_times = []
        self.step_sizes = []
        self.step_times = []
        self.transition_stats = {}
        self.effective_transmat = None  # 添加有效转移矩阵属性
        
        # 添加参数变化标志，避免重复更新
        self.updating = False
        
        self.setWindowTitle(f"Interactive HMM Analysis - Force: {self.force:.2f} pN")
        self.resize(1400, 800)
        
        self._setup_ui()
        self._train_initial_model()
        self._update_analysis()
    
    def _setup_ui(self):
        """设置增强的UI界面，控制面板在右侧"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 创建主布局（水平布局）
        main_layout = QHBoxLayout(central_widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        
        # 左侧图形面板
        self.figure_panel = QWidget()
        figure_layout = QVBoxLayout(self.figure_panel)
        figure_layout.setContentsMargins(0, 0, 0, 0)
        
        # 创建图形
        self.fig = plt.figure(figsize=(10, 8))
        self.canvas = FigureCanvasQTAgg(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.toolbar = NavigationToolbar2QT(self.canvas, self)
        
        figure_layout.addWidget(self.canvas, 1)
        figure_layout.addWidget(self.toolbar, 0)
        
        # 右侧控制面板
        self.control_panel = QWidget()
        self.control_panel.setFixedWidth(350)
        control_layout = QVBoxLayout(self.control_panel)
        
        # 创建标签页控件
        self.tab_widget = QTabWidget()
        
        # 创建各个控制标签页
        self._create_parameter_tab()
        self._create_analysis_tab()
        self._create_export_tab()
        
        control_layout.addWidget(self.tab_widget)
        
        # 添加面板到主布局
        main_layout.addWidget(self.figure_panel, 3)  # 图形面板占更多空间
        main_layout.addWidget(self.control_panel, 1)
    
    def _create_parameter_tab(self):
        """创建参数调整标签页"""
        param_tab = QWidget()
        param_layout = QVBoxLayout(param_tab)
        
        # HMM参数组
        hmm_group = QGroupBox("HMM Parameters")
        hmm_layout = QFormLayout(hmm_group)
        
        # 转移惩罚
        self.transition_penalty_spin = QSpinBox()
        self.transition_penalty_spin.setRange(1, 100)
        self.transition_penalty_spin.setValue(20)
        self.transition_penalty_spin.setToolTip("Higher values reduce state transitions")
        self.transition_penalty_spin.valueChanged.connect(self._on_parameter_changed)
        hmm_layout.addRow("Transition Penalty:", self.transition_penalty_spin)
        
        # 迭代次数
        self.n_iter_spin = QSpinBox()
        self.n_iter_spin.setRange(50, 500)
        self.n_iter_spin.setValue(200)
        self.n_iter_spin.setSingleStep(50)
        self.n_iter_spin.setToolTip("Number of EM algorithm iterations")
        self.n_iter_spin.valueChanged.connect(self._on_parameter_changed)
        hmm_layout.addRow("Iterations:", self.n_iter_spin)
        
        param_layout.addWidget(hmm_group)
        
        # 后处理参数组
        postprocess_group = QGroupBox("Post-processing Parameters")
        postprocess_layout = QFormLayout(postprocess_group)
        
        # 最小驻留时间
        self.min_dwell_time_spin = QSpinBox()
        self.min_dwell_time_spin.setRange(1, 200)
        self.min_dwell_time_spin.setValue(35)
        self.min_dwell_time_spin.setSuffix(" points")
        self.min_dwell_time_spin.setToolTip("Minimum time a state must be maintained")
        self.min_dwell_time_spin.valueChanged.connect(self._on_parameter_changed)
        postprocess_layout.addRow("Min Dwell Time:", self.min_dwell_time_spin)
        
        # 平滑窗口大小
        self.smoothing_window_spin = QSpinBox()
        self.smoothing_window_spin.setRange(3, 31)
        self.smoothing_window_spin.setValue(9)
        self.smoothing_window_spin.setSingleStep(2)
        self.smoothing_window_spin.setToolTip("Window size for state sequence smoothing")
        self.smoothing_window_spin.valueChanged.connect(self._on_parameter_changed)
        postprocess_layout.addRow("Smoothing Window:", self.smoothing_window_spin)
        
        # 启用平滑
        self.use_smoothing_check = QCheckBox("Enable Smoothing")
        self.use_smoothing_check.setChecked(True)
        self.use_smoothing_check.setToolTip("Apply smoothing to reduce noise-induced transitions")
        self.use_smoothing_check.stateChanged.connect(self._on_parameter_changed)
        postprocess_layout.addRow("", self.use_smoothing_check)
        
        param_layout.addWidget(postprocess_group)
        
        # 重新训练按钮
        retrain_button = QPushButton("Retrain Model")
        retrain_button.setToolTip("Retrain HMM model with current parameters")
        retrain_button.clicked.connect(self._retrain_model)
        param_layout.addWidget(retrain_button)
        
        param_layout.addStretch(1)
        
        self.tab_widget.addTab(param_tab, "Parameters")
    
    def _create_analysis_tab(self):
        """创建分析结果标签页"""
        analysis_tab = QWidget()
        analysis_layout = QVBoxLayout(analysis_tab)
        
        # 模型质量组
        quality_group = QGroupBox("Model Quality")
        quality_layout = QFormLayout(quality_group)
        
        self.log_likelihood_label = QLabel("N/A")
        quality_layout.addRow("Log Likelihood:", self.log_likelihood_label)
        
        self.aic_label = QLabel("N/A")
        quality_layout.addRow("AIC:", self.aic_label)
        
        self.bic_label = QLabel("N/A")
        quality_layout.addRow("BIC:", self.bic_label)
        
        analysis_layout.addWidget(quality_group)
        
        # 转换统计组
        transition_group = QGroupBox("Transition Statistics")
        transition_layout = QFormLayout(transition_group)
        
        self.total_transitions_label = QLabel("0")
        transition_layout.addRow("Total Transitions:", self.total_transitions_label)
        
        self.transition_rate_label = QLabel("0.000")
        transition_layout.addRow("Transition Rate (/s):", self.transition_rate_label)
        
        self.avg_dwell_time_label = QLabel("0.000")
        transition_layout.addRow("Avg Dwell Time (s):", self.avg_dwell_time_label)
        
        analysis_layout.addWidget(transition_group)
        
        # 状态统计组
        state_group = QGroupBox("State Statistics")
        state_layout = QVBoxLayout(state_group)
        
        # 创建状态统计标签的容器
        self.state_stats_container = QWidget()
        self.state_stats_layout = QFormLayout(self.state_stats_container)
        state_layout.addWidget(self.state_stats_container)
        
        analysis_layout.addWidget(state_group)
        
        analysis_layout.addStretch(1)
        
        self.tab_widget.addTab(analysis_tab, "Analysis")
    
    def _create_export_tab(self):
        """创建导出选项标签页"""
        export_tab = QWidget()
        export_layout = QVBoxLayout(export_tab)
        
        # 导出选项组
        export_options_group = QGroupBox("Export Options")
        export_options_layout = QVBoxLayout(export_options_group)
        
        self.save_original_check = QCheckBox("Save Original Data")
        self.save_original_check.setChecked(True)
        export_options_layout.addWidget(self.save_original_check)
        
        self.save_fitted_check = QCheckBox("Save Fitted States")
        self.save_fitted_check.setChecked(True)
        export_options_layout.addWidget(self.save_fitted_check)
        
        self.save_transitions_check = QCheckBox("Save Transition Matrix")
        self.save_transitions_check.setChecked(True)
        export_options_layout.addWidget(self.save_transitions_check)
        
        self.save_dwell_times_check = QCheckBox("Save Dwell Times")
        self.save_dwell_times_check.setChecked(True)
        export_options_layout.addWidget(self.save_dwell_times_check)
        
        self.save_figure_check = QCheckBox("Save Figure")
        self.save_figure_check.setChecked(True)
        export_options_layout.addWidget(self.save_figure_check)
        
        export_layout.addWidget(export_options_group)
        
        # 图形格式选择
        format_group = QGroupBox("Figure Format")
        format_layout = QHBoxLayout(format_group)
        
        format_label = QLabel("Format:")
        self.figure_format_combo = QComboBox()
        self.figure_format_combo.addItems(["PNG", "PDF", "SVG", "EPS"])
        format_layout.addWidget(format_label)
        format_layout.addWidget(self.figure_format_combo)
        
        export_layout.addWidget(format_group)
        
        # 导出按钮
        export_button = QPushButton("Export Results")
        export_button.clicked.connect(self.save_results)
        export_layout.addWidget(export_button)
        
        export_layout.addStretch(1)
        
        self.tab_widget.addTab(export_tab, "Export")
    
    def _update_poly_order_visibility(self):
        """更新多项式阶数控件的可见性"""
        self.hmm_poly_order_spin.setEnabled(
            not self.use_current_filter_check.isChecked() and 
            self.hmm_filter_type_combo.currentText() == "Savitzky-Golay"
        )
    
    def _preprocess_data_for_hmm(self, data, time_data):
        """为HMM分析预处理数据"""
        # 根据初始化时传入的滤波设置选择处理方式
        if self.used_current_filter:
            # 使用可视化中当前使用的滤波设置
            filter_type = self.parent.filterselection.currentText()
            kernel_size = self.parent.kernel_size_box.value()
            
            # 对于 Savitzky-Golay 滤波，还需要多项式阶数
            if filter_type == "Savitzky-Golay" and hasattr(self.parent, 'poly_order_spin'):
                poly_order = self.parent.poly_order_spin.value()
                filtered_data = self.parent._apply_filter_to_data(
                    data, filter_type, kernel_size, poly_order)
            else:
                filtered_data = self.parent._apply_filter_to_data(
                    data, filter_type, kernel_size)
        else:
            # 使用默认的中值滤波
            kernel_size = 15  # 默认窗口大小
            filtered_data = scipy.signal.medfilt(data, kernel_size=kernel_size)
        
        # 可选的下采样
        if len(filtered_data) > 10000:
            # 下采样因子，根据数据长度动态调整
            if len(filtered_data) > 50000:  # 非常大的数据集
                ds_factor = max(1, int(len(filtered_data) / 10000))
            elif len(filtered_data) > 10000:
                ds_factor = max(1, int(len(filtered_data) / 5000))
            else:
                ds_factor = 1
            ds_data = filtered_data[::ds_factor]
            ds_time = time_data[::ds_factor]
        else:
            ds_data = filtered_data
            ds_time = time_data
        
        return ds_data.reshape(-1, 1), ds_time

    def _train_initial_model(self):
        """训练初始HMM模型"""
        # 预处理数据
        processed_data, processed_time = self._preprocess_data_for_hmm(
            self.extension_data, self.time_data)
        
        # 保存处理后的数据供后续使用
        self.processed_extension_data = processed_data
        self.processed_time_data = processed_time
        
        self.model = hmm.GaussianHMM(
            n_components=self.n_states, 
            covariance_type="full", 
            n_iter=self.n_iter_spin.value(),
            init_params='tc'
        )
        
        # 使用处理后的数据拟合模型
        if self.initial_means is not None:
            self.model.means_ = self.initial_means.reshape(-1, 1)
        
        self.model.fit(processed_data)
        
        # 保存原始转移矩阵
        self.original_transmat = self.model.transmat_.copy()
        
        # 计算模型质量指标
        self._calculate_model_quality()
    
    def _calculate_model_quality(self):
        """计算模型质量指标"""
        try:
            # 计算对数似然
            log_likelihood = self.model.score(self.extension_data_for_hmm)
            
            # 计算参数数量
            n_params = (self.n_states * (self.n_states - 1) +  # 转移概率
                       self.n_states * 2)  # 均值和方差
            
            # 计算AIC和BIC
            n_samples = len(self.extension_data_for_hmm)
            aic = 2 * n_params - 2 * log_likelihood
            bic = np.log(n_samples) * n_params - 2 * log_likelihood
            
            # 更新UI
            self.log_likelihood_label.setText(f"{log_likelihood:.2f}")
            self.aic_label.setText(f"{aic:.2f}")
            self.bic_label.setText(f"{bic:.2f}")
            
        except Exception as e:
            print(f"Failed to calculate model quality: {e}")
    
    def _on_parameter_changed(self):
        """参数改变时的回调函数"""
        if self.model is not None and not self.updating:
            self.updating = True
            try:
                self._update_analysis()
            finally:
                self.updating = False
    
    def _retrain_model(self):
        """重新训练模型"""
        self._train_initial_model()
        self._update_analysis()
    
    def _update_analysis(self):
        """更新分析结果 - 确保状态映射一致性"""
        if self.model is None or self.original_transmat is None:
            return
        
        # 恢复原始转移矩阵
        self.model.transmat_ = self.original_transmat.copy()
        
        # 应用转移惩罚
        penalty_factor = self.transition_penalty_spin.value() / 100.0
        if penalty_factor > 0:
            enhanced_transmat = (self.original_transmat * (1 - penalty_factor) + 
                            np.eye(self.n_states) * penalty_factor)
            # 归一化转移矩阵
            enhanced_transmat = enhanced_transmat / enhanced_transmat.sum(axis=1, keepdims=True)
            self.model.transmat_ = enhanced_transmat
        
        # 重新预测状态序列
        self.hidden_states = self.model.predict(self.extension_data_for_hmm)
        
        # 应用最小驻留时间过滤
        min_dwell_points = self.min_dwell_time_spin.value()
        if min_dwell_points > 1:
            self.hidden_states = self._apply_minimum_dwell_time_filter(
                self.hidden_states, min_dwell_points
            )
        
        # 应用平滑
        if self.use_smoothing_check.isChecked():
            window_size = self.smoothing_window_spin.value()
            if window_size > 1:
                self.hidden_states = self._smooth_state_sequence(
                    self.hidden_states, window_size
                )
        
        # 创建状态映射和方波
        self._create_state_mapping_and_square_wave()
        
        # 使用映射后的状态序列计算有效转移矩阵 - 确保使用正确的映射
        mapped_states = np.array([self.state_map.get(s, s) for s in self.hidden_states])
        self.effective_transmat = self._calculate_effective_transition_matrix(mapped_states)
        
        # 计算统计信息
        self._calculate_statistics()
        
        # 更新图形
        self._update_plot()
        
        # 更新UI显示
        self._update_ui_statistics()
    
    def _create_state_mapping_and_square_wave(self):
        """创建状态映射和方波表示"""
        # 获取状态均值并排序
        state_means = []
        unique_states = np.unique(self.hidden_states)
        
        for i in unique_states:
            mask = (self.hidden_states == i)
            if np.any(mask):
                mean = np.mean(self.extension_data[mask])
                state_means.append((i, mean))
        
        # 确保我们有足够的状态
        if len(state_means) < self.n_states:
            print(f"警告: 只找到 {len(state_means)} 个状态，预期 {self.n_states} 个")
        
        # 按均值排序
        state_means.sort(key=lambda x: x[1])
        
        # 创建从原始状态到排序后状态的映射
        self.state_map = {old_state: new_state for new_state, (old_state, _) in enumerate(state_means)}
        
        # 创建方波表示
        self.square_wave = np.zeros_like(self.extension_data)
        for old_state, mean_val in state_means:
            mask = (self.hidden_states == old_state)
            self.square_wave[mask] = mean_val
    
    def _calculate_statistics(self):
        """计算各种统计信息"""
        # 計算驻留时间和步长大小
        self.dwell_times, self.step_sizes, self.step_times = self._calculate_dwell_time_and_steps()
        
        # 计算转换统计
        self.transition_stats = self._calculate_transition_rates(self.hidden_states, self.time_data)
    
    def _update_plot(self):
        """更新图形显示 - 修复转移矩阵显示"""
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        
        # 绘制原始数据
        ax.plot(self.time_data, self.extension_data, 'gray', alpha=0.5, label='Original Data', linewidth=1)
        
        # 绘制状态模型
        ax.plot(self.time_data, self.square_wave, 'r-', linewidth=2, label='State Model')
        
        # 添加状态水平线和标签 - 确保按映射后的状态顺序显示
        colors = ['blue', 'green', 'purple']
        
        # 获取唯一状态和它们的平均值
        unique_states = np.unique(self.hidden_states)
        state_means = []
        
        for state in unique_states:
            mask = (self.hidden_states == state)
            mean_val = np.mean(self.square_wave[mask])
            mapped_state = self.state_map.get(state, state)
            state_means.append((state, mapped_state, mean_val))
        
        # 按照平均值排序，确保从下到上显示
        state_means.sort(key=lambda x: x[2])
        
        for idx, (orig_state, mapped_state, mean_val) in enumerate(state_means):
            color = colors[idx % len(colors)]
            
            # 添加水平线
            ax.axhline(y=mean_val, color=color, linestyle='--', alpha=0.7)
            
            # 添加标签 - 使用映射后的状态索引
            if mapped_state < len(self.state_names):
                state_name = self.state_names[mapped_state]
            else:
                state_name = f"State {mapped_state+1}"
                
            label_text = f"{state_name} ({mean_val:.2f} nm)"
            ax.text(self.time_data[0] + (self.time_data[-1] - self.time_data[0]) * 0.05, 
                mean_val, label_text, color=color, va='center', ha='left', fontsize=9, 
                bbox=dict(facecolor='white', alpha=0.9, edgecolor='none'))
        
        ax.set_xlabel('Time (s)')
        ax.set_ylabel('Extension (nm)')
        ax.set_title(f'HMM Analysis - {self.n_states} States, Force: {self.force:.2f} pN')
        ax.legend()
        
        # 添加转移概率文本 - 修复状态名称映射
        textstr = "Transition Matrix:\n"
        
        # 对于转移矩阵的每个元素
        for i in range(self.n_states):
            for j in range(self.n_states):
                # 获取状态名称 - 直接使用索引，因为effective_transmat已经是基于映射状态
                from_name = self.state_names[i] if i < len(self.state_names) else f"State {i+1}"
                to_name = self.state_names[j] if j < len(self.state_names) else f"State {j+1}"
                
                # 显示概率值
                textstr += f"P({from_name[:4]} → {to_name[:4]}) = {self.effective_transmat[i, j]:.3f}\n"
        
        props = dict(boxstyle='round', facecolor='wheat', alpha=0.5)
        ax.text(0.05, 0.95, textstr, transform=ax.transAxes, fontsize=8,
                verticalalignment='top', bbox=props)
        
        self.canvas.draw_idle()
    
    def _update_ui_statistics(self):
        """更新UI中的统计信息显示"""
        # 更新转换统计
        self.total_transitions_label.setText(str(self.transition_stats.get('total_transitions', 0)))
        self.transition_rate_label.setText(f"{self.transition_stats.get('transition_rate', 0):.3f}")
        self.avg_dwell_time_label.setText(f"{self.transition_stats.get('avg_dwell_time', 0):.3f}")
        
        # 清除旧的状态统计
        while self.state_stats_layout.count():
            item = self.state_stats_layout.takeAt(0)
            if item:
                if item.widget():
                    item.widget().deleteLater()
        
        # 添加新的状态统计
        for state_idx in range(self.n_states):
            # 使用映射后的状态索引
            mapped_idx = self.state_map.get(state_idx, state_idx)
            if mapped_idx < len(self.state_names):
                state_name = self.state_names[mapped_idx]
                
                # 计算该状态的统计信息
                mask = (self.hidden_states == state_idx)
                if np.any(mask):
                    occupancy = np.sum(mask) / len(self.hidden_states) * 100
                    mean_extension = np.mean(self.extension_data[mask])
                    
                    # 添加到UI
                    occupancy_label = QLabel(f"{occupancy:.1f}%")
                    mean_label = QLabel(f"{mean_extension:.2f} nm")
                    
                    self.state_stats_layout.addRow(f"{state_name} Occupancy:", occupancy_label)
                    self.state_stats_layout.addRow(f"{state_name} Mean:", mean_label)

    def _apply_minimum_dwell_time_filter(self, states, min_dwell_time):
        """应用最小驻留时间过滤器 - 优化版本"""
        if min_dwell_time <= 1:
            return states
            
        filtered_states = states.copy()
        
        # 找到所有状态转换点
        change_points = np.where(np.diff(states) != 0)[0] + 1
        
        if len(change_points) == 0:
            return filtered_states
        
        # 添加起始和结束点
        all_boundaries = np.concatenate(([0], change_points, [len(states)]))
        
        # 迭代处理短片段，直到没有更多短片段为止
        changed = True
        max_iterations = 10  # 防止无限循环
        iteration = 0
        
        while changed and iteration < max_iterations:
            changed = False
            iteration += 1
            
            # 重新计算边界
            change_points = np.where(np.diff(filtered_states) != 0)[0] + 1
            if len(change_points) == 0:
                break
                
            all_boundaries = np.concatenate(([0], change_points, [len(filtered_states)]))
            
            # 检查每个状态片段的长度
            for i in range(len(all_boundaries) - 1):
                start_idx = all_boundaries[i]
                end_idx = all_boundaries[i + 1]
                segment_length = end_idx - start_idx
                
                # 如果片段太短，将其合并到相邻的较长片段
                if segment_length < min_dwell_time:
                    current_state = filtered_states[start_idx]
                    
                    # 找到相邻的状态
                    prev_state = filtered_states[start_idx - 1] if start_idx > 0 else None
                    next_state = filtered_states[end_idx] if end_idx < len(filtered_states) else None
                    
                    # 选择合并到哪个相邻状态
                    merge_state = None
                    if prev_state is not None and next_state is not None:
                        # 选择相邻片段中较长的那个状态
                        prev_segment_length = self._get_segment_length_at_index(filtered_states, start_idx - 1)
                        next_segment_length = self._get_segment_length_at_index(filtered_states, end_idx)
                        merge_state = prev_state if prev_segment_length >= next_segment_length else next_state
                    elif prev_state is not None:
                        merge_state = prev_state
                    elif next_state is not None:
                        merge_state = next_state
                    
                    # 执行合并
                    if merge_state is not None and merge_state != current_state:
                        filtered_states[start_idx:end_idx] = merge_state
                        changed = True
                        break  # 重新开始检查
        
        return filtered_states
    
    def _get_segment_length_at_index(self, states, index):
        """获取指定索引处状态片段的长度"""
        if index < 0 or index >= len(states):
            return 0
            
        current_state = states[index]
        
        # 向前找到片段开始
        start = index
        while start > 0 and states[start - 1] == current_state:
            start -= 1
        
        # 向后找到片段结束
        end = index
        while end < len(states) - 1 and states[end + 1] == current_state:
            end += 1
        
        return end - start + 1
    
    def _smooth_state_sequence(self, states, window_size=9):
        """对状态序列进行平滑处理 - 增强版"""
        if len(states) < window_size or window_size <= 1:
            return states
        
        # 确保窗口大小为奇数
        if window_size % 2 == 0:
            window_size += 1
        
        smoothed_states = states.copy()
        half_window = window_size // 2
        
        # 获取有效状态集合
        valid_states = np.unique(states)
        
        for i in range(half_window, len(states) - half_window):
            # 提取窗口内的状态
            window = states[i - half_window:i + half_window + 1]
            
            # 使用众数（最频繁的状态）作为当前点的状态
            unique_states, counts = np.unique(window, return_counts=True)
            most_frequent_state = unique_states[np.argmax(counts)]
            
            # 只有当众数与当前状态不同，且众数占窗口的大部分时才改变
            max_count = np.max(counts)
            if max_count > window_size * 0.6 and most_frequent_state in valid_states:  # 至少60%的点是同一状态
                smoothed_states[i] = most_frequent_state
        
        return smoothed_states
    
    def _calculate_transition_rates(self, states, time_data):
        """计算状态转换速率 - 优化版本"""
        # 计算转换次数
        transitions = np.sum(np.diff(states) != 0)
        total_time = time_data[-1] - time_data[0]
        transition_rate = transitions / total_time if total_time > 0 else 0
        
        # 计算平均驻留时间
        change_points = np.where(np.diff(states) != 0)[0] + 1
        if len(change_points) > 0:
            all_boundaries = np.concatenate(([0], change_points, [len(states) - 1]))
            dwell_times = []
            
            for i in range(len(all_boundaries) - 1):
                start_idx = all_boundaries[i]
                end_idx = all_boundaries[i + 1]
                
                # 确保索引有效
                if end_idx < len(time_data):
                    dwell_time = time_data[end_idx] - time_data[start_idx]
                    if dwell_time > 0:  # 只包含有效的驻留时间
                        dwell_times.append(dwell_time)
            
            if dwell_times:
                avg_dwell_time = np.mean(dwell_times)
                min_dwell_time = np.min(dwell_times)
            else:
                avg_dwell_time = total_time
                min_dwell_time = total_time
        else:
            avg_dwell_time = total_time
            min_dwell_time = total_time
        
        return {
            'transition_rate': transition_rate,
            'avg_dwell_time': avg_dwell_time,
            'min_dwell_time': min_dwell_time,
            'total_transitions': transitions
        }
    
    def _calculate_effective_transition_matrix(self, states):
        """计算状态序列的有效转移矩阵"""
        if len(states) < 2:
            return np.eye(self.n_states)
            
        # 初始化转移计数矩阵
        trans_count = np.zeros((self.n_states, self.n_states))
        
        # 计算转移计数
        for i in range(len(states) - 1):
            from_state = int(states[i])
            to_state = int(states[i + 1])
            if 0 <= from_state < self.n_states and 0 <= to_state < self.n_states:
                trans_count[from_state, to_state] += 1
        
        # 计算转移概率
        state_counts = np.sum(trans_count, axis=1)
        effective_transmat = np.zeros_like(trans_count, dtype=float)
        
        for i in range(self.n_states):
            if state_counts[i] > 0:
                effective_transmat[i] = trans_count[i] / state_counts[i]
            else:
                # 如果没有该状态的转移，设置为自转移概率为1
                effective_transmat[i, i] = 1.0
        
        return effective_transmat

    def save_results(self):
        """保存分析结果 - 确保与显示一致"""
        # 获取基本文件名
        base_filename = f"{self.parent.base_name}_{self.bead_name}_HMM_{self.n_states}states_{self.force:.2f}pN"
        file_path = os.path.join(self.parent.Data_Saved_Path, f"{base_filename}.xlsx")
        
        try:
            # 保存图形
            if self.save_figure_check.isChecked():
                figure_format = self.figure_format_combo.currentText().lower()
                figure_path = os.path.join(self.parent.Data_Saved_Path, f"{base_filename}.{figure_format}")
                self.fig.savefig(figure_path, format=figure_format, dpi=300, bbox_inches='tight')
            
            # 创建Excel文件
            if (self.save_original_check.isChecked() or self.save_fitted_check.isChecked() or 
                self.save_transitions_check.isChecked() or self.save_dwell_times_check.isChecked()):
                
                wb = openpyxl.Workbook()
                
                # 参数信息工作表 - 保存当前使用的所有参数
                ws_params = wb.active
                ws_params.title = "Analysis Parameters"
                ws_params.append(["Parameter", "Value"])
                ws_params.append(["Force (pN)", f"{self.force:.2f}"])
                ws_params.append(["Number of States", self.n_states])
                ws_params.append(["Transition Penalty", f"{self.transition_penalty_spin.value()}"])
                ws_params.append(["Min Dwell Time", f"{self.min_dwell_time_spin.value()} points"])
                ws_params.append(["Smoothing Window", f"{self.smoothing_window_spin.value()}"])
                ws_params.append(["Smoothing Enabled", str(self.use_smoothing_check.isChecked())])
                ws_params.append(["Used Current Filter", str(self.used_current_filter)])
                ws_params.append(["Analysis Date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
                
                # 状态信息工作表 - 添加状态映射信息
                ws_states = wb.create_sheet("State Information")
                ws_states.append(["State Index", "State Name", "Mean Value (nm)", "Occupancy (%)"])
                
                # 添加每个状态的信息
                for i in range(self.n_states):
                    state_name = self.state_names[i] if i < len(self.state_names) else f"State {i+1}"
                    mask = (np.array([self.state_map.get(s, s) for s in self.hidden_states]) == i)
                    if np.any(mask):
                        mean_val = np.mean(self.extension_data[mask])
                        occupancy = np.sum(mask) / len(self.hidden_states) * 100
                        ws_states.append([i+1, state_name, mean_val, occupancy])
                
                # 主数据工作表
                ws_main = wb.create_sheet(f"HMM Analysis {self.force:.2f}pN")
                
                # 添加标题
                headers = ["Time (s)"]
                if self.save_original_check.isChecked():
                    headers.append("Original Extension (nm)")
                if self.save_fitted_check.isChecked():
                    headers.append("Fitted State (nm)")
                    headers.append("Original State Index")
                    headers.append("Mapped State Index")
                    headers.append("State Name")
                
                ws_main.append(headers)
                
                # 添加数据 - 包括映射信息
                mapped_states = np.array([self.state_map.get(s, s) for s in self.hidden_states])
                for i in range(len(self.time_data)):
                    row = [self.time_data[i]]
                    if self.save_original_check.isChecked():
                        row.append(self.extension_data[i])
                    if self.save_fitted_check.isChecked():
                        row.append(self.square_wave[i])
                        row.append(int(self.hidden_states[i]))
                        mapped_state = int(mapped_states[i])
                        row.append(mapped_state)
                        # 添加状态名称
                        if mapped_state < len(self.state_names):
                            row.append(self.state_names[mapped_state])
                        else:
                            row.append(f"State {mapped_state+1}")
                    ws_main.append(row)
                
                # 转移矩阵工作表 - 使用有效转移矩阵
                if self.save_transitions_check.isChecked():
                    ws_trans = wb.create_sheet("Transition Matrix")
                    
                    # 添加表头 - 使用与显示相同的状态名称
                    header_row = [""]
                    for i in range(self.n_states):
                        if i < len(self.state_names):
                            header_row.append(f"To {self.state_names[i]}")
                        else:
                            header_row.append(f"To State {i+1}")
                    ws_trans.append(header_row)
                    
                    # 添加转移矩阵数据 - 确保与显示一致
                    for i in range(self.n_states):
                        row = []
                        if i < len(self.state_names):
                            row.append(f"From {self.state_names[i]}")
                        else:
                            row.append(f"From State {i+1}")
                            
                        for j in range(self.n_states):
                            # 使用有效转移矩阵 - 与图形显示相同
                            row.append(self.effective_transmat[i, j])
                        ws_trans.append(row)
                    
                    # 添加转移矩阵说明
                    ws_trans.append([])
                    ws_trans.append(["Note: This transition matrix is calculated from the processed state sequence",
                                "after applying minimum dwell time filtering and smoothing (if enabled)."])
                    ws_trans.append(["It represents the actual transitions in the displayed results."])
                
                # 驻留时间工作表 - 使用映射后的状态
                if self.save_dwell_times_check.isChecked():
                    # 重新计算驻留时间，确保与显示一致
                    mapped_dwell_times = []
                    
                    # 使用映射后的状态序列
                    mapped_seq = np.array([self.state_map.get(s, s) for s in self.hidden_states])
                    
                    # 找到所有状态转换点
                    change_points = np.where(np.diff(mapped_seq) != 0)[0] + 1
                    
                    if len(change_points) > 0:
                        # 添加起始和结束点
                        all_boundaries = np.concatenate(([0], change_points, [len(mapped_seq)]))
                        
                        # 计算每个状态区间的驻留时间
                        for i in range(len(all_boundaries) - 1):
                            start_idx = all_boundaries[i]
                            end_idx = all_boundaries[i+1]
                            
                            # 获取当前状态
                            current_state = mapped_seq[start_idx]
                            
                            # 计算驻留时间
                            if end_idx < len(self.time_data):
                                dwell_time = self.time_data[end_idx] - self.time_data[start_idx]
                                if dwell_time > 0:  # 只包含有效的驻留时间
                                    mapped_dwell_times.append((current_state, dwell_time))
                    
                    # 为每个状态创建工作表
                    for state_idx in range(self.n_states):
                        state_dwells = [(dt, i) for i, (st, dt) in enumerate(mapped_dwell_times) if st == state_idx]
                        
                        if state_dwells:
                            state_name = self.state_names[state_idx] if state_idx < len(self.state_names) else f"State {state_idx+1}"
                            ws_state = wb.create_sheet(f"{state_name} Dwell Times")
                            ws_state.append(["Index", "Dwell Time (s)"])
                            
                            for idx, (dt, orig_idx) in enumerate(state_dwells):
                                ws_state.append([idx+1, dt])
                    
                    # 步长工作表 - 重新计算以确保与显示一致
                    ws_steps = wb.create_sheet("Step Sizes")
                    ws_steps.append(["Index", "From State", "To State", "Step Size (nm)", "Time (s)"])
                    
                    # 计算步长
                    step_idx = 1
                    for i in range(len(change_points)):
                        if i < len(change_points):
                            idx = change_points[i]
                            if idx > 0 and idx < len(mapped_seq):
                                from_state = mapped_seq[idx-1]
                                to_state = mapped_seq[idx]
                                
                                # 获取状态名称
                                from_name = self.state_names[from_state] if from_state < len(self.state_names) else f"State {from_state+1}"
                                to_name = self.state_names[to_state] if to_state < len(self.state_names) else f"State {to_state+1}"
                                
                                # 计算步长
                                before_mean = np.mean(self.extension_data[max(0, idx-10):idx])
                                after_mean = np.mean(self.extension_data[idx:min(idx+10, len(self.extension_data))])
                                step_size = after_mean - before_mean
                                
                                ws_steps.append([step_idx, from_name, to_name, step_size, self.time_data[idx]])
                                step_idx += 1
                
                wb.save(file_path)
                
                # 显示成功消息
                msg = "Saved Files:\n"
                if (self.save_original_check.isChecked() or self.save_fitted_check.isChecked() or 
                    self.save_transitions_check.isChecked() or self.save_dwell_times_check.isChecked()):
                    msg += f"- Data: {file_path}\n"
                if self.save_figure_check.isChecked():
                    msg += f"- Figure: {figure_path}"
                
                QMessageBox.information(self, "Save Complete", msg)
                
        except Exception as e:
            QMessageBox.critical(self, "Save Error", f"Failed to save results: {str(e)}")

    def _calculate_dwell_time_and_steps(self, min_threshold=1.0, max_threshold=20.0):
        """计算驻留时间和步长大小，类似于MATLAB的dwelltime.m
        
        Args:
            min_threshold: 最小步长阈值，小于此值的变化不考虑为状态跳变
            max_threshold: 最大步长阈值，大于此值的变化被视为噪声尖峰
            
        Returns:
            tuple: (驻留时间列表, 步长大小列表, 步长时间列表)
        """
        dwell_times = []  # 格式: [(状态, 驻留时间), ...]
        step_sizes = []   # 步长大小
        step_times = []   # 步长发生的时间点
        
        # 获取重新映射的状态序列
        mapped_states = np.array([self.state_map.get(s, s) for s in self.hidden_states])
        
        # 找到所有状态转换点
        change_points = np.where(np.diff(mapped_states) != 0)[0]
        
        if len(change_points) == 0:
            return dwell_times, step_sizes, step_times
        
        # 添加起始点和结束点来计算完整的驻留时间
        all_points = np.concatenate(([0], change_points, [len(mapped_states)-1]))
        
        # 计算每个状态区间的驻留时间和步长
        for i in range(len(all_points)-1):
            start_idx = all_points[i]
            end_idx = all_points[i+1]
            
            # 获取当前状态
            current_state = mapped_states[start_idx]
            
            # 计算驻留时间
            if i < len(all_points) - 2:  # 非最后一个区间
                dwell_time = self.time_data[end_idx] - self.time_data[start_idx]
                dwell_times.append((current_state, dwell_time))
                
                # 计算状态转换的步长大小
                next_state = mapped_states[end_idx + 1]
                step_mean_before = np.mean(self.extension_data[start_idx:end_idx+1])
                step_mean_after = np.mean(self.extension_data[end_idx+1:min(end_idx+11, len(self.extension_data))])
                step_size = step_mean_after - step_mean_before
                
                # 只记录有效步长
                abs_step = abs(step_size)
                if min_threshold <= abs_step <= max_threshold:
                    step_sizes.append(step_size)
                    step_times.append(self.time_data[end_idx])
        
        return dwell_times, step_sizes, step_times