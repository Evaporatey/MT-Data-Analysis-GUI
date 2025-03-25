# author: Ye Yang
# For force-extension analysis of MT data
# Inculde WLC and eWLC models for fitting
# Jump detection and event analysis
# Save data and figures
# 注意：多段拟合存在bug尚未修复，每选择一个区间拟合后，使用中键随意选择两个点来重置，否则会卡住，问题未知。


import math
import os

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import numpy as np
import scipy
from PySide6.QtWidgets import (QComboBox, QHBoxLayout, QPushButton, QSizePolicy, QVBoxLayout, QWidget, QLabel,
                               QCheckBox, QSpinBox, QRadioButton, QButtonGroup, QMessageBox, QFormLayout, QLineEdit, QGroupBox, QTabWidget,
                               QGridLayout)
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT, FigureCanvasQTAgg
from PySide6.QtCore import Qt

from tdms_reader import read_tdms_file
from scipy.interpolate import interp1d
from scipy.optimize import curve_fit


def WLC_inv(F, Lo, Lp, T, flag=1):
    """
    Worm-like chain model for polymer extension analysis
    
    Args:
        F: Force (pN)
        Lo: Contour length (nm)
        Lp: Persistence length (nm)
        T: Temperature (K)
        flag: Whether to include additional tension term
        
    Returns:
        Predicted extension length
    """
    kB = 1.38064852e-23  # Boltzmann constant
    T_K = T + 273.15 if T < 100 else T  # Ensure temperature in Kelvin
    kBT = kB * T_K * 1e21  # Convert to pN·nm units
    
    # 防止除零错误
    F_safe = np.maximum(F, 0.01)  # 最小力值设为0.01pN
    
    extension = Lo * (1 - 0.5 * np.sqrt(kBT / (F_safe * Lp)) + F_safe / 50) if flag else Lo * (1 - 0.5 * np.sqrt(kBT / (F_safe * Lp)))
    return extension

def eWLC_inv(F, Lo, Lp, T, Ko, flag=1):
    """
    Extensible worm-like chain model, considering elastic stretching
    
    Args:
        F: Force (pN)
        Lo: Contour length (nm)
        Lp: Persistence length (nm)
        T: Temperature (K)
        Ko: Stretch modulus (pN)
        flag: Whether to include additional tension term
        
    Returns:
        Predicted extension length
    """
    kB = 1.38064852e-23  # Boltzmann constant
    T_K = T + 273.15 if T < 100 else T  # Ensure temperature in Kelvin
    kBT = kB * T_K * 1e21  # Convert to pN·nm units
    
    # 防止除零错误
    F_safe = np.maximum(F, 0.01)  # 最小力值设为0.01pN
    
    extension = Lo * (1 - 0.5 * np.sqrt(kBT / (F_safe * Lp)) + F_safe / Ko) if flag else Lo * (1 - 0.5 * np.sqrt(kBT / (F_safe * Lp)))
    return extension

# WLC拟合函数，用于curve_fit
def WLC_fit(F, Lo, Lp):
    """用于拟合的WLC模型函数"""
    T = 300  # 默认温度300K
    return WLC_inv(F, Lo, Lp, T)

# eWLC拟合函数，用于curve_fit
def eWLC_fit(F, Lo, Lp, Ko):
    """用于拟合的eWLC模型函数"""
    T = 300  # 默认温度300K
    return eWLC_inv(F, Lo, Lp, T, Ko)


class FigureView(QWidget):
    def __init__(self, data_for_figure):
        super().__init__()

        # 初始化数据
        self._init_data(data_for_figure)
        
        # 初始化界面
        self._init_ui()
        
        # 初始化拟合曲线列表
        self.fit_curves = []
        
        # 初始化选择拟合区间模式
        self.selecting_fit_range = False
        self.fit_range_points = []
        
        # 添加拟合曲线颜色列表，用于区分不同的拟合段
        self.fit_colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', 
                           '#8c564b', '#e377c2', '#7f7f7f', '#bcbd22', '#17becf']
        
        # 绘制图表
        self.plotfig()

    def _init_ui(self):
        """初始化用户界面"""
        # 设置窗口属性
        self.setWindowTitle("Force-Extension Analysis")
        self.resize(1600, 900)  # 更大的窗口尺寸
        
        # 创建主布局为水平布局
        self.main_layout = QHBoxLayout(self)
        
        # 创建右侧控制面板（宽度固定）
        self.control_panel = QWidget()
        self.control_panel.setFixedWidth(300)  # 控制面板宽度固定
        self.control_layout = QVBoxLayout(self.control_panel)
        
        # 创建左侧图形面板（可伸缩）
        self.figure_panel = QWidget()
        self.figure_layout = QVBoxLayout(self.figure_panel)
        self.figure_layout.setContentsMargins(0, 0, 0, 0)
        
        # 初始化控制UI组件 - 移到这里，确保在初始化图表前创建控件
        self._init_control_ui()
        
        # 初始化图表 - 移到控件初始化之后
        self._init_figure()
        
        # 添加面板到主布局 - 调整顺序，图形在左，控制在右
        self.main_layout.addWidget(self.figure_panel, 1)  # 图形面板占据更多空间
        self.main_layout.addWidget(self.control_panel)

    def _init_data(self, data_for_figure):
        """初始化数据"""
        self.Data_Saved_Path = data_for_figure['Data_Saved_Path']
        self.file_name = data_for_figure['file_name']
        self.file_type = data_for_figure['file_type']
        self.file_info = data_for_figure['self.file_info']
        self.base_name = data_for_figure['base_name']

        # Load TDMS data
        self.tdms_data_frame = read_tdms_file(self.file_name,need_force=True)
        self.tdms_data_store = self.tdms_data_frame
        self.beads_list = self.tdms_data_frame.columns.values.tolist()
        
        # 处理磁铁移动状态数据
        str_magnet_move = str(self.beads_list[-1])
        magnet_move_state = self.tdms_data_store[str_magnet_move].values.tolist()
        new_magnet_move_state = [i for i in magnet_move_state if not math.isnan(i)]
        self.new_int_magnet_move_state = [int(i) for i in new_magnet_move_state]
        self.num_of_state = len(self.new_int_magnet_move_state) // 2
        self.final_slice_magnet_move_state = [self.new_int_magnet_move_state[i:i+2] 
                                             for i in range(0, len(self.new_int_magnet_move_state), 2)]
        self.num_of_sliced_data = len(self.final_slice_magnet_move_state)

    def _init_control_ui(self):
        """Initialize control UI components in sidebar"""
        # 创建标签页控件
        self.tab_widget = QTabWidget()
        
        # 创建"Basic & Filter"标签页（合并基本控制和滤波设置）
        basic_tab = QWidget()
        basic_layout = QVBoxLayout(basic_tab)
        
        # 添加数据选择下拉框
        data_group = QGroupBox("Data Selection")
        data_layout = QVBoxLayout(data_group)
        
        # Y轴选择
        y_layout = QHBoxLayout()
        y_label = QLabel("Y-Axis:")
        self.y_axis_box = QComboBox()
        for i in range(4, len(self.beads_list) - 1):
            self.y_axis_box.addItem(str(self.beads_list[i]))
        y_layout.addWidget(y_label)
        y_layout.addWidget(self.y_axis_box)
        data_layout.addLayout(y_layout)
        
        # X轴选择
        x_layout = QHBoxLayout()
        x_label = QLabel("X-Axis:")
        self.x_axis_box = QComboBox()
        for i in range(0, 3):
            self.x_axis_box.addItem(str(self.beads_list[i]))
        x_layout.addWidget(x_label)
        x_layout.addWidget(self.x_axis_box)
        data_layout.addLayout(x_layout)
        
        # 分段选择
        segment_layout = QHBoxLayout()
        segment_label = QLabel("Segment:")
        self.chose_sliced_data_box = QComboBox()
        self.chose_sliced_data_box.addItems([str(i) for i in range(1, self.num_of_sliced_data + 1)])
        segment_layout.addWidget(segment_label)
        segment_layout.addWidget(self.chose_sliced_data_box)
        data_layout.addLayout(segment_layout)
        
        basic_layout.addWidget(data_group)
        
        # 添加复选框
        options_group = QGroupBox("Options")
        options_layout = QVBoxLayout(options_group)
        
        self.sliced_data_box = QCheckBox("Image Segmentation")
        self.sliced_data_box.stateChanged.connect(self.plotfig)
        options_layout.addWidget(self.sliced_data_box)
        
        self.check_fitted_data_box = QCheckBox("Filtered Data")
        self.check_fitted_data_box.setChecked(True)
        self.check_fitted_data_box.stateChanged.connect(self.plotfig)
        options_layout.addWidget(self.check_fitted_data_box)
        
        # 新增平均简化数据选项
        self.average_data_box = QCheckBox("Average Simplified Data")
        self.average_data_box.setEnabled(False)  # 默认禁用
        self.average_data_box.stateChanged.connect(self.plotfig)
        options_layout.addWidget(self.average_data_box)
        
        basic_layout.addWidget(options_group)
        
        # 添加滤波器设置
        filter_group = QGroupBox("Filter Settings")
        filter_layout = QVBoxLayout(filter_group)
        
        # 滤波器类型
        self.filter_button_group = QButtonGroup(filter_group)
        self.median_filter_radio = QRadioButton('Median Filter')
        self.moving_avg_radio = QRadioButton('Moving Average')
        self.median_filter_radio.setChecked(True)
        
        self.filter_button_group.addButton(self.median_filter_radio)
        self.filter_button_group.addButton(self.moving_avg_radio)
        
        filter_layout.addWidget(self.median_filter_radio)
        filter_layout.addWidget(self.moving_avg_radio)
        
        # 核大小
        kernel_layout = QHBoxLayout()
        kernel_label = QLabel("Kernel Size:")
        self.kernel_size_box = QSpinBox()
        self.kernel_size_box.setRange(1, 100)
        self.kernel_size_box.setValue(3)
        self.kernel_size_box.setSingleStep(2)
        kernel_layout.addWidget(kernel_label)
        kernel_layout.addWidget(self.kernel_size_box)
        filter_layout.addLayout(kernel_layout)
        
        basic_layout.addWidget(filter_group)
        
        # 保存按钮
        save_group = QGroupBox("Save Options")
        save_layout = QVBoxLayout(save_group)
        
        self.save_FE_data_button = QPushButton("Save Point Data")
        self.save_FE_data_button.clicked.connect(self.save_data_to_excel)
        save_layout.addWidget(self.save_FE_data_button)
        
        self.save_FE_fig_button = QPushButton("Save All Data")
        self.save_FE_fig_button.clicked.connect(self.save_fig)
        save_layout.addWidget(self.save_FE_fig_button)
        
        basic_layout.addWidget(save_group)
        
        # 创建"Model"标签页
        model_tab = QWidget()
        model_layout = QVBoxLayout(model_tab)
        model_widget = self._create_model_controls()
        model_layout.addWidget(model_widget)
        
        # 创建"Advanced"标签页
        advanced_tab = QWidget()
        advanced_layout = QVBoxLayout(advanced_tab)
        advanced_widget = self._create_advanced_controls()
        advanced_layout.addWidget(advanced_widget)
        
        # 将标签页添加到标签页控件
        self.tab_widget.addTab(basic_tab, "Basic & Filter")
        self.tab_widget.addTab(model_tab, "Model")
        self.tab_widget.addTab(advanced_tab, "Advanced")
        
        # 添加数据信息显示
        info_group = QGroupBox("Data Info")
        info_layout = QVBoxLayout(info_group)
        
        force_layout = QHBoxLayout()
        force_label = QLabel("Force (pN):")
        self.force_data_info = QLabel("")
        force_layout.addWidget(force_label)
        force_layout.addWidget(self.force_data_info)
        info_layout.addLayout(force_layout)
        
        ext_layout = QHBoxLayout()
        ext_label = QLabel("Extension (nm):")
        self.extension_data_info = QLabel("")
        ext_layout.addWidget(ext_label)
        ext_layout.addWidget(self.extension_data_info)
        info_layout.addLayout(ext_layout)
        
        # 添加控件到控制面板
        self.control_layout.addWidget(self.tab_widget)
        self.control_layout.addWidget(info_group)
        
        # 连接信号
        self.y_axis_box.currentTextChanged.connect(self.plotfig)
        self.x_axis_box.currentTextChanged.connect(self.plotfig)
        self.chose_sliced_data_box.currentTextChanged.connect(self.plotfig)
        self.median_filter_radio.toggled.connect(self.plotfig)
        self.moving_avg_radio.toggled.connect(self.plotfig)
        self.kernel_size_box.valueChanged.connect(self.plotfig)

    def _create_basic_controls(self):
        """创建基本控制控件"""
        # Create horizontal layout for controls
        controlsWidget = QWidget()
        controlsLayout = QHBoxLayout(controlsWidget)
        controlsLayout.setContentsMargins(5, 5, 5, 5)
        controlsLayout.setSpacing(10)

        # Add radio buttons group
        self._add_radio_buttons(controlsLayout)
        
        # Add combo boxes and other controls
        self._add_combo_boxes(controlsLayout)
        self._add_checkboxes_and_spinbox(controlsLayout)
        self._add_labels_and_buttons(controlsLayout)
        
        return controlsWidget, controlsLayout

    def _create_filter_controls(self):
        """创建滤波控件组"""
        filterWidget = QWidget()
        filterLayout = QVBoxLayout(filterWidget)
        filterLayout.setContentsMargins(5, 5, 5, 5)
        filterLayout.setSpacing(10)
        
        # 添加滤波器选择单选按钮组
        self.filter_button_group = QButtonGroup(filterWidget)
        
        self.median_filter_radio = QRadioButton('Median Filter')
        self.median_filter_radio.setObjectName('median_filter_radio')
        self.median_filter_radio.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.median_filter_radio.setChecked(True)  # 默认选择中值滤波
        
        self.moving_avg_radio = QRadioButton('Moving Average')
        self.moving_avg_radio.setObjectName('moving_avg_radio')
        self.moving_avg_radio.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        
        self.filter_button_group.addButton(self.median_filter_radio)
        self.filter_button_group.addButton(self.moving_avg_radio)
        
        filterLayout.addWidget(self.median_filter_radio)
        filterLayout.addWidget(self.moving_avg_radio)
        
        # 创建内核大小控件和布局
        kernelWidget = QWidget()
        kernelLayout = QHBoxLayout(kernelWidget)
        kernelLayout.setContentsMargins(0, 0, 0, 0)
        
        kernelLabel = QLabel("Kernel Size:")
        kernelLayout.addWidget(kernelLabel)
        
        self.kernel_size_box = QSpinBox()
        self.kernel_size_box.setObjectName("kernel_size_box")
        self.kernel_size_box.setRange(1, 100)
        self.kernel_size_box.setValue(3)
        self.kernel_size_box.setSingleStep(2)
        self.kernel_size_box.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.kernel_size_box.valueChanged.connect(self.plotfig)
        kernelLayout.addWidget(self.kernel_size_box)
        
        filterLayout.addWidget(kernelWidget)
        
        return filterWidget

    def _create_model_controls(self):
        """Create model controls widget"""
        modelWidget = QWidget()
        modelLayout = QVBoxLayout(modelWidget)
        modelLayout.setContentsMargins(5, 5, 5, 5)
        modelLayout.setSpacing(10)
        
        # Add model selection label and dropdown
        modelLabel = QLabel("Theoretical Model:")
        modelLayout.addWidget(modelLabel)
        
        self.model_combo = QComboBox()
        self.model_combo.addItems(["None", "DNA", "RNA", "Protein", "Custom"])
        self.model_combo.currentTextChanged.connect(self.update_model_parameters)
        modelLayout.addWidget(self.model_combo)
        
        # Parameters group
        self.params_widget = QWidget()
        self.params_layout = QFormLayout(self.params_widget)
        self.params_layout.setVerticalSpacing(10)  # 增加垂直间距
        self.params_layout.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)  # 允许字段增长
        
        # Contour length
        self.Lo_input = QLineEdit("0")
        self.Lo_input.setToolTip("Contour length of molecule being studied (not including handles)")
        self.params_layout.addRow("Contour Length (nm):", self.Lo_input)
        
        # Persistence length
        self.Lp_input = QLineEdit("0")
        self.Lp_input.setToolTip("Persistence length defines how flexible the molecule is")
        self.params_layout.addRow("Persistence Length (nm):", self.Lp_input)
        
        # Temperature
        self.T_input = QLineEdit("300")
        self.T_input.setToolTip("Temperature in Kelvin (300K = 27°C)")
        self.params_layout.addRow("Temperature (K):", self.T_input)
        
        # Stretch modulus (only for eWLC)
        self.Ko_input = QLineEdit("400")
        self.Ko_input.setToolTip("Elastic stretch modulus (higher = stiffer)")
        self.params_layout.addRow("Stretch Modulus (pN):", self.Ko_input)
        
        # 添加DNA手柄长度输入框
        self.handle_length_input = QLineEdit("500")
        self.handle_length_input.setToolTip("Length of DNA handles used to tether the molecule")
        self.params_layout.addRow("DNA Handle Length (nm):", self.handle_length_input)
        
        # 改用垂直布局来避免模型类型选择控件的重叠
        model_type_group_box = QGroupBox("Model Type")
        model_type_layout = QVBoxLayout(model_type_group_box)
        model_type_layout.setContentsMargins(10, 10, 10, 10)
        model_type_layout.setSpacing(10)
        
        self.model_type_group = QButtonGroup()
        self.wlc_radio = QRadioButton("WLC Model")
        self.ewlc_radio = QRadioButton("eWLC Model")
        self.wlc_radio.setChecked(True)
        self.model_type_group.addButton(self.wlc_radio)
        self.model_type_group.addButton(self.ewlc_radio)
        
        model_type_layout.addWidget(self.wlc_radio)
        model_type_layout.addWidget(self.ewlc_radio)
        
        # 添加模型类型组
        self.params_layout.addRow(model_type_group_box)
        
        # 使用单独行添加复选框
        self.use_fit_checkbox = QCheckBox("Use Fitted Data")
        self.params_layout.addRow(self.use_fit_checkbox)
        
        # 添加WLC correction复选框
        self.wlc_correction_checkbox = QCheckBox("WLC Correction")
        self.wlc_correction_checkbox.setToolTip("Subtract DNA handle extension from measured extension")
        self.wlc_correction_checkbox.setEnabled(False)  # 默认禁用
        self.wlc_correction_checkbox.stateChanged.connect(self.plotfig)
        self.params_layout.addRow(self.wlc_correction_checkbox)
        
        # 按钮使用垂直布局
        button_layout = QVBoxLayout()
        button_layout.setSpacing(10)
        
        self.apply_model_button = QPushButton("Apply Model")
        self.apply_model_button.setMinimumHeight(30)  # 增加按钮高度
        self.apply_model_button.clicked.connect(self.apply_model)
        button_layout.addWidget(self.apply_model_button)
        
        self.fit_model_button = QPushButton("Fit Model")
        self.fit_model_button.setMinimumHeight(30)  # 增加按钮高度
        self.fit_model_button.clicked.connect(self.fit_model)
        button_layout.addWidget(self.fit_model_button)
        
        # 添加区间选择和多段拟合相关按钮
        self.select_range_button = QPushButton("Select Fitting Range")
        self.select_range_button.setMinimumHeight(30)
        self.select_range_button.clicked.connect(self.select_fitting_range)
        button_layout.addWidget(self.select_range_button)
        
        self.clear_fits_button = QPushButton("Clear All Fits")
        self.clear_fits_button.setMinimumHeight(30)
        self.clear_fits_button.clicked.connect(self.clear_all_fits)
        button_layout.addWidget(self.clear_fits_button)
        
        # 添加导出拟合参数和曲线按钮
        self.export_fit_button = QPushButton("Export Fit Results")
        self.export_fit_button.setMinimumHeight(30)
        self.export_fit_button.clicked.connect(self.export_fit_results)
        button_layout.addWidget(self.export_fit_button)
        
        button_widget = QWidget()
        button_widget.setLayout(button_layout)
        self.params_layout.addRow(button_widget)
        
        # Disable parameter controls until model is selected
        self.params_widget.setEnabled(False)
        
        modelLayout.addWidget(self.params_widget)
        
        # Status hint label
        self.model_status_label = QLabel("")
        self.model_status_label.setWordWrap(True)  # Allow text to wrap
        modelLayout.addWidget(self.model_status_label)
        
        return modelWidget

    def _create_advanced_controls(self):
    
        advancedWidget = QWidget()
        advancedLayout = QVBoxLayout(advancedWidget)
        advancedLayout.setContentsMargins(5, 5, 5, 5)
        advancedLayout.setSpacing(10)
        
        # 添加事件分析组
        event_group = QGroupBox("Event Analysis")
        event_layout = QVBoxLayout(event_group)
        
        # 力跳变阈值（增加上限）
        threshold_widget = QWidget()
        threshold_layout = QGridLayout(threshold_widget)
        
        threshold_layout.addWidget(QLabel("Min Threshold (nm):"), 0, 0)
        self.jump_threshold_input = QLineEdit("2.0")
        threshold_layout.addWidget(self.jump_threshold_input, 0, 1)
        
        threshold_layout.addWidget(QLabel("Max Threshold (nm):"), 1, 0)
        self.jump_threshold_max_input = QLineEdit("20.0")
        threshold_layout.addWidget(self.jump_threshold_max_input, 1, 1)
        
        self.detect_jumps_button = QPushButton("Detect Jumps")
        self.detect_jumps_button.clicked.connect(self.detect_force_jumps)
        threshold_layout.addWidget(self.detect_jumps_button, 2, 0, 1, 2)
        
        event_layout.addWidget(threshold_widget)
        
        # 移除几何校正相关代码
        
        advancedLayout.addWidget(event_group)
        
        # 添加数据导出按钮
        export_button = QPushButton("Export Analysis Results")
        export_button.clicked.connect(self.export_analysis_data)
        advancedLayout.addWidget(export_button)
        
        return advancedWidget

    def update_model_parameters(self):
        """Update parameters based on selected model"""
        selected_model = self.model_combo.currentText()
        
        if (selected_model == "None"):
            self.params_widget.setEnabled(False)
            return
        
        # 检查是否满足应用模型的条件
        is_force_x_axis = self.x_axis_box.currentText() and "force" in self.x_axis_box.currentText().lower()
        is_segmentation_enabled = self.sliced_data_box.isChecked()
        
        if not (is_force_x_axis and is_segmentation_enabled):
            QMessageBox.warning(self, "Model Application Conditions", 
                              "Models can only be applied when Force is the X-axis and Image Segmentation is enabled")
            self.model_combo.setCurrentText("None")
            return
        
        self.params_widget.setEnabled(True)
        
        # Preset parameter values
        if selected_model == "DNA":
            self.Lo_input.setText("340")  # nm/bp * 1000bp
            self.Lp_input.setText("50")   # ~50nm
            self.Ko_input.setText("1000") # pN
            self.ewlc_radio.setChecked(True)
            self.wlc_correction_checkbox.setEnabled(True)  # 启用WLC校正
        else:
            self.wlc_correction_checkbox.setEnabled(False)  # 其他模型禁用
        if selected_model == "RNA":
            self.Lo_input.setText("320")  # nm
            self.Lp_input.setText("1")    # lower persistence length
            self.Ko_input.setText("500")  # pN
            self.wlc_radio.setChecked(True)
        elif selected_model == "Protein":
            self.Lo_input.setText("0.4")  # nm/aa
            self.Lp_input.setText("0.6")  # nm
            self.Ko_input.setText("500")  # pN
            self.wlc_radio.setChecked(True)

    def apply_model(self):
        """Apply model and show theoretical curve"""
        # 检查是否满足应用模型的条件
        is_force_x_axis = self.x_axis_box.currentText() and "force" in self.x_axis_box.currentText().lower()
        is_segmentation_enabled = self.sliced_data_box.isChecked()
        
        if not (is_force_x_axis and is_segmentation_enabled):
            QMessageBox.warning(self, "Model Application Conditions", 
                               "Models can only be applied when Force is the X-axis and Image Segmentation is enabled")
            return
        
        # Get parameters
        try:
            Lo = float(self.Lo_input.text())
            Lp = float(self.Lp_input.text())
            T = float(self.T_input.text())
            Ko = float(self.Ko_input.text())
        except ValueError:
            QMessageBox.warning(self, "Parameter Error", "Please enter valid numerical parameters")
            return
        
        # Get current data
        extension_data, time_data, force_data = self._get_current_data()
        
        # Get current axis limits
        ax = self.fig.axes[0]
        x_min, x_max = ax.get_xlim()
        y_min, y_max = ax.get_ylim()
        
        # Save current lines to restore later
        existing_lines = []
        for line in ax.lines:
            if not hasattr(line, '_model_curve') or not line._model_curve:
                existing_lines.append((line.get_xdata(), line.get_ydata(), line.get_color(), line.get_label()))
        
        # Clear all lines
        ax.clear()
        
        # Restore original data lines
        for x_data, y_data, color, label in existing_lines:
            ax.plot(x_data, y_data, color=color, label=label)
        
        # Generate force data points (force as X-axis)
        force_range = np.linspace(0.1, max(force_data) * 1.1, 500)  # from 0.1 to 1.1 times max force
        
        # Calculate extension based on model type
        if self.wlc_radio.isChecked():
            extension = WLC_inv(force_range, Lo, Lp, T)
            model_name = "WLC Model"
        else:
            extension = eWLC_inv(force_range, Lo, Lp, T, Ko)
            model_name = "eWLC Model"
        
        # Add new theoretical curve (force as X-axis, extension as Y-axis)
        line = ax.plot(force_range, extension, 'g--', label=model_name)[0]
        line._model_curve = True
        
        # Set axis labels
        ax.set_xlabel("Force (pN)")
        ax.set_ylabel("Extension (nm)")
        ax.set_title("Force-Extension Relationship")
        
        # Update legend
        ax.legend()
        
        self.canvas.draw_idle()

    def fit_model(self):
        """Fit theoretical model to data"""
        # Check if conditions for applying model are met
        # 使用横坐标是否为力来判断，不再使用不存在的force_ramp_radio_box
        is_force_x_axis = self.x_axis_box.currentText() and "force" in self.x_axis_box.currentText().lower()
        is_segmentation_enabled = self.sliced_data_box.isChecked()
        
        if not (is_force_x_axis and is_segmentation_enabled):
            QMessageBox.warning(self, "Model Application Conditions", 
                              "Models can only be applied when Force is the X-axis and Image Segmentation is enabled")
            return
        
        # Get current data
        extension_data, time_data, force_data = self._get_current_data()
        
        # Ensure data is valid
        if len(force_data) < 5 or len(extension_data) < 5:
            QMessageBox.warning(self, "Insufficient Data", "Not enough data points for fitting")
            return
        
        # Perform fitting
        try:
            if self.wlc_radio.isChecked():
                # Initial parameters [Lo, Lp]
                p0 = [float(self.Lo_input.text()), float(self.Lp_input.text())]
                bounds = ([0, 0], [10000, 1000])  # Parameter ranges
                
                # Fit WLC model (note: force_data is x, extension_data is y)
                popt, pcov = curve_fit(WLC_fit, force_data, extension_data, p0=p0, bounds=bounds)
                Lo_fit, Lp_fit = popt
                
                # Update text fields
                self.Lo_input.setText(f"{Lo_fit:.2f}")
                self.Lp_input.setText(f"{Lp_fit:.2f}")
                
                # Display fit results
                model_name = f"WLC Fit (Lo={Lo_fit:.2f}nm, Lp={Lp_fit:.2f}nm)"
                
            else:
                # Initial parameters [Lo, Lp, Ko]
                p0 = [float(self.Lo_input.text()), float(self.Lp_input.text()), float(self.Ko_input.text())]
                bounds = ([0, 0, 0], [10000, 1000, 10000])
                
                # Define fit function
                def fit_func(F, Lo, Lp, Ko):
                    T = float(self.T_input.text())
                    return eWLC_inv(F, Lo, Lp, T, Ko)
                
                # Fit eWLC model
                popt, pcov = curve_fit(fit_func, force_data, extension_data, p0=p0, bounds=bounds)
                Lo_fit, Lp_fit, Ko_fit = popt
                
                # Update text fields
                self.Lo_input.setText(f"{Lo_fit:.2f}")
                self.Lp_input.setText(f"{Lp_fit:.2f}")
                self.Ko_input.setText(f"{Ko_fit:.2f}")
                
                # Display fit results
                model_name = f"eWLC Fit (Lo={Lo_fit:.2f}nm, Lp={Lp_fit:.2f}nm, Ko={Ko_fit:.2f}pN)"
            
            # Apply model (using fitted parameters)
            self.apply_model()
            
            QMessageBox.information(self, "Fit Success", f"Model fitting complete, parameters updated\n{model_name}")
            
        except Exception as e:
            QMessageBox.critical(self, "Fit Error", f"Model fitting failed: {str(e)}")

    def _init_figure(self):
        """初始化图表及相关事件"""
        self.fig = plt.figure(figsize=(10, 8))  # 更大的初始图形尺寸
        self.canvas = FigureCanvasQTAgg(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.toolbar = NavigationToolbar2QT(self.canvas, self)
        
        # 添加到图形面板
        self.figure_layout.addWidget(self.canvas)
        self.figure_layout.addWidget(self.toolbar)
        
        # 处理鼠标滚轮事件
        def wheelEvent(event):
            target = None
            step = 0
            
            if self.y_axis_box.underMouse():
                target = self.y_axis_box
            elif self.x_axis_box.underMouse():
                target = self.x_axis_box
            elif self.chose_sliced_data_box.underMouse():
                target = self.chose_sliced_data_box
            elif self.kernel_size_box.underMouse():
                if event.angleDelta().y() > 0:
                    self.kernel_size_box.setValue(self.kernel_size_box.value() - 2)
                else:
                    self.kernel_size_box.setValue(self.kernel_size_box.value() + 2)
                return
            else:
                super(FigureView, self).wheelEvent(event)
                return
                
            if target:
                step = -1 if event.angleDelta().y() > 0 else 1
                target.setCurrentIndex((target.currentIndex() + step) % target.count())
                
        self.y_axis_box.wheelEvent = wheelEvent
        self.x_axis_box.wheelEvent = wheelEvent
        self.chose_sliced_data_box.wheelEvent = wheelEvent
        
        # 处理缩放事件
        def zoom_event(event):
            axtemp = event.inaxes
            if not axtemp:
                return
                
            x_min, x_max = axtemp.get_xlim()
            y_min, y_max = axtemp.get_ylim()
            x_range = x_max - x_min
            y_range = y_max - y_min
            x_zoom = x_range / 10
            y_zoom = y_range / 10

            if event.button == 'up':
                axtemp.set(xlim=(x_min + x_zoom, x_max - x_zoom),
                           ylim=(y_min + y_zoom, y_max - y_zoom))
            elif event.button == 'down':
                axtemp.set(xlim=(x_min - x_zoom, x_max + x_zoom),
                           ylim=(y_min - y_zoom, y_max + y_zoom))
            self.canvas.draw_idle()

        self.canvas.mpl_connect('scroll_event', zoom_event)
        
        # 处理右键拖动事件
        self.lastx = 0
        self.lasty = 0
        self.press = False

        def on_press(event):
            if event.inaxes and event.button == 3:
                self.lastx = event.xdata
                self.lasty = event.ydata
                self.press = True

        def on_move(event):
            axtemp = event.inaxes
            if axtemp and self.press:
                x = event.xdata - self.lastx
                y = event.ydata - self.lasty
                x_min, x_max = axtemp.get_xlim()
                y_min, y_max = axtemp.get_ylim()
                axtemp.set(xlim=(x_min - x, x_max - x), ylim=(y_min - y, y_max - y))
                self.canvas.draw_idle()

        def on_release(event):
            self.press = False

        self.canvas.mpl_connect('button_press_event', on_press)
        self.canvas.mpl_connect('button_release_event', on_release)
        self.canvas.mpl_connect('motion_notify_event', on_move)
        
        # 处理中键点击事件（获取点）
        self.points_for_save = []
        
        def get_point(event):
            if event.inaxes and event.button == 2:
                if len(self.points_for_save) == 2:
                    self.points_for_save.clear()
                
                # 获取点击位置
                x_click = event.xdata
                y_click = event.ydata
                
                # 检查是否在选择拟合区间模式
                if hasattr(self, 'selecting_fit_range') and self.selecting_fit_range:
                    self.fit_range_points.append(x_click)
                    
                    if len(self.fit_range_points) == 2:
                        x_min, x_max = sorted(self.fit_range_points)
                        
                        # 应用模型到选定区间
                        self.apply_model_to_range(x_min, x_max)
                        
                        # 成功应用后，apply_model_to_range会调用reset_fitting_selection
                        # 不需要在这里重复重置
                    
                    return
                
                # 基于数据类型选择最近点的逻辑
                if len(self.points_for_save) == 0:
                    # 首先检查是否使用平均简化数据
                    # 1. WLC校正数据
                    if self.wlc_correction_checkbox.isChecked() and hasattr(self, 'corrected_x') and hasattr(self, 'corrected_y'):
                        closest_idx = self._find_closest_point(x_click, y_click, self.corrected_x, self.corrected_y)
                        if closest_idx is not None:
                            self.x1 = self.corrected_x[closest_idx]
                            self.y1 = self.corrected_y[closest_idx]
                        else:
                            self.x1 = x_click
                            self.y1 = y_click
                    elif self.average_data_box.isChecked() and self.average_data_box.isEnabled() and hasattr(self, 'avg_x') and hasattr(self, 'avg_y'):
                        # 使用平均数据中的最近点
                        closest_idx = self._find_closest_point(x_click, y_click, self.avg_x, self.avg_y)
                        if closest_idx is not None:
                            self.x1 = self.avg_x[closest_idx]
                            self.y1 = self.avg_y[closest_idx]
                        else:
                            self.x1 = x_click
                            self.y1 = y_click
                    elif self.check_fitted_data_box.isChecked():
                        # 使用滤波后数据中的最近点
                        # 获取当前绘制的数据
                        ax = self.fig.axes[0]
                        for line in ax.lines:
                            if line.get_label() and 'Filter' in line.get_label():
                                filter_x = line.get_xdata()
                                filter_y = line.get_ydata()
                                closest_idx = self._find_closest_point(x_click, y_click, filter_x, filter_y)
                                if closest_idx is not None:
                                    self.x1 = filter_x[closest_idx]
                                    self.y1 = filter_y[closest_idx]
                                    break
                        else:  # 如果没有找到滤波数据线
                            self.x1 = x_click
                            self.y1 = y_click
                    else:
                        # 使用原始曲线中的最近点
                        self.x1 = x_click
                        self.y1 = y_click
                        
                    self.points_for_save.append((self.x1, self.y1))
                else:
                    # 为第二个点应用相同的逻辑
                    # 1. WLC校正数据
                    if self.wlc_correction_checkbox.isChecked() and hasattr(self, 'corrected_x') and hasattr(self, 'corrected_y'):
                        closest_idx = self._find_closest_point(x_click, y_click, self.corrected_x, self.corrected_y)
                        if closest_idx is not None:
                            self.x2 = self.corrected_x[closest_idx]
                            self.y2 = self.corrected_y[closest_idx]
                        else:
                            self.x2 = x_click
                            self.y2 = y_click
                    elif self.average_data_box.isChecked() and self.average_data_box.isEnabled() and hasattr(self, 'avg_x') and hasattr(self, 'avg_y'):
                        closest_idx = self._find_closest_point(x_click, y_click, self.avg_x, self.avg_y)
                        if closest_idx is not None:
                            self.x2 = self.avg_x[closest_idx]
                            self.y2 = self.avg_y[closest_idx]
                        else:
                            self.x2 = x_click
                            self.y2 = y_click
                    elif self.check_fitted_data_box.isChecked():
                        ax = self.fig.axes[0]
                        for line in ax.lines:
                            if line.get_label() and 'Filter' in line.get_label():
                                filter_x = line.get_xdata()
                                filter_y = line.get_ydata()
                                closest_idx = self._find_closest_point(x_click, y_click, filter_x, filter_y)
                                if closest_idx is not None:
                                    self.x2 = filter_x[closest_idx]
                                    self.y2 = filter_y[closest_idx]
                                    break
                        else:
                            self.x2 = x_click
                            self.y2 = y_click
                    else:
                        self.x2 = x_click
                        self.y2 = y_click
                        
                    self.points_for_save.append((self.x2, self.y2))
                    self.extension_data = self.y2 - self.y1
                    
                    # 根据横坐标类型决定获取力数据的方式
                    is_force_x_axis = self.x_axis_box.currentText() and "force" in self.x_axis_box.currentText().lower()
                    
                    if is_force_x_axis:
                        # 如果横坐标是力，直接使用x1作为力值
                        self.force_data = self.x1
                    else:
                        # 如果横坐标是时间，通过时间点找对应的力值
                        time_point = min(self.xx, key=lambda x: abs(x - self.x1))
                        time_point_data = self.xx.tolist().index(time_point)
                        self.force_data_list = self.tdms_data_store.iloc[:, 1].values
                        self.force_data = self.force_data_list[time_point_data]
                    
                    self.force_data_info.setText(f"{self.force_data:.2f} pN")
                    self.extension_data_info.setText(f"{self.extension_data:.2f} nm")

                    # 在图表上显示选中的点
                    ax = self.fig.axes[0]
                    # 移除旧的标记点
                    for line in ax.lines:
                        if hasattr(line, '_selected_point') and line._selected_point:
                            line.remove()
                    # 添加新的标记点
                    for i, (px, py) in enumerate(self.points_for_save):
                        pt_label = f"Point {i+1}"
                        line = ax.plot(px, py, 'o', markersize=8, label=pt_label)[0]
                        line._selected_point = True
                    
                    self.canvas.draw_idle()

        self.canvas.mpl_connect('button_press_event', get_point)

    def _find_closest_point(self, x, y, x_data, y_data):
        """找到数据中离给定点最近的点的索引，处理NaN值和边界情况"""
        if len(x_data) == 0 or len(y_data) == 0:
            return None
        
        # 确保数据长度一致
        if len(x_data) != len(y_data):
            min_len = min(len(x_data), len(y_data))
            x_data = x_data[:min_len]
            y_data = y_data[:min_len]
        
        # 创建有效数据的掩码(非NaN值)
        valid_mask = ~np.isnan(x_data) & ~np.isnan(y_data)
        
        # 检查是否有足够的有效数据
        if np.sum(valid_mask) == 0:
            return None
        
        # 只使用有效数据计算距离
        valid_x = x_data[valid_mask]
        valid_y = y_data[valid_mask]
        
        # 计算欧几里得距离
        distances = np.sqrt((valid_x - x)**2 + (valid_y - y)**2)
        
        # 找到最小距离的索引
        min_dist_idx = np.argmin(distances)
        
        # 将索引映射回原始数据集
        original_indices = np.where(valid_mask)[0]
        closest_idx = original_indices[min_dist_idx]
        
        return closest_idx

    def _setup_figure_layout(self):
        """设置图表布局"""
        # 只清空图形，不尝试重建布局
        self.fig.clear()
        # 如果plotfig()是在__init__中首次调用，不需要删除现有widget
        # 直接使用figure_layout来管理canvas和toolbar


    def plotfig(self):
        """绘制图形"""
        # 获取所选数据
        self.chosen_bead = self.y_axis_box.currentText()
        self.chosen_x_data = self.x_axis_box.currentText()
        self.x_axis = self.tdms_data_store[str(self.chosen_x_data)]
        self.y_axis = self.tdms_data_store[str(self.chosen_bead)]
        self.xx = self.x_axis.values
        self.yy = self.y_axis.values
        
        # 检查是否同时满足平均简化数据的条件: 
        # 1. 选择force为横坐标
        # 2. 开启Image Segmentation
        is_force_x_axis = "force" in self.chosen_x_data.lower()
        is_segmentation_enabled = self.sliced_data_box.isChecked()
        self.average_data_box.setEnabled(is_force_x_axis and is_segmentation_enabled)
        
        plt.clf()
        self._setup_figure_layout()
        ax = self.fig.add_subplot(111)
        
        # 选择绘图数据
        if (self.sliced_data_box.isChecked()):
            self.chose_sliced_data_box.setEnabled(True)
            selected_sliced_data_num = int(self.chose_sliced_data_box.currentText()) - 1
            
            section = self.final_slice_magnet_move_state[selected_sliced_data_num]
            start_point = section[0] - 1 if section[0] != 0 else 0
            end_point = section[1] - 1
            
            axis_x = self.xx[start_point:end_point]
            axis_y = self.yy[start_point:end_point]
            
            # 获取力和延伸数据用于后续处理
            force_data = axis_x
            extension_data = axis_y
            
            # 首先独立处理WLC校正
            if self.wlc_correction_checkbox.isChecked() and self.wlc_correction_checkbox.isEnabled():
                corrected_extension = self.apply_wlc_correction(force_data, extension_data)
                self.corrected_x = force_data
                self.corrected_y = corrected_extension
                # 更新轴数据为校正后的数据
                axis_y = corrected_extension
            else:
                # 不使用WLC校正时清除属性
                if hasattr(self, 'corrected_x'):
                    del self.corrected_x
                if hasattr(self, 'corrected_y'):
                    del self.corrected_y
            
            # 然后处理平均简化数据
            if self.average_data_box.isChecked() and self.average_data_box.isEnabled():
                # 保存原始数据用于对比
                original_extension = axis_y.copy()  # 这里的axis_y可能已经是校正后的数据
                mag_height = self.tdms_data_store['mag height mm'].values[start_point:end_point]
                
                # 计算平均简化数据
                avg_x, avg_y = self._calculate_averaged_data(mag_height, force_data, axis_y)
                
                # 保存平均数据供其他函数使用
                self.avg_x = avg_x
                self.avg_y = avg_y
                
                # 绘制原始数据点（浅色）
                ax.plot(force_data, axis_y, 'o', color='lightgray', alpha=0.3, markersize=3, label='Original Points')
                
                # 更新轴数据为平均后的数据
                axis_x = avg_x
                axis_y = avg_y
            else:
                # 不使用平均数据时清除属性
                if hasattr(self, 'avg_x'):
                    del self.avg_x
                if hasattr(self, 'avg_y'):
                    del self.avg_y
        else:
            self.chose_sliced_data_box.setEnabled(False)
            self.average_data_box.setEnabled(False)
            axis_x = self.xx
            axis_y = self.yy
            
            # 不使用平均数据时清除属性
            if hasattr(self, 'avg_x'):
                del self.avg_x
            if hasattr(self, 'avg_y'):
                del self.avg_y
        
        # 绘制原始数据
        if not (self.average_data_box.isChecked() and self.average_data_box.isEnabled()):
            ax.plot(axis_x, axis_y, color='darkgrey', label='Raw Data')
        elif self.average_data_box.isChecked() and self.average_data_box.isEnabled():
            # 平均数据点已经在上面绘制，这里绘制连线
            ax.plot(axis_x, axis_y, 'o-', color='blue', markersize=5, label='Averaged Data')
        
        # 绘制拟合数据
        if self.check_fitted_data_box.isChecked():
            self.kernel_size_box.setEnabled(True)
            set_size = self.kernel_size_box.value()
            
            # 使用选定的滤波方法
            y_fitted = self._apply_filter(axis_y, set_size)
            
            # 根据滤波器类型设置标签
            filter_type = "Median Filter" if self.median_filter_radio.isChecked() else "Moving Average"
            ax.plot(axis_x, y_fitted, color='red', label=f'Filtered Data ({filter_type})')
        else:
            self.kernel_size_box.setEnabled(False)
        
        # 添加图例
        plt.legend(frameon=False, loc='upper right')
        
        # 在绘图代码之后添加坐标轴标签设置
        ax.set_xlabel(self.chosen_x_data)
        ax.set_ylabel(self.chosen_bead)
        ax.set_title(f"{self.chosen_bead} vs {self.chosen_x_data}")
        
        # 安全设置坐标轴范围，处理可能的NaN或Inf值
        # 对X轴范围
        x_min = np.nanmin(axis_x) if not np.all(np.isnan(axis_x)) and len(axis_x) > 0 else 0
        x_max = np.nanmax(axis_x) if not np.all(np.isnan(axis_x)) and len(axis_x) > 0 else 100
        
        # 确保x_min和x_max不是NaN或Inf
        if np.isnan(x_min) or np.isinf(x_min): x_min = 0
        if np.isnan(x_max) or np.isinf(x_max): x_max = 100
        if x_min == x_max: x_max = x_min + 1  # 避免相等的边界
        
        # 对Y轴范围
        y_min = np.nanmin(axis_y) if not np.all(np.isnan(axis_y)) and len(axis_y) > 0 else 0
        y_max = np.nanmax(axis_y) if not np.all(np.isnan(axis_y)) and len(axis_y) > 0 else 100
        
        # 确保y_min和y_max不是NaN或Inf
        if np.isnan(y_min) or np.isinf(y_min): y_min = 0
        if np.isnan(y_max) or np.isinf(y_max): y_max = 100
        if y_min == y_max: y_max = y_min + 1  # 避免相等的边界
        
        ax.set_xlim(x_min, x_max)
        ax.set_ylim(y_min, y_max)
        
        self.canvas.draw_idle()
        
        # 检查模型按钮的状态
        self.check_model_availability()
        
        # 设置WLC Correction复选框状态
        self.wlc_correction_checkbox.setEnabled(is_force_x_axis and is_segmentation_enabled and self.model_combo.currentText() == "DNA")

    def check_model_availability(self):
        """检查当前是否满足应用模型的条件"""
        # 检查是否满足应用模型的条件：force为横坐标 且 开启Image Segmentation
        is_force_x_axis = self.x_axis_box.currentText() and "force" in self.x_axis_box.currentText().lower()
        is_segmentation_enabled = self.sliced_data_box.isChecked()
        
        # 设置模型应用按钮和状态
        model_available = is_force_x_axis and is_segmentation_enabled
        
        if not is_force_x_axis:
            self.model_status_label.setText("Please select Force as X-axis to apply models")
            self.model_status_label.setStyleSheet("color: red")
        elif not is_segmentation_enabled:
            self.model_status_label.setText("Please enable Image Segmentation to apply models")
            self.model_status_label.setStyleSheet("color: red")
        else:
            self.model_status_label.setText("Theoretical models can be applied")
            self.model_status_label.setStyleSheet("color: green")

    def _get_current_data(self):
        """获取当前选中的数据，按优先级返回最合适的数据"""
        self.chosen_bead = self.y_axis_box.currentText()
        extension_data = self.tdms_data_store[str(self.chosen_bead)].values
        time_data = self.tdms_data_store.iloc[:, 0].values
        force_data = self.tdms_data_store.iloc[:, 1].values
        
        if self.sliced_data_box.isChecked():
            selected_num = int(self.chose_sliced_data_box.currentText()) - 1
            section = self.final_slice_magnet_move_state[selected_num]
            start_point = section[0] - 1 if section[0] != 0 else 0
            end_point = section[1] - 1
            
            segment_extension = extension_data[start_point:end_point]
            segment_time = time_data[start_point:end_point]
            segment_force = force_data[start_point:end_point]
            
            # 按照优先级返回数据
            # 1. WLC校正数据（独立于平均数据）
            if hasattr(self, 'corrected_x') and hasattr(self, 'corrected_y') and self.wlc_correction_checkbox.isChecked():
                return self.corrected_y, segment_time, self.corrected_x
            
            # 2. 平均简化数据
            elif hasattr(self, 'avg_x') and hasattr(self, 'avg_y') and self.average_data_box.isChecked():
                return self.avg_y, segment_time, self.avg_x
            
            # 3. 滤波数据
            elif self.check_fitted_data_box.isChecked():
                set_size = self.kernel_size_box.value()
                filtered_extension = self._apply_filter(segment_extension, set_size)
                return filtered_extension, segment_time, segment_force
            
            # 4. 原始数据
            else:
                return segment_extension, segment_time, segment_force
        else:
            return extension_data, time_data, force_data

    def _save_to_excel(self, file_path, sheet_name, headers, data_columns):
        """Save data to Excel file with proper formatting"""
        try:
            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(file_path)
                if sheet_name in workbook.sheetnames:
                    # If sheet exists, ask if user wants to overwrite
                    response = QMessageBox.question(
                        self, 'Sheet exists',
                        f'Sheet "{sheet_name}" already exists. Overwrite?',
                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                    )
                    if response == QMessageBox.Yes:
                        # Remove existing sheet and create a new one
                        idx = workbook.sheetnames.index(sheet_name)
                        workbook.remove(workbook[sheet_name])
                        worksheet = workbook.create_sheet(sheet_name, idx)
                    else:
                        # Append number to sheet name to make it unique
                        i = 1
                        new_name = f"{sheet_name}_{i}"
                        while new_name in workbook.sheetnames:
                            i += 1
                            new_name = f"{sheet_name}_{i}"
                        sheet_name = new_name
                        worksheet = workbook.create_sheet(sheet_name)
                else:
                    worksheet = workbook.create_sheet(sheet_name)
            else:
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = sheet_name
            
            # Add headers
            worksheet.append(headers)
            
            # Add data rows
            for i in range(len(data_columns[0])):
                row_data = [col[i] for col in data_columns if i < len(col)]
                worksheet.append(row_data)
                    
            workbook.save(file_path)
            QMessageBox.information(self, "Success", f"Data saved to {file_path}, sheet: {sheet_name}")
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save data: {str(e)}")
            return False

    def save_data_to_excel(self):
        """Save clicked point data to Excel"""
        if not hasattr(self, 'force_data') or not hasattr(self, 'extension_data'):
            QMessageBox.warning(self, "Warning", "Please use middle mouse button to select two points first")
            return
            
        xlsx_file_path = os.path.join(self.Data_Saved_Path, self.base_name + '.xlsx')
        sheet_name = "Points_Data"  # 统一使用一个sheet
        
        if os.path.exists(xlsx_file_path):
            workbook = openpyxl.load_workbook(xlsx_file_path)
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.create_sheet(sheet_name)
                # 新的表头包含Bead名称和Segment
                worksheet.append(['Bead', 'Segment', 'Force(pN)', 'Extension(nm)', 'Time'])
        else:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            worksheet.append(['Bead', 'Segment', 'Force(pN)', 'Extension(nm)', 'Time'])
        
        # 添加数据行，包含bead名称和segment信息
        segment = self.chose_sliced_data_box.currentText() if self.sliced_data_box.isChecked() else "N/A"
        bead = self.y_axis_box.currentText()
        import datetime
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        worksheet.append([bead, segment, self.force_data, self.extension_data, current_time])
        workbook.save(xlsx_file_path)
        
        QMessageBox.information(self, "Save Successful", 
                               f"Point data has been saved to {xlsx_file_path} in the {sheet_name} sheet")

    def save_fig(self):
        """Save data to Excel with all current data points and save figure"""
        # 保存数据到Excel（原有功能）
        xlsx_file_path = os.path.join(self.Data_Saved_Path, self.base_name + '_raw_and_fit.xlsx')
        
        if self.sliced_data_box.isChecked():
            sheet_name = f"{self.y_axis_box.currentText()}-{self.chose_sliced_data_box.currentText()}"
            # 获取分段的原始数据
            selected_num = int(self.chose_sliced_data_box.currentText()) - 1
            section = self.final_slice_magnet_move_state[selected_num]
            start_point = section[0] - 1 if section[0] != 0 else 0
            end_point = section[1] - 1
            
            # 获取原始数据（不经过_get_current_data处理）
            self.chosen_bead = self.y_axis_box.currentText()
            raw_extension = self.tdms_data_store[str(self.chosen_bead)].values[start_point:end_point]
            raw_time = self.tdms_data_store.iloc[:, 0].values[start_point:end_point]
            raw_force = self.tdms_data_store.iloc[:, 1].values[start_point:end_point]
        else:
            sheet_name = self.y_axis_box.currentText()
            # 获取全部原始数据
            self.chosen_bead = self.y_axis_box.currentText()
            raw_extension = self.tdms_data_store[str(self.chosen_bead)].values
            raw_time = self.tdms_data_store.iloc[:, 0].values
            raw_force = self.tdms_data_store.iloc[:, 1].values
        
        # 初始化headers和data_columns，始终使用原始数据作为基础
        headers = ['Time(s)', 'Force(pN)', 'Raw Extension(nm)']
        data_columns = [raw_time, raw_force, raw_extension]
        
        # 获取处理后的数据，用于添加额外的列
        processed_extension, processed_time, processed_force = self._get_current_data()
        
        # 检查是否使用滤波数据
        if self.check_fitted_data_box.isChecked():
            set_size = self.kernel_size_box.value()
            # 使用选定的滤波方法对原始数据进行滤波
            fitted_data = self._apply_filter(raw_extension, set_size)
            # 添加滤波器类型信息
            filter_type = "Median" if self.median_filter_radio.isChecked() else "MovingAvg"
            headers.append(f'{filter_type} Filtered Extension(nm)')
            data_columns.append(fitted_data)
        
        # 检查是否使用平均简化数据
        if self.average_data_box.isChecked() and self.average_data_box.isEnabled():
            if hasattr(self, 'avg_x') and hasattr(self, 'avg_y'):
                headers.extend(['Avg Force(pN)', 'Avg Extension(nm)'])
                
                # 调整数组长度，确保不重叠且对齐
                max_len = max(len(col) for col in data_columns)
                avg_force_padded = np.pad(self.avg_x, (0, max_len - len(self.avg_x)), 
                                         'constant', constant_values=np.nan)
                avg_extension_padded = np.pad(self.avg_y, (0, max_len - len(self.avg_y)), 
                                             'constant', constant_values=np.nan)
                data_columns.extend([avg_force_padded, avg_extension_padded])
        
        # 检查是否有WLC校正数据
        if hasattr(self, 'corrected_x') and hasattr(self, 'corrected_y') and self.wlc_correction_checkbox.isChecked():
            headers.append('WLC Corrected Extension(nm)')
            # 调整数组长度
            max_len = max(len(col) for col in data_columns)
            corrected_padded = np.pad(self.corrected_y, (0, max_len - len(self.corrected_y)), 
                                     'constant', constant_values=np.nan)
            data_columns.append(corrected_padded)
        
        # 保存Excel
        excel_saved = self._save_to_excel(xlsx_file_path, sheet_name, headers, data_columns)
        
        # 只有成功保存Excel后才保存图片
        if excel_saved:
            # 定义图片文件路径
            fig_file_path = os.path.join(self.Data_Saved_Path, f"{self.base_name}_{sheet_name}.png")
            
            # 检查图片文件是否已存在
            if os.path.exists(fig_file_path):
                # 如果图片已存在，询问用户是否覆盖、重命名或取消
                msg_box = QMessageBox()
                msg_box.setWindowTitle("图片文件已存在")
                msg_box.setText(f"图片文件 '{os.path.basename(fig_file_path)}' 已存在。")
                msg_box.setInformativeText("您想如何处理？")
                
                # 添加覆盖、重命名和取消按钮
                overwrite_button = msg_box.addButton("覆盖", QMessageBox.AcceptRole)
                rename_button = msg_box.addButton("重命名", QMessageBox.ActionRole)
                cancel_button = msg_box.addButton("取消", QMessageBox.RejectRole)
                
                msg_box.exec()
                
                clicked_button = msg_box.clickedButton()
                
                if clicked_button == cancel_button:
                    # 用户选择取消，不保存图片
                    QMessageBox.information(self, "部分保存成功", 
                                          f"数据已保存到: {xlsx_file_path}, 工作表: {sheet_name}\n"
                                          f"图片保存已取消。")
                    return
                elif clicked_button == rename_button:
                    # 用户选择重命名，生成新文件名
                    i = 1
                    base_name, ext = os.path.splitext(fig_file_path)
                    new_file_path = f"{base_name}_{i}{ext}"
                    while os.path.exists(new_file_path):
                        i += 1
                        new_file_path = f"{base_name}_{i}{ext}"
                    fig_file_path = new_file_path
            
            # 保存图片
            try:
                self.fig.savefig(fig_file_path, dpi=300, bbox_inches='tight')
                # 更新成功消息，显示数据和图片都保存成功
                QMessageBox.information(self, "保存成功", 
                                      f"数据保存到: {xlsx_file_path}, 工作表: {sheet_name}\n\n"
                                      f"图片保存到: {fig_file_path}")
            except Exception as e:
                QMessageBox.critical(self, "图片保存失败", f"保存图片时出错: {str(e)}\n"
                                   f"数据已保存到: {xlsx_file_path}")

    def _apply_filter(self, data, kernel_size):
        """根据选择的滤波器类型应用滤波算法
        
        Args:
            data: 输入数据
            kernel_size: 滤波器大小
            
        Returns:
            滤波后的数据
        """
        import numpy as np
        
        if self.median_filter_radio.isChecked():
            # 应用中值滤波
            return scipy.signal.medfilt(data, kernel_size=kernel_size)
        else:
            # 应用滑动平均滤波
            # 确保kernel_size是奇数
            if (kernel_size % 2) == 0:
                kernel_size += 1
                
            # 创建卷积核
            kernel = np.ones(kernel_size) / kernel_size
            
            # 应用卷积（等效于滑动平均）
            # 使用mode='same'确保输出长度与输入相同
            return np.convolve(data, kernel, mode='same')

    def detect_force_jumps(self):
        """检测力跳变事件"""
        if not hasattr(self, 'xx') or len(self.xx) == 0:
            QMessageBox.warning(self, "Data Error", "Please load and select data first")
            return
        
        # 确保已开启Image Segmentation
        if not self.sliced_data_box.isChecked():
            QMessageBox.warning(self, "Mode Error", "Please enable Image Segmentation to detect jumps")
            return
        
        # 获取当前数据
        extension_data, time_data, force_data = self._get_current_data()
        
        # 应用滤波减少噪声
        filtered_data = self._apply_filter(extension_data, self.kernel_size_box.value())
        
        # 检测跳变（使用一阶导数和阈值）
        diff_data = np.diff(filtered_data)
        try:
            min_threshold = float(self.jump_threshold_input.text())
            max_threshold = float(self.jump_threshold_max_input.text()) if hasattr(self, 'jump_threshold_max_input') else 20.0
        except (ValueError, AttributeError):
            min_threshold = 2.0
            max_threshold = 20.0
        
        # 使用最小和最大阈值范围检测跳变
        jump_indices = np.where((np.abs(diff_data) > min_threshold) & 
                            (np.abs(diff_data) < max_threshold))[0]
        
        # 合并临近的跳变（如果相距小于10个点）
        if len(jump_indices) > 0:
            merged_jumps = [jump_indices[0]]
            for idx in jump_indices[1:]:
                if idx - merged_jumps[-1] > 10:
                    merged_jumps.append(idx)
            jump_indices = np.array(merged_jumps)
        
        # 在图上标记跳变位置
        ax = self.fig.axes[0]
        
        # 移除旧的标记
        for line in ax.lines:
            if hasattr(line, '_jump_marker') and line._jump_marker:
                line.remove()
        
        # 添加新的标记
        for idx in jump_indices:
            if idx >= len(time_data) or idx >= len(filtered_data) or idx >= len(force_data):
                continue  # 避免索引超出范围
                
            # 检查是否将force作为x轴
            is_force_x_axis = self.x_axis_box.currentText() and "force" in self.x_axis_box.currentText().lower()
            
            if not is_force_x_axis:
                # 在时间轴上标记
                line = ax.axvline(x=time_data[idx], color='r', linestyle='-', alpha=0.5)
            else:
                # 在力-延伸图上标记
                line = ax.plot(force_data[idx], filtered_data[idx], 'ro', markersize=8)[0]
            line._jump_marker = True
        
        # 更新图形
        self.canvas.draw_idle()
        
        # 显示结果
        if len(jump_indices) > 0:
            QMessageBox.information(self, "Event Detection", 
                                f"Detected {len(jump_indices)} jump events")
            
            # 保存跳变事件数据
            self.jumps_data = {
                'time': time_data[jump_indices],
                'extension': filtered_data[jump_indices],
                'force': force_data[jump_indices],
                'jump_size': diff_data[jump_indices]
            }
        else:
            QMessageBox.information(self, "Event Detection", "No jump events detected")

    def _add_advanced_analysis_tab(self):
        """添加高级分析标签页"""
        # 创建高级分析面板
        self.advanced_panel = QWidget()
        advanced_layout = QVBoxLayout(self.advanced_panel)
        
        # 添加模型分析组
        model_group = QGroupBox("Theoretical Model Analysis")
        model_layout = QVBoxLayout(model_group)
        
        # 添加模型选择标签和下拉菜单
        modelLabel = QLabel("Theoretical Model:")
        model_layout.addWidget(modelLabel)
        
        model_combo = QComboBox()
        model_combo.addItems(["None", "DNA", "RNA", "Protein", "Custom"])
        model_combo.currentTextChanged.connect(self.update_model_parameters)
        model_layout.addWidget(model_combo)
        self.model_combo_advanced = model_combo  # 使用不同的名称避免冲突
        
        advanced_layout.addWidget(model_group)
        
        # 添加事件分析组
        event_group = QGroupBox("Event Analysis")
        event_layout = QVBoxLayout(event_group)
        
        # 力跳变阈值
        threshold_widget = QWidget()
        threshold_layout = QHBoxLayout(threshold_widget)
        threshold_layout.setContentsMargins(0, 0, 0, 0)
        
        threshold_label = QLabel("Jump Threshold (nm):")
        self.jump_threshold_input = QLineEdit("2.0")
        self.detect_jumps_button = QPushButton("Detect Jumps")
        self.detect_jumps_button.clicked.connect(self.detect_force_jumps)
        
        threshold_layout.addWidget(threshold_label)
        threshold_layout.addWidget(self.jump_threshold_input)
        threshold_layout.addWidget(self.detect_jumps_button)
        
        event_layout.addWidget(threshold_widget)
        
        # 移除几何校正相关代码
        
        advanced_layout.addWidget(event_group)
        
        # 添加数据导出按钮
        export_button = QPushButton("Export Analysis Results")
        export_button.clicked.connect(self.export_analysis_data)
        advanced_layout.addWidget(export_button)
        
        # 添加到主控制区
        self.controlsLayout.addWidget(self.advanced_panel)

    def export_analysis_data(self):
        """导出分析结果到Excel文件"""
        # 获取当前数据
        extension_data, time_data, force_data = self._get_current_data()
        
        # 确保有跳变事件数据
        if not hasattr(self, 'jumps_data') or len(self.jumps_data['time']) == 0:
            QMessageBox.warning(self, "No Data", "No jump events detected to export")
            return
        
        # 创建DataFrame
        df = pd.DataFrame(self.jumps_data)
        
        # 保存到Excel
        xlsx_file_path = os.path.join(self.Data_Saved_Path, self.base_name + '_jumps.xlsx')
        try:
            df.to_excel(xlsx_file_path, index=False)
            QMessageBox.information(self, "Export Successful", f"Jump events data exported to {xlsx_file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", f"Failed to export data: {str(e)}")

    def _calculate_averaged_data(self, mag_height, force_data, extension_data):
        """计算平均简化数据
        通过对磁铁高度进行差分，识别磁铁滞留区域，并计算每个区域的平均力和平均延伸值
        
        Args:
            mag_height: 磁铁高度数组
            force_data: 力数据数组
            extension_data: 延伸数据数组
            
        Returns:
            (avg_force, avg_extension): 平均力和平均延伸值数组
        """
        # 计算磁铁高度的差分
        height_diff = np.abs(np.diff(mag_height))
        
        # 寻找变化点（高度变化超过阈值的地方）
        change_threshold = 0.01  # 可以根据实际情况调整
        change_indices = np.where(height_diff > change_threshold)[0]
        
        # 添加开始和结束点以便处理所有段
        all_indices = np.concatenate(([0], change_indices, [len(mag_height)-1]))
        
        # 初始化结果数组
        avg_force = []
        avg_extension = []
        
        # 对每一段计算平均值
        for i in range(len(all_indices) - 1):
            start_idx = all_indices[i]
            end_idx = all_indices[i+1]
            
            # 只有当这一段包含足够多的点时才计算平均值
            if end_idx - start_idx > 5:  # 至少需要5个点
                segment_force = force_data[start_idx:end_idx]
                segment_extension = extension_data[start_idx:end_idx]
                
                # 计算平均值
                avg_f = np.nanmean(segment_force)
                avg_e = np.nanmean(segment_extension)
                
                # 添加到结果
                avg_force.append(avg_f)
                avg_extension.append(avg_e)
                
        return np.array(avg_force), np.array(avg_extension)

    def apply_wlc_correction(self, force_data, extension_data):
        """应用WLC correction，从测量的extension中减去DNA手柄的理论延伸长度，并自动平移到正值区域
        
        Args:
            force_data: 力数据数组
            extension_data: 延伸数据数组
            
        Returns:
            corrected_extension: 校正后的延伸数据（已平移到正值区域）
        """
        try:
            Lo = float(self.handle_length_input.text())
            Lp = float(self.Lp_input.text())
            T = float(self.T_input.text())
            
            # 计算DNA手柄的理论延伸长度
            handle_extension = WLC_inv(force_data, Lo, Lp, T)
            
            # 从测量的extension中减去手柄的延伸长度
            corrected_extension = extension_data - handle_extension
            
            # 如果校正后的最小值为负，将整个曲线平移到正值区域（最小值为1 nm）
            min_corrected = np.min(corrected_extension)
            if min_corrected < 0:
                offset = abs(min_corrected) + 1.0  # 额外加1 nm确保所有值为正
                corrected_extension += offset
            
            # 显示校正前后的差异信息
            avg_correction = np.mean(handle_extension)
            QMessageBox.information(self, "WLC Correction Applied", 
                                   f"Average correction: {avg_correction:.2f} nm\n"
                                   f"Original range: {np.min(extension_data):.2f}-{np.max(extension_data):.2f} nm\n"
                                   f"Corrected range: {np.min(corrected_extension):.2f}-{np.max(corrected_extension):.2f} nm\n"
                                   f"Applied offset: {offset:.2f} nm" if min_corrected < 0 else "No offset needed")
            
            return corrected_extension
        except ValueError as e:
            QMessageBox.warning(self, "Parameter Error", f"WLC correction failed: {str(e)}")
            return extension_data  # 如果发生错误，返回原始数据

    def select_fitting_range(self):
        """选择数据区间用于拟合"""
        # 先取消之前未完成的选择操作
        self.reset_fitting_selection()
        
        QMessageBox.information(self, "Select Fitting Range", 
                              "Please use middle mouse button to select two points on the plot to define the fitting range.\n"
                              "Press ESC key to cancel selection.")
        
        # 设置标志，表示我们正在选择拟合区间
        self.selecting_fit_range = True
        self.fit_range_points = []
        
        # 添加ESC键取消选择的事件处理
        if not hasattr(self, 'key_press_handler_connected'):
            self.canvas.mpl_connect('key_press_event', self._on_key_press)
            self.key_press_handler_connected = True

    def reset_fitting_selection(self):
        """重置拟合区间选择状态"""
        if hasattr(self, 'selecting_fit_range') and self.selecting_fit_range:
            self.selecting_fit_range = False
            self.fit_range_points = []
            QMessageBox.information(self, "Selection Reset", "Fitting range selection has been reset.")

    def _on_key_press(self, event):
        """处理键盘事件"""
        if event.key == 'escape' and hasattr(self, 'selecting_fit_range') and self.selecting_fit_range:
            self.reset_fitting_selection()
            QMessageBox.information(self, "Selection Cancelled", "Fitting range selection has been cancelled.")

    def apply_model_to_range(self, x_min, x_max):
        """对选定范围的数据应用模型并拟合"""
        # 获取当前数据
        extension_data, time_data, force_data = self._get_current_data()
        
        # 确保数据有效
        if len(force_data) < 5 or len(extension_data) < 5:
            QMessageBox.warning(self, "Insufficient Data", "Not enough data points for fitting")
            return
        
        # 获取选定范围内的数据
        range_mask = (force_data >= x_min) & (force_data <= x_max)
        range_force = force_data[range_mask]
        range_extension = extension_data[range_mask]
        
        # 确保选定范围内有足够的数据点
        if len(range_force) < 5 or len(range_extension) < 5:
            QMessageBox.warning(self, "Insufficient Data", "Not enough data points in the selected range for fitting")
            return
        
        # 获取参数
        try:
            Lo = float(self.Lo_input.text())
            Lp = float(self.Lp_input.text())
            T = float(self.T_input.text())
            Ko = float(self.Ko_input.text())
        except ValueError:
            QMessageBox.warning(self, "Parameter Error", "Please enter valid numerical parameters")
            return
        
        # 执行模型拟合
        try:
            if self.wlc_radio.isChecked():
                # 初始参数 [Lo, Lp]
                p0 = [Lo, Lp]
                bounds = ([0, 0], [10000, 1000])  # 参数范围
                
                # 拟合WLC模型
                popt, pcov = curve_fit(WLC_fit, range_force, range_extension, p0=p0, bounds=bounds)
                Lo_fit, Lp_fit = popt
                
                # 更新文本框
                self.Lo_input.setText(f"{Lo_fit:.2f}")
                self.Lp_input.setText(f"{Lp_fit:.2f}")
                
                # 生成拟合曲线
                force_range = np.linspace(x_min, x_max, 500)
                extension_range = WLC_inv(force_range, Lo_fit, Lp_fit, T)
                
                # 拟合结果标签
                fit_label = f"WLC Fit {len(self.fit_curves)+1} (Lo={Lo_fit:.2f}nm, Lp={Lp_fit:.2f}nm)"
                fit_params = {'Lo': Lo_fit, 'Lp': Lp_fit, 'model': 'WLC'}
            else:
                # 初始参数 [Lo, Lp, Ko]
                p0 = [Lo, Lp, Ko]
                bounds = ([0, 0, 0], [10000, 1000, 10000])
                
                # 拟合eWLC模型
                def fit_func(F, Lo, Lp, Ko):
                    return eWLC_inv(F, Lo, Lp, T, Ko)
                
                popt, pcov = curve_fit(fit_func, range_force, range_extension, p0=p0, bounds=bounds)
                Lo_fit, Lp_fit, Ko_fit = popt
                
                # 更新文本框
                self.Lo_input.setText(f"{Lo_fit:.2f}")
                self.Lp_input.setText(f"{Lp_fit:.2f}")
                self.Ko_input.setText(f"{Ko_fit:.2f}")
                
                # 生成拟合曲线
                force_range = np.linspace(x_min, x_max, 500)
                extension_range = eWLC_inv(force_range, Lo_fit, Lp_fit, T, Ko_fit)
                
                # 拟合结果标签
                fit_label = f"eWLC Fit {len(self.fit_curves)+1} (Lo={Lo_fit:.2f}nm, Lp={Lp_fit:.2f}nm, Ko={Ko_fit:.2f}pN)"
                fit_params = {'Lo': Lo_fit, 'Lp': Lp_fit, 'Ko': Ko_fit, 'model': 'eWLC'}
            
            # 为拟合曲线选择颜色和样式
            fit_index = len(self.fit_curves) % len(self.fit_colors)
            color = self.fit_colors[fit_index]
            # 为WLC和eWLC使用不同的线型，但颜色循环使用
            linestyle = '--' if self.wlc_radio.isChecked() else '-'
            style = dict(color=color, linestyle=linestyle, linewidth=2)
            
            # 绘制拟合曲线
            ax = self.fig.axes[0]
            line = ax.plot(force_range, extension_range, label=fit_label, **style)[0]
            ax.legend()
            self.canvas.draw_idle()
            
            # 存储拟合曲线数据
            fit_curve = {
                'x': force_range,
                'y': extension_range,
                'style': style,
                'label': fit_label,
                'params': fit_params,
                'range': [x_min, x_max],
                'data': {'force': range_force.tolist(), 'extension': range_extension.tolist()}
            }
            self.fit_curves.append(fit_curve)
            
            QMessageBox.information(self, "Fit Success", 
                                  f"Model fitted to data range [{x_min:.2f}, {x_max:.2f}]\n{fit_label}")
            
        except Exception as e:
            QMessageBox.critical(self, "Fit Error", f"Model fitting to range failed: {str(e)}")
            
        # 确保完成后清除选择状态
        self.reset_fitting_selection()

    def clear_all_fits(self):
        """清除所有拟合曲线"""
        if hasattr(self, 'fit_curves'):
            self.fit_curves = []
            self.plotfig()  # 重新绘制图表
            QMessageBox.information(self, "Fits Cleared", "All fit curves have been cleared")

    def export_fit_results(self):
        """导出拟合参数和曲线数据"""
        if not hasattr(self, 'fit_curves') or len(self.fit_curves) == 0:
            QMessageBox.warning(self, "No Fits", "No fit results to export")
            return
        
        xlsx_file_path = os.path.join(self.Data_Saved_Path, self.base_name + '_fit_results.xlsx')
        
        try:
            workbook = openpyxl.Workbook()
            params_sheet = workbook.active
            params_sheet.title = "Fit Parameters"
            # 添加样式颜色信息
            params_sheet.append(["Fit Number", "Model", "Range Start", "Range End", "Lo (nm)", "Lp (nm)", "Ko (pN)", "R^2", "Color", "Line Style"])
            
            for i, curve in enumerate(self.fit_curves):
                fit_num = i + 1
                model = curve['params']['model']
                range_start, range_end = curve['range']
                Lo = curve['params']['Lo']
                Lp = curve['params']['Lp']
                Ko = curve['params'].get('Ko', 'N/A')
                
                # 获取样式信息
                color = curve['style']['color'] if isinstance(curve['style'], dict) and 'color' in curve['style'] else 'Unknown'
                linestyle = curve['style']['linestyle'] if isinstance(curve['style'], dict) and 'linestyle' in curve['style'] else 'Unknown'
                
                # 计算R^2（拟合优度）- 改进版本
                if len(curve['data']['force']) > 0:
                    try:
                        force_data = np.array(curve['data']['force'])
                        ext_data = np.array(curve['data']['extension'])
                        
                        # 过滤掉可能导致问题的力值
                        valid_idx = force_data > 0.05  # 避免极小力值
                        if np.sum(valid_idx) > 5:  # 确保有足够的有效点
                            force_data = force_data[valid_idx]
                            ext_data = ext_data[valid_idx]
                            
                            if model == 'WLC':
                                predicted = WLC_fit(force_data, Lo, Lp)
                            else:
                                T = float(self.T_input.text())
                                predicted = eWLC_inv(force_data, Lo, Lp, T, Ko)
                            
                            # 检查预测值中是否有无效值
                            valid_pred = ~np.isnan(predicted) & ~np.isinf(predicted)
                            if np.sum(valid_pred) > 5:
                                ext_valid = ext_data[valid_pred]
                                pred_valid = predicted[valid_pred]
                                
                                ss_res = np.sum((ext_valid - pred_valid) ** 2)
                                ss_tot = np.sum((ext_valid - np.mean(ext_valid)) ** 2)
                                r2 = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0
                            else:
                                r2 = 'N/A (预测值无效)'
                        else:
                            r2 = 'N/A (有效点太少)'
                    except Exception as calc_err:
                        r2 = f'N/A (计算错误: {str(calc_err)[:20]})'
                else:
                    r2 = 'N/A (无数据)'
                
                # 写入参数到总表（包含样式信息）
                params_sheet.append([fit_num, model, range_start, range_end, Lo, Lp, Ko, r2, color, linestyle])
                
                # 创建单独的工作表
                data_sheet = workbook.create_sheet(f"Fit {fit_num}")
                data_sheet.append(["Force (pN)", "Extension (nm)", ""])
                
                # 限制点数量以避免过大的Excel文件
                max_points = 1000
                data_points = len(curve['data']['force'])
                step = max(1, data_points // max_points)
                
                for j in range(0, len(curve['data']['force']), step):
                    data_sheet.append([curve['data']['force'][j], curve['data']['extension'][j], ''])
                
                # 写入拟合曲线 - 同样限制点数
                data_sheet.append(['', '', ''])  # 空行分隔
                data_sheet.append(["Curve Force (pN)", "Curve Extension (nm)", ''])
                
                curve_points = len(curve['x'])
                curve_step = max(1, curve_points // max_points)
                
                for j in range(0, len(curve['x']), curve_step):
                    data_sheet.append([curve['x'][j], curve['y'][j], ''])
            
            workbook.save(xlsx_file_path)
            
            # 保存图像时降低dpi减少资源消耗
            fig_file_path = os.path.join(self.Data_Saved_Path, self.base_name + '_fit_results.png')
            self.fig.savefig(fig_file_path, dpi=150, bbox_inches='tight')
            
            QMessageBox.information(self, "Export Successful", 
                                   f"Fit results exported to:\n{xlsx_file_path}\n\nFigure saved to:\n{fig_file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export fit results: {str(e)}")
