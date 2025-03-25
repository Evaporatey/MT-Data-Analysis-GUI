# author: Ye Yang
# Force Calibration Module
# This module is used to calibrate the force data from magnetic beads
# The data contains the x, y, z direction data of the magnetic beads, and the detailed state of the magnet movement
# -*- coding: utf-8 -*-



import math
import os
import matplotlib
matplotlib.rcParams['font.sans-serif'] = ['Arial Unicode MS', 'SimHei', 'DejaVu Sans']
matplotlib.rcParams['axes.unicode_minus'] = False

import matplotlib.pyplot as plt
import nptdms as nt
import numpy as np
import openpyxl
import pandas as pd
import scipy
from PySide6.QtWidgets import (QComboBox, QHBoxLayout, QPushButton, QSizePolicy, QVBoxLayout, QWidget, QCheckBox,
                               QSpinBox, QMessageBox, QRadioButton, QButtonGroup, QLineEdit, QLabel)
from matplotlib.backends.backend_qtagg import NavigationToolbar2QT, FigureCanvasQTAgg
from scipy.optimize import curve_fit
from scipy import signal
from tdms_reader import read_tdms_chunk, read_tdms_file


# 这里的数据包含磁球x,y,z方向的数据，同时包含了磁铁运动的详细状态，可以将每一步的数据进行拆分并校准。
# x方向的PSD校准还没有完善，比较复杂

def force_calibration_func(x, a1, b1, a2, b2, c):
    return a1 * np.exp(b1 * x) + a2 * np.exp(b2 * x) + c

# PSD分析相关函数
def lorentzian(f, A, fc, offset):
    """洛伦兹函数用于拟合PSD曲线
    f: 频率
    A: 振幅
    fc: 特征频率 (角频率)
    offset: 偏移量
    """
    return A / (1 + (f/fc)**2) + offset

def calculate_psd(position_data, sampling_rate):
    """计算位置数据的功率谱密度
    position_data: 位置数据时间序列
    sampling_rate: 采样率 (Hz)
    返回: 频率和PSD值
    """
    # 去除线性趋势
    detrended_data = signal.detrend(position_data)
    
    # 使用Welch方法计算PSD
    frequencies, psd = signal.welch(detrended_data, fs=sampling_rate, 
                                   nperseg=min(1024, len(detrended_data)//8),
                                   scaling='density')
    return frequencies, psd

def calculate_force_from_psd(fc, gamma, T=298):
    """根据PSD拟合得到的特征频率计算力
    fc: 特征频率 (Hz)
    gamma: 阻尼系数 (取决于磁珠尺寸和液体粘度)
    T: 温度 (K)，默认为室温298K
    """
    kb = 1.38e-23  # 玻尔兹曼常数 (J/K)
    force = 2 * np.pi * kb * T * fc / gamma  # 单位为牛顿(N)
    return force * 1e12  # 转换为pN


class ForceCalibration(QWidget):  # create new figure view widget
    def __init__(self, data_for_figure):  # initialize figure view widget
        super().__init__()  # call super class

        self.Data_Saved_Path = data_for_figure['Data_Saved_Path']
        self.file_name = data_for_figure['file_name']
        self.file_type = data_for_figure['file_type']
        self.file_info = data_for_figure['self.file_info']
        self.base_name = data_for_figure['base_name']

        self.tdms_data_frame = pd.DataFrame()
        self.tdms_data_frame = read_tdms_file(self.file_name, need_force=False)
        self.tdms_data_store = self.tdms_data_frame

        self.num_of_data = (len(self.tdms_data_frame.columns) - 7) / 3

        self.beads_list = []
        self.beads_list = self.tdms_data_frame.columns.values.tolist()

        self.beads_list_x = []
        self.beads_list_y = []
        self.beads_list_z = []
        for i in range(4, 4 + int(self.num_of_data)):
            self.beads_list_x.append(self.beads_list[i])
        for i in range(5 + int(self.num_of_data), 5 + int(self.num_of_data) * 2):
            self.beads_list_y.append(self.beads_list[i])
        for i in range(6 + int(self.num_of_data) * 2, 6 + int(self.num_of_data) * 3):
            self.beads_list_z.append(self.beads_list[i])

        self.tdms_data_store_x = self.tdms_data_store[self.beads_list_x]
        self.tdms_data_store_y = self.tdms_data_store[self.beads_list_y]
        self.tdms_data_store_z = self.tdms_data_store[self.beads_list_z]

        self.setWindowTitle("Force-Extension Analysis")  # set window title
        self.resize(2000, 2000)  # resize widget
        # ---------------------------------create central widget of FigureView-------------------------
        self.centralwidget = QWidget()  # create central widget of main window
        self.centralwidget.setObjectName(u"Figure View")  # set object name of central widget
        self.centralwidgetlayout = QVBoxLayout(self.centralwidget)

        self.horizontalLayoutWidget = QWidget(self.centralwidget)
        self.horizontalLayoutWidget.setObjectName(u"horizontalLayoutWidget")

        self.horizontalLayout = QHBoxLayout(self.horizontalLayoutWidget)
        self.horizontalLayout.setObjectName(u"horizontalLayout")
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)

        # 添加校准方法选择
        self.method_group = QButtonGroup(self.horizontalLayoutWidget)
        self.method_group.setObjectName('method_group')
        
        self.traditional_method = QRadioButton('Variance Method')
        self.traditional_method.setObjectName('traditional_method')
        self.traditional_method.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.traditional_method.setChecked(True)
        
        self.psd_method = QRadioButton('PSD Analysis Method')
        self.psd_method.setObjectName('psd_method')
        self.psd_method.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        
        self.method_group.addButton(self.traditional_method)
        self.method_group.addButton(self.psd_method)
        
        self.horizontalLayout.addWidget(self.traditional_method)
        self.horizontalLayout.addWidget(self.psd_method)
        
        # 方向选择
        self.radio_box_calibration_group = QButtonGroup(self.horizontalLayout)
        self.radio_box_calibration_group.setObjectName('radio_box_calibration_group')

        self.radio_box_y_calibration = QRadioButton('Y Direction Calibration')
        self.radio_box_y_calibration.setObjectName('radio_y_calibration')
        self.radio_box_y_calibration.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.radio_box_x_calibration = QRadioButton('X Direction calibration')
        self.radio_box_x_calibration.setObjectName('radio_x_calibration')
        self.radio_box_x_calibration.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.radio_box_calibration_group.addButton(self.radio_box_y_calibration)
        self.radio_box_calibration_group.addButton(self.radio_box_x_calibration)

        self.radio_box_y_calibration.setChecked(True)
        self.horizontalLayout.addWidget(self.radio_box_y_calibration)
        self.horizontalLayout.addWidget(self.radio_box_x_calibration)
        
        # 添加采样率设置
        self.sampling_rate_label = QLabel("Sampling Rate (Hz):")
        self.sampling_rate_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.horizontalLayout.addWidget(self.sampling_rate_label)
        
        self.sampling_rate_input = QLineEdit(self.horizontalLayoutWidget)
        self.sampling_rate_input.setObjectName("sampling_rate_input")
        self.sampling_rate_input.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.sampling_rate_input.setText("200")  # 默认采样率
        self.horizontalLayout.addWidget(self.sampling_rate_input)

        # 添加DNA长度输入框
        self.dna_length_label = QLabel("DNA Length (nm):")
        self.dna_length_label.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.horizontalLayout.addWidget(self.dna_length_label)

        self.dna_length_input = QLineEdit(self.horizontalLayoutWidget)
        self.dna_length_input.setObjectName("dna_length_input")
        self.dna_length_input.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.dna_length_input.setText("3400")  # 默认10kbp DNA长度
        self.horizontalLayout.addWidget(self.dna_length_input)

        self.magnet_beads_radius = QLineEdit(self.horizontalLayoutWidget)
        self.magnet_beads_radius.setObjectName("magnet_bead_radius")
        self.magnet_beads_radius.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.magnet_beads_radius.setText(
            "Input Radius of Magnet Beads in X Direction Calibration (nm) [M270/M280 = 1400 nm]")

        self.horizontalLayout.addWidget(self.magnet_beads_radius)

        self.y_axis_box = QComboBox(self.horizontalLayoutWidget)
        self.y_axis_box.setObjectName(u"y_axis_box")
        self.y_axis_box.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.horizontalLayout.addWidget(self.y_axis_box)

        self.x_axis_box = QComboBox(self.horizontalLayoutWidget)
        self.x_axis_box.setObjectName(u"x_axis_box")
        self.x_axis_box.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.horizontalLayout.addWidget(self.x_axis_box)
        self.x_axis_box.setEnabled(False)

        for i in range(0, 3):
            self.x_axis_box.addItem(str(self.beads_list[i]))

        for i in range(len(self.beads_list_z)):
            self.y_axis_box.addItem(str(self.beads_list_z[i]))

        # ---------------------------data slicing--------------------------------------------------
        str_magnet_move = str(self.beads_list[-1])

        magnet_move_state = self.tdms_data_store[str(str_magnet_move)].values.tolist()

        new_magnet_move_state = [i for i in magnet_move_state if math.isnan(i) == False]

        self.new_int_magnet_move_state = [int(i) for i in new_magnet_move_state]

        self.num_of_state = len(self.new_int_magnet_move_state) / 2

        self.final_slice_magnet_move_state = [self.new_int_magnet_move_state[i: i + 2] for i in range(0,
                                                                                                      len(self.new_int_magnet_move_state),
                                                                                                      2)]
        self.num_of_sliced_data = len(self.final_slice_magnet_move_state)
        sliced_data_num_list = list(range(1, self.num_of_sliced_data + 1))

        self.y_axis_box.currentTextChanged.connect(self.plotfig)
        self.x_axis_box.currentTextChanged.connect(self.plotfig)
        
        # 为PSD方法添加UI更新
        self.psd_method.toggled.connect(self.update_ui_for_method)
        self.traditional_method.toggled.connect(self.update_ui_for_method)

        self.remove_bad_data_button = QPushButton(self.horizontalLayoutWidget)
        self.remove_bad_data_button.setObjectName(u"remove_bad_data_button")
        self.remove_bad_data_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.horizontalLayout.addWidget(self.remove_bad_data_button)
        self.remove_bad_data_button.setText("Remove")
        self.remove_bad_data_button.clicked.connect(self.remove_bad_data)

        self.check_fitted_data_box = QCheckBox(self.horizontalLayoutWidget)
        self.check_fitted_data_box.setObjectName(u"check_fitted_data_box")
        self.horizontalLayout.addWidget(self.check_fitted_data_box)
        self.check_fitted_data_box.setChecked(True)
        self.check_fitted_data_box.stateChanged.connect(self.plotfig)
        self.check_fitted_data_box.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.check_fitted_data_box.setText("Fitted Data")

        self.kernel_size_box = QSpinBox(self.horizontalLayoutWidget)
        self.kernel_size_box.setObjectName(u"kernel_size_box")
        self.horizontalLayout.addWidget(self.kernel_size_box)
        self.kernel_size_box.setRange(1, 100)
        self.kernel_size_box.setValue(3)
        self.kernel_size_box.setSingleStep(2)
        self.kernel_size_box.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.kernel_size_box.valueChanged.connect(self.plotfig)
        
        # 将UI初始化移到这里，确保所有UI控件都已创建
        self.update_ui_for_method()

        self.calibration_button = QPushButton(self.horizontalLayoutWidget)
        self.calibration_button.setObjectName(u"calibration_button")
        self.horizontalLayout.addWidget(self.calibration_button)
        self.calibration_button.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)
        self.calibration_button.setText("Calibration")
        self.calibration_button.clicked.connect(self.force_calibration_calculation)

        self.centralwidgetlayout.addWidget(self.horizontalLayoutWidget)
        self.setLayout(self.centralwidgetlayout)

        self.fig = plt.figure()  # create figure

        def wheelEvent(event):
            if self.y_axis_box.underMouse():
                if event.angleDelta().y() > 0:
                    self.y_axis_box.setCurrentIndex((self.y_axis_box.currentIndex() - 1) % self.y_axis_box.count())
                else:
                    self.y_axis_box.setCurrentIndex((self.y_axis_box.currentIndex() + 1) % self.y_axis_box.count())
            elif self.kernel_size_box.underMouse():
                if event.angleDelta().y() > 0:
                    self.kernel_size_box.setValue(self.kernel_size_box.value() - 2)
                else:
                    self.kernel_size_box.setValue(self.kernel_size_box.value() + 2)
            else:
                super().wheelEvent(event)

        self.y_axis_box.wheelEvent = wheelEvent

        # ------------------zoom in and out------------------------
        def zoom_event(event):  # zoom event
            axtemp = event.inaxes  # get axes
            x_min, x_max = axtemp.get_xlim()  # get x limits
            y_min, y_max = axtemp.get_ylim()  # get y limits
            x_range = x_max - x_min  # get x range
            y_range = y_max - y_min  # get y range
            x_zoom = x_range / 10  # get x zoom
            y_zoom = y_range / 10  # get y zoom

            if event.button == 'up':  # if scroll up
                axtemp.set(xlim=(x_min + x_zoom, x_max - x_zoom),
                           ylim=(y_min + y_zoom, y_max - y_zoom))  # zoom in
            elif event.button == 'down':  # if scroll down
                axtemp.set(xlim=(x_min - x_zoom, x_max + x_zoom),
                           ylim=(y_min - y_zoom, y_max + y_zoom))  # zoom out
            self.canvas.draw_idle()  # draw canvas

        self.canvas = FigureCanvasQTAgg(self.fig)
        self.canvas.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.toolbar = NavigationToolbar2QT(self.canvas, self)  # create toolbar for figure view widget and add

        self.canvas.mpl_connect('scroll_event', zoom_event)  # connect zoom event to canvas

        # -------------move figure with right-click--------------------
        self.lastx = 0  # initialize last x
        self.lasty = 0  # initialize last y
        self.press = False  # initialize press

        def on_press(event):
            if event.inaxes != None:
                if event.button == 3:
                    self.lastx = event.xdata
                    self.lasty = event.ydata
                    self.press = True

        def on_move(event):
            axtemp = event.inaxes
            if axtemp != None:
                if self.press == True:
                    x = event.xdata - self.lastx
                    y = event.ydata - self.lasty

                    x_min, x_max = axtemp.get_xlim()
                    y_min, y_max = axtemp.get_ylim()

                    x_min -= x
                    x_max -= x
                    y_min -= y
                    y_max -= y

                    axtemp.set(xlim=(x_min, x_max), ylim=(y_min, y_max))
                    self.canvas.draw_idle()

        def on_release(event):
            if self.press:
                self.press = False

        self.canvas.mpl_connect('button_press_event', on_press)
        self.canvas.mpl_connect('button_release_event', on_release)
        self.canvas.mpl_connect('motion_notify_event', on_move)

        self.vlayoutwidget = QWidget(self)  # create vertical layout widget
        self.vlayoutwidget.setObjectName(u"vlayoutwidget")  # set object name of vertical layout widget

        self.plotfig()

        # 连接radio按钮的信号
        self.radio_box_y_calibration.toggled.connect(self.update_direction_ui)
        self.radio_box_x_calibration.toggled.connect(self.update_direction_ui)

        # 初始化UI状态
        self.update_ui_for_method()
        self.update_direction_ui()

    def update_ui_for_method(self):
        """根据选择的校准方法更新UI元素的状态"""
        is_psd = self.psd_method.isChecked()
        
        # PSD方法设置项可见性
        self.sampling_rate_label.setVisible(is_psd)
        self.sampling_rate_input.setVisible(is_psd)
        self.dna_length_label.setVisible(is_psd)
        self.dna_length_input.setVisible(is_psd)
        
        # PSD方法强制使用y方向校准
        if is_psd:
            self.radio_box_y_calibration.setChecked(True)
            self.radio_box_x_calibration.setEnabled(False)
            self.kernel_size_box.setEnabled(False)
            self.magnet_beads_radius.setEnabled(False)
        else:
            self.radio_box_x_calibration.setEnabled(True)
            self.kernel_size_box.setEnabled(True)
            self.magnet_beads_radius.setEnabled(True)

    def update_direction_ui(self):
        """根据选择的方向更新UI元素状态"""
        is_x_direction = self.radio_box_x_calibration.isChecked()
        
        # X方向校准需要磁珠半径参数
        self.magnet_beads_radius.setVisible(is_x_direction)
        
        # 如果选择了PSD方法，强制使用Y方向
        if self.psd_method.isChecked() and is_x_direction:
            self.radio_box_y_calibration.setChecked(True)

    def remove_bad_data(self):
        self.current_index_num = self.y_axis_box.findText(self.y_axis_box.currentText())

        self.tdms_data_store_x = self.tdms_data_store_x.drop(str(self.beads_list_x[self.current_index_num]), axis=1)
        self.beads_list_x.remove(self.beads_list_x[self.current_index_num])
        self.tdms_data_store_y = self.tdms_data_store_y.drop(str(self.beads_list_y[self.current_index_num]), axis=1)
        self.beads_list_y.remove(self.beads_list_y[self.current_index_num])
        self.tdms_data_store_z = self.tdms_data_store_z.drop(str(self.beads_list_z[self.current_index_num]), axis=1)
        self.beads_list_z.remove(self.beads_list_z[self.current_index_num])

        self.y_axis_box.setCurrentIndex((self.y_axis_box.currentIndex() + 1) % self.y_axis_box.count())
        self.y_axis_box.removeItem(self.current_index_num)

    def plotfig(self):  # plot figure
        self.chosen_bead = self.y_axis_box.currentText()
        self.chosen_x_data = self.x_axis_box.currentText()

        self.x_axis = self.tdms_data_store[str(self.chosen_x_data)]
        self.y_axis = self.tdms_data_store[str(self.chosen_bead)]
        self.xxx = self.x_axis.drop(0)
        self.xx = self.xxx.values
        self.yyy = self.y_axis.drop(0)
        self.yy = self.yyy.values

        plt.clf()

        if self.check_fitted_data_box.isChecked():
            self.kernel_size_box.setEnabled(True)

            if self.vlayoutwidget.isVisible():
                self.vlayoutwidget.deleteLater()
            else:
                pass

            self.vlayoutwidget = QWidget(self)  # create vertical layout widget
            self.vlayoutwidget.setObjectName(u"vlayoutwidget")  # set object name of vertical layout widget
            self.vlayout = QVBoxLayout(self.vlayoutwidget)  # create vertical layout
            self.vlayout.setObjectName(u"vlayout")  # set object name of vertical layout
            self.vlayout.setContentsMargins(0, 0, 0, 0)  # set contents margins of vertical layout
            self.vlayout.addWidget(self.canvas)  # add canvas to vertical layout
            self.vlayout.addWidget(self.toolbar)  # add toolbar to vertical layout
            self.centralwidgetlayout.addWidget(self.vlayoutwidget)

            ax = self.fig.add_subplot(111)  # add subplot to figure
            axis_x = self.xx  # set x axis
            axis_y = self.yy  # set y axis

            ax.plot(axis_x, axis_y, color='darkgrey', label='Raw Data')  # plot x and y axis

            set_size = self.kernel_size_box.value()  # get kernel size

            y_fitted = scipy.signal.medfilt(axis_y, kernel_size=set_size)  # apply median filter to y axis
            ax.plot(axis_x, y_fitted, color='red', label='Fitted Data')

            plt.legend(frameon=False, loc='upper right')  # show legend

        else:
            self.kernel_size_box.setEnabled(False)

            if self.vlayoutwidget.isVisible():
                self.vlayoutwidget.deleteLater()
            else:
                pass

            self.vlayoutwidget = QWidget(self)  # create vertical layout widget
            self.vlayoutwidget.setObjectName(u"vlayoutwidget")  # set object name of vertical layout widget
            self.vlayout = QVBoxLayout(self.vlayoutwidget)  # create vertical layout
            self.vlayout.setObjectName(u"vlayout")  # set object name of vertical layout
            self.vlayout.setContentsMargins(0, 0, 0, 0)  # set contents margins of vertical layout
            self.vlayout.addWidget(self.canvas)  # add canvas to vertical layout
            self.vlayout.addWidget(self.toolbar)  # add toolbar to vertical layout
            self.centralwidgetlayout.addWidget(self.vlayoutwidget)

            ax = self.fig.add_subplot(111)  # add subplot to figure
            axis_x = self.xx  # set x axis
            axis_y = self.yy  # set y axis

            ax.plot(axis_x, axis_y, color='darkgrey', label='Raw Data')  # plot x and y axis

            plt.legend(frameon=False, loc='upper right')  # show legend

    def force_calibration_calculation(self):
        """根据选择的方法进行力校准"""
        # 如果选择了PSD方法且尝试进行X方向校准，提示并返回
        if self.psd_method.isChecked() and self.radio_box_x_calibration.isChecked():
            QMessageBox.warning(self, 'Warning', 'PSD method only supports Y direction calibration. Please select Y direction calibration.')
            return
            
        # 根据选择的方法执行不同的校准函数
        if self.psd_method.isChecked():
            self.psd_force_calibration()
        else:
            self.traditional_force_calibration()  # 原有的方差法校准
    
    def psd_force_calibration(self):
        """使用PSD方法进行力校准(仅支持Y方向)"""
        try:
            # 获取UI输入参数
            sampling_rate = float(self.sampling_rate_input.text())
            if sampling_rate <= 0:
                raise ValueError("Sampling rate must be positive")
                
            # 从输入框获取DNA长度
            L_DNA = float(self.dna_length_input.text())  # nm
            if L_DNA <= 0:
                raise ValueError("DNA length must be positive")
                
            # 默认使用1400nm作为磁珠半径 (M270/M280磁珠)
            Rbead = 1400  # nm
                
        except ValueError as e:
            QMessageBox.warning(self, 'Warning', f'Please enter valid parameters! {str(e)}')
            return
        
        # 配置参数
        kb = 1.38e-23      # 玻尔兹曼常数 (J/K)
        T = 298            # 温度 (K)
        kbT = kb * T       # 热能 (J)
        eta = 8.9e-4       # 水的粘度 (Pa·s)
        
        # 获取磁铁高度数据
        self.magnets_height_index = self.beads_list[2]
        self.magnets_height = self.tdms_data_store[str(self.magnets_height_index)]
        self.final_magnets_height_segment = []
        
        # 处理磁铁高度数据
        for i in range(len(self.final_slice_magnet_move_state)):
            section_of_selected_data = self.final_slice_magnet_move_state[i]
            start_point = section_of_selected_data[0] - 1 if section_of_selected_data[0] != 0 else 0
            end_point = section_of_selected_data[1] - 1
            
            magnet_height_segment = self.magnets_height[start_point:end_point]
            if len(magnet_height_segment) > 2000:
                # 保存磁铁高度，单位保持mm
                magnets_height_mm = np.mean(magnet_height_segment)
                self.final_magnets_height_segment.append(magnets_height_mm)
        
        # 定义表面修正系数计算函数
        def calc_surface_correction(L, R):
            """计算表面修正系数"""
            Lr = L/R
            C_par = 1/(1 - 9/16*(1+Lr)**(-1) + 1/8*(1+Lr)**(-3) - 45/256*(1+Lr)**(-4) - 1/16*(1+Lr)**(-5))
            C_rot = 1 + 5/16*(1+Lr)**(-3)
            return C_par, C_rot
        
        # 定义双模式耦合PSD函数 (简化版本，基于Daldrop et al. 2015)
        def coupled_psd_model(f, F, A, offset):
            """双模式耦合PSD函数，包含别名修正"""
            # 计算表面修正系数
            C_par, C_rot = calc_surface_correction(L, Rbead)
            
            # 计算阻尼系数
            gamma_y = 6*np.pi*eta*Rbead*1e-9 * C_par
            gamma_phi = 8*np.pi*eta*(Rbead*1e-9)**3 * C_rot
            
            # 转换单位
            F_N = F * 1e-12  # 力 (N)
            L_m = L * 1e-9   # 长度 (m)
            R_m = Rbead * 1e-9  # 半径 (m)
            
            # 计算特征频率
            term1 = F_N/(L_m)/(2*np.pi)
            term2 = ((L_m+R_m)*R_m)/(2*gamma_phi) + 1/(2*gamma_y)
            term3 = np.sqrt(((L_m+R_m)*R_m/gamma_phi + 1/gamma_y)**2 - 4*L_m*R_m/(gamma_y*gamma_phi))
            
            f_low = term1 * (term2 - 0.5*term3)
            f_high = term1 * (term2 + 0.5*term3)
            C = 2*np.pi*f_low*L_m/F_N - (L_m+R_m)*R_m/gamma_phi
            
            # 计算PSD基础值
            psd_base = 4*kbT/(2*np.pi)**2/(1+C**2*gamma_y*gamma_phi/R_m**2)
            
            # 初始化PSD数组
            psd = np.zeros_like(f)
            
            # 考虑别名效应 (主频率和-1别名)
            for n in [0, -1]:
                f_alias = np.abs(f + n*sampling_rate)
                # Lorentzian项
                lorentzian_term = (gamma_phi*C**2/R_m**2/(f_low**2+f_alias**2) + 
                                1/gamma_y/(f_high**2+f_alias**2))
                # 采样sinc修正
                sinc_term = np.ones_like(f_alias)
                nonzero = f_alias != 0
                sinc_term[nonzero] = (np.sin(np.pi*f_alias[nonzero]/sampling_rate)/
                                    (np.pi*f_alias[nonzero]/sampling_rate))**2
                
                # 添加到总PSD
                psd += A * psd_base * lorentzian_term * sinc_term
            
            # 返回并添加白噪声背景
            return psd + offset
        
        # 创建Excel保存结果
        xlsx_file_path = os.path.join(self.Data_Saved_Path, f"{self.base_name}_psd_calibration.xlsx")
        if os.path.exists(xlsx_file_path):
            workbook = openpyxl.load_workbook(xlsx_file_path)
        else:
            workbook = openpyxl.Workbook()
            # 删除默认的Sheet
            if 'Sheet' in workbook.sheetnames:
                workbook.remove(workbook['Sheet'])
        
        # 处理每个珠子数据
        for j in range(len(self.beads_list_z)):
            bead_name_z = self.beads_list_z[j]
            bead_name_y = self.beads_list_y[j]
            
            current_bead_name = bead_name_z.split('/')[-1]  # 提取珠子名称
            
            # 为当前珠子创建专用文件夹
            bead_folder = os.path.join(self.Data_Saved_Path, f"{self.base_name}_psd_{current_bead_name}")
            os.makedirs(bead_folder, exist_ok=True)
            
            # 存储结果数据
            heights = []        # Magnet heights
            forces_y = []       # Calculated forces
            z_positions = []    # Bead average z positions
            f_low_values = []   # Low frequency characteristic frequencies
            f_high_values = []  # High frequency characteristic frequencies
            
            # 处理每一段数据（不同的磁铁高度）
            for i in range(len(self.final_slice_magnet_move_state)):
                if i >= len(self.final_magnets_height_segment):
                    continue
                    
                section_of_selected_data = self.final_slice_magnet_move_state[i]
                start_point = section_of_selected_data[0] - 1 if section_of_selected_data[0] != 0 else 0
                end_point = section_of_selected_data[1] - 1
                
                # 获取y,z方向数据
                z_data = self.tdms_data_store_z[bead_name_z].values[start_point:end_point]
                y_data = self.tdms_data_store_y[bead_name_y].values[start_point:end_point]
                
                if len(y_data) < 2000:  # 数据太少，跳过
                    continue
                
                # 计算DNA长度 (z位置+零点校正)
                z_mean = np.mean(z_data)
                z_positions.append(z_mean)
                L = max(100, z_mean if z_mean > 0 else L_DNA)  # Ensure DNA length is positive
                
                # 记录当前高度
                current_height = self.final_magnets_height_segment[i]
                heights.append(current_height)
                
                # 数据预处理 - 使用线性去趋势
                y_detrended = signal.detrend(y_data, type='linear')
                
                # 高效计算PSD - 使用多段平均提高稳定性
                segment_len = min(8192, len(y_detrended)//4)  # 更大的FFT窗口提高低频分辨率
                
                # 使用Welch方法计算PSD (重叠50%)
                freqs, psd_total = signal.welch(y_detrended, fs=sampling_rate, 
                                            nperseg=segment_len, 
                                            noverlap=segment_len//2, 
                                            scaling='density')
                
                try:
                    # 估算初始力值 (使用方差法)
                    var_y = np.var(y_detrended)
                    F_init = kb * T * L * 1e21 / var_y  # 转换为pN
                    F_init = min(max(0.1, F_init), 50)  # 限制在合理范围内
                    
                    # 选择适当的频率范围进行拟合
                    # 低频段：0.5-10Hz (避开DC和高频噪声)
                    low_freq_mask = (freqs >= 0.5) & (freqs <= min(10, sampling_rate/4))
                    
                    # 中高频段：10-Nyquist/2 (避开最高频噪声和别名)
                    high_freq_mask = (freqs > 10) & (freqs < sampling_rate/2.5)
                    
                    # 组合频率掩码
                    freq_mask = low_freq_mask | high_freq_mask
                    
                    # 平滑PSD曲线
                    window_size = min(11, len(psd_total)//10*2+1)
                    if window_size > 3:  # 确保窗口大小足够进行平滑
                        psd_smooth = signal.savgol_filter(psd_total, window_size, 3)
                    else:
                        psd_smooth = psd_total
                    
                    # 改进初始参数估计
                    A_init = np.max(psd_smooth[low_freq_mask]) * 0.5  # 振幅初始估计
                    offset_init = np.median(psd_smooth[-10:])  # 白噪声底线
                    
                    # 使用更稳健的拟合方法 (先使用低频数据拟合)
                    try:
                        popt, pcov = curve_fit(
                            lambda f, F, A, offset: coupled_psd_model(f, F, A, offset),
                            freqs[freq_mask],
                            psd_smooth[freq_mask],
                            p0=[F_init, A_init, offset_init],
                            bounds=([0.01, 0, 0], [100, np.inf, np.inf]),
                            maxfev=50000
                        )
                        
                        # 提取拟合结果
                        F_fit, A_fit, offset_fit = popt
                        
                        # 计算特征频率
                        C_par, C_rot = calc_surface_correction(L, Rbead)
                        gamma_y = 6*np.pi*eta*Rbead*1e-9 * C_par
                        gamma_phi = 8*np.pi*eta*(Rbead*1e-9)**3 * C_rot
                        
                        F_N = F_fit * 1e-12
                        L_m = L * 1e-9
                        R_m = Rbead * 1e-9
                        
                        term1 = F_N/(L_m)/(2*np.pi)
                        term2 = ((L_m+R_m)*R_m)/(2*gamma_phi) + 1/(2*gamma_y)
                        term3 = np.sqrt(((L_m+R_m)*R_m/gamma_phi + 1/gamma_y)**2 - 4*L_m*R_m/(gamma_y*gamma_phi))
                        
                        f_low = term1 * (term2 - 0.5*term3)
                        f_high = term1 * (term2 + 0.5*term3)
                        
                        # 存储结果
                        forces_y.append(F_fit)
                        f_low_values.append(f_low)
                        f_high_values.append(f_high)
                        
                        # 生成拟合曲线并绘图
                        plt.figure(figsize=(10, 6))
                        
                        # 使用对数刻度绘制PSD
                        plt.loglog(freqs, psd_total, 'b-', alpha=0.5, label='Original PSD')
                        
                        # 高亮显示拟合区域
                        plt.loglog(freqs[freq_mask], psd_smooth[freq_mask], 'g.', alpha=0.7, label='Fitting Data Points')
                        
                        # 计算拟合曲线
                        freqs_fit = np.logspace(np.log10(freqs[1]), np.log10(freqs[-1]), 1000)
                        psd_fit = coupled_psd_model(freqs_fit, F_fit, A_fit, offset_fit)
                        plt.loglog(freqs_fit, psd_fit, 'r-', linewidth=2, label='Fitted Curve')
                        
                        # 标记特征频率
                        plt.axvline(x=f_low, color='cyan', linestyle='--', label=f'Low Freq.: {f_low:.2f} Hz')
                        plt.axvline(x=f_high, color='purple', linestyle='--', label=f'High Freq.: {f_high:.2f} Hz')
                        
                        # 添加图例和标签
                        plt.title(f'Magnet Height: {current_height:.2f} mm | Force: {F_fit:.2f} pN | DNA Length: {L:.0f} nm')
                        plt.xlabel('Frequency (Hz)')
                        plt.ylabel('PSD (nm²/Hz)')
                        plt.xlim([0.1, sampling_rate/2])
                        plt.ylim([np.min(psd_smooth)/10, np.max(psd_smooth)*5])
                        plt.legend(loc='best', fontsize=9)
                        plt.grid(True, which='both', linestyle='--', alpha=0.3)
                        
                        # 保存图像到珠子专用文件夹
                        plt.savefig(os.path.join(bead_folder, f"height_{current_height:.2f}mm_force_{F_fit:.2f}pN.png"), dpi=150)
                        plt.close()
                        
                    except Exception as e:
                        print(f"Fitting error (Height {current_height:.2f} mm): {str(e)}")
                        continue
                
                except Exception as e:
                    print(f"PSD calculation error (Height {current_height:.2f} mm): {str(e)}")
                    continue
            
            # 结果为空的情况处理
            if len(forces_y) == 0:
                QMessageBox.warning(self, 'Warning', f'Bead {current_bead_name} has no valid fitting results!')
                continue
                    
            # 在Excel中保存结果
            if current_bead_name in workbook.sheetnames:
                workbook.remove(workbook[current_bead_name])
            worksheet = workbook.create_sheet(current_bead_name)
                
            # 设置列标题
            worksheet.cell(1, 1, "Magnet Height (mm)")
            worksheet.cell(1, 2, "Force (pN)")
            worksheet.cell(1, 3, "Low Freq. (Hz)")
            worksheet.cell(1, 4, "High Freq. (Hz)")
            worksheet.cell(1, 5, "DNA Length (nm)")
            
            # 保存数据
            for i in range(len(heights)):
                worksheet.cell(i + 2, 1, heights[i])
                worksheet.cell(i + 2, 2, forces_y[i])
                worksheet.cell(i + 2, 3, f_low_values[i])
                worksheet.cell(i + 2, 4, f_high_values[i])
                worksheet.cell(i + 2, 5, z_positions[i])
            
            # 拟合力与磁铁高度关系
            try:
                # 按高度排序数据
                sorted_indices = np.argsort(heights)
                heights_sorted = [heights[i] for i in sorted_indices]
                forces_sorted = [forces_y[i] for i in sorted_indices]
                
                # 拟合力和高度关系
                popt, _ = curve_fit(force_calibration_func, 
                                np.array(heights_sorted), 
                                np.array(forces_sorted),
                                p0=[10, -0.5, 10, -0.1, 0.1],
                                bounds=([0, -10, 0, -10, -5], [100, 0, 100, 0, 5]),
                                maxfev=20000)
                    
                # 保存拟合参数
                worksheet.cell(1, 7, "Force-Height Fit Parameters: F=a₁*exp(b₁*h)+a₂*exp(b₂*h)+c")
                var_names = ['a₁', 'b₁', 'a₂', 'b₂', 'c']
                for i, name in enumerate(var_names):
                    worksheet.cell(2, 7+i, name)
                    worksheet.cell(3, 7+i, f"{popt[i]:.6f}")
                    
                # 绘制力-高度关系图
                plt.figure(figsize=(10, 8))
                plt.scatter(heights_sorted, forces_sorted, s=70, marker='o', color='blue', 
                        label='PSD Analysis Results')
                    
                # 生成拟合曲线 (使用更密集的点以显示平滑曲线)
                h_fit = np.linspace(min(heights_sorted)-0.5, max(heights_sorted)+0.5, 200)
                f_fit = force_calibration_func(h_fit, *popt)
                plt.plot(h_fit, f_fit, 'r-', linewidth=2, label='Fitted Curve')
                    
                # 添加数据点标签
                for i, (h, f) in enumerate(zip(heights_sorted, forces_sorted)):
                    plt.text(h, f+0.2, f'{f:.1f}pN', fontsize=8, ha='center')
                    
                plt.xlabel('Magnet Height (mm)', fontsize=12)
                plt.ylabel('Force (pN)', fontsize=12)
                plt.title(f'PSD Force Calibration - Bead: {current_bead_name}', fontsize=14)
                plt.legend(fontsize=12)
                plt.grid(True, alpha=0.3)
                    
                # 添加参数注释
                equation = f'F = {popt[0]:.2f}*exp({popt[1]:.3f}*h) + {popt[2]:.2f}*exp({popt[3]:.3f}*h) + {popt[4]:.2f}'
                plt.annotate(equation, xy=(0.05, 0.95), xycoords='axes fraction', 
                            bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="gray", alpha=0.8))
                    
                # 保存到珠子专用文件夹
                plt.savefig(os.path.join(bead_folder, "force_height_calibration.png"), dpi=300)
                plt.close()
                    
                # 创建汇总图表
                plt.figure(figsize=(12, 10))
                
                # 创建2x2网格
                gs = plt.GridSpec(2, 2, height_ratios=[2, 1])
                
                # 1. 力-高度关系图
                ax1 = plt.subplot(gs[0, :])
                ax1.scatter(heights_sorted, forces_sorted, s=70, marker='o', color='blue')
                ax1.plot(h_fit, f_fit, 'r-', linewidth=2)
                ax1.set_xlabel('Magnet Height (mm)')
                ax1.set_ylabel('Force (pN)')
                ax1.set_title(f'Force-Height Calibration - {current_bead_name}')
                ax1.grid(True, alpha=0.3)
                
                # 2. 特征频率随高度变化
                ax2 = plt.subplot(gs[1, 0])
                ax2.scatter([heights[i] for i in sorted_indices], [f_low_values[i] for i in sorted_indices], 
                        label='Low Freq.')
                ax2.scatter([heights[i] for i in sorted_indices], [f_high_values[i] for i in sorted_indices], 
                        label='High Freq.')
                ax2.set_xlabel('Magnet Height (mm)')
                ax2.set_ylabel('Characteristic Freq. (Hz)')
                ax2.legend()
                ax2.grid(True, alpha=0.3)
                
                # 3. DNA长度随高度变化
                ax3 = plt.subplot(gs[1, 1])
                ax3.scatter([heights[i] for i in sorted_indices], [z_positions[i] for i in sorted_indices])
                ax3.set_xlabel('Magnet Height (mm)')
                ax3.set_ylabel('DNA Length (nm)')
                ax3.grid(True, alpha=0.3)
                
                plt.tight_layout()
                plt.savefig(os.path.join(bead_folder, "calibration_summary.png"), dpi=300)
                plt.close()
                    
            except Exception as e:
                print(f"Force-height fitting error: {str(e)}")
                    
        # 保存Excel
        workbook.save(xlsx_file_path)
        
        QMessageBox.information(self, 'Complete', f'PSD analysis force calibration completed! Results saved to each bead folder.')


        
    def traditional_force_calibration(self):
        """Original traditional variance method for force calibration"""
        final_beads_num = len(self.tdms_data_store_z.columns.tolist())

        self.magnets_height_index = self.beads_list[2]
        self.magnets_height = self.tdms_data_store[str(self.magnets_height_index)]
        self.final_magnets_height_segment = []

        for i in range(len(self.final_slice_magnet_move_state)):
            section_of_selected_data = self.final_slice_magnet_move_state[i]
            start_point = section_of_selected_data[0]
            if start_point != 0:
                start_point = section_of_selected_data[0] - 1
            else:
                pass
            end_point = section_of_selected_data[1] - 1
            self.magnets_height_segment = self.magnets_height[start_point:end_point]
            if len(self.magnets_height_segment) > 2000:
                # 先存储原始值（单位：mm）
                self.magnets_height_segment_mean_mm = np.mean(self.magnets_height_segment)
                # 转换为nm用于内部计算
                self.magnets_height_segment_mean = np.mean(self.magnets_height_segment) * 1e6
                self.final_magnets_height_segment.append(self.magnets_height_segment_mean)

        self.force_calibration_results = []

        for j in range(final_beads_num):
            self.beads_name_on_calibration_z = self.beads_list_z[j]
            self.beads_name_on_calibration_y = self.beads_list_y[j]
            self.beads_name_on_calibration_x = self.beads_list_x[j]

            self.z_data_segment_final = []
            self.y_data_segment_final = []
            self.x_data_segment_final = []

            self.force_calibration_x_axis_data = []
            self.force_calibration_y_axis_data = []
            self.force_calibration_y = []
            self.force_calibration_x = []

            self.current_bead_name = self.tdms_data_store_z.columns.tolist()[j]
            self.on_calibration_data_z = self.tdms_data_store_z[str(self.beads_name_on_calibration_z)].values.tolist()
            self.on_calibration_data_y = self.tdms_data_store_y[str(self.beads_name_on_calibration_y)].values.tolist()
            self.on_calibration_data_x = self.tdms_data_store_x[str(self.beads_name_on_calibration_x)].values.tolist()

            for i in range(len(self.final_slice_magnet_move_state)):
                section_of_selected_data = self.final_slice_magnet_move_state[i]
                start_point = section_of_selected_data[0]
                if start_point != 0:
                    start_point = section_of_selected_data[0] - 1
                else:
                    pass
                end_point = section_of_selected_data[1] - 1

                self.z_data_segment = self.on_calibration_data_z[start_point:end_point]
                if len(self.z_data_segment) > 2000:
                    self.z_data_segment_final.append(self.z_data_segment)
                
                self.y_data_segment = self.on_calibration_data_y[start_point:end_point]
                if len(self.y_data_segment) > 2000:
                    self.y_data_segment_final.append(self.y_data_segment)
                
                self.x_data_segment = self.on_calibration_data_x[start_point: end_point]
                if len(self.x_data_segment) > 2000:
                    self.x_data_segment_final.append(self.x_data_segment)

            self.start_z_segment = self.z_data_segment_final[0]
            self.end_z_segment = self.z_data_segment_final[-1]

            if min(self.start_z_segment) > min(self.end_z_segment):
                self.zero_point = min(self.end_z_segment)
            else:
                self.zero_point = min(self.start_z_segment)

            self.total_z_data = []
            self.total_y_data = []
            self.total_x_data = []
            for i in range(len(self.z_data_segment_final)):
                self.current_z_segment = self.z_data_segment_final[i]
                for j in range(len(self.current_z_segment)):
                    self.current_z_segment[j] = self.current_z_segment[j] - self.zero_point
                self.total_z_data.append(np.mean(self.current_z_segment))
                self.total_y_data.append(np.var(self.y_data_segment_final[i]))
                self.total_x_data.append(np.var(self.x_data_segment_final[i]))

            self.x_data = self.final_magnets_height_segment
            self.y_data = []  # 初始化为空列表
            
            # 计算力值
            self.force_calibration_y = []  # 确保初始化这些列表
            self.force_calibration_x = []
            
            kb = 1.38e-23  # 玻尔兹曼常数 (J/K)
            T = 298  # 室温 (K)
            kbT = kb * T * 1e21  # 转换为pN·nm

            for i in range(len(self.final_magnets_height_segment)):
                if self.radio_box_y_calibration.isChecked():
                    # 确保除数不为零
                    if self.total_y_data[i] > 0:
                        self.force_calibration_y_axis_data = kbT * self.total_z_data[i] / self.total_y_data[i]
                        self.force_calibration_y.append(self.force_calibration_y_axis_data)
                elif self.radio_box_x_calibration.isChecked():
                    try:
                        self.beads_radius = float(self.magnet_beads_radius.text())
                        if self.total_x_data[i] > 0:
                            self.force_calibration_x_axis_data = kbT * (self.total_z_data[i] + self.beads_radius) / self.total_x_data[i]
                            self.force_calibration_x.append(self.force_calibration_x_axis_data)
                    except ValueError:
                        QMessageBox.warning(self, 'Warning', 'Please enter a valid bead radius!')
                        return

            # 确保正确设置 self.y_data
            if self.radio_box_y_calibration.isChecked():
                self.y_data = self.force_calibration_y
            else:
                self.y_data = self.force_calibration_x
                
            # 检查 self.y_data 是否为空
            if not self.y_data:
                QMessageBox.warning(self, 'Warning', 'No valid force data! Please check parameter settings.')
                return
                
            # 拟合部分保持不变...
            try:
                popt_y, pcov_y = curve_fit(force_calibration_func, self.x_data, self.y_data, 
                                          p0=[0, 0, 0, 0, 0], maxfev=50000)
            except RuntimeError:
                QMessageBox.warning(self, 'Warning', 'Fitting failed, please check data and try again.')
                return

            # 保存结果到Excel文件
            direction = 'y' if self.radio_box_y_calibration.isChecked() else 'x'
            xlsx_file_path = os.path.join(self.Data_Saved_Path, f"{self.base_name}_{direction}_calibration.xlsx")

            if os.path.exists(xlsx_file_path):
                workbook = openpyxl.load_workbook(xlsx_file_path)
                if self.current_bead_name in workbook.sheetnames:
                    workbook.remove(workbook[self.current_bead_name])
                
                worksheet = workbook.create_sheet(self.current_bead_name)
            else:
                workbook = openpyxl.Workbook()
                worksheet = workbook.active
                worksheet.title = self.current_bead_name

            # 设置列标题
            worksheet.cell(1, 1, "Magnet Height (mm)")
            worksheet.cell(1, 2, "Force (pN)")
            worksheet.cell(1, 4, "Fitting Parameters (Unit: x in nm, F in pN)")

            # 保存拟合参数在右侧列
            var_names = ['a1', 'b1', 'a2', 'b2', 'c']
            for i, name in enumerate(var_names):
                worksheet.cell(2, 4+i, name)
                worksheet.cell(3, 4+i, f"{popt_y[i]:.6f}")

            # 调整拟合参数为mm单位的值
            popt_y_mm = popt_y.copy()
            popt_y_mm[1] = popt_y[1] * 1e6  # b1调整
            popt_y_mm[3] = popt_y[3] * 1e6  # b2调整

            worksheet.cell(5, 4, "Fitting Parameters (for x in mm):")
            for i, name in enumerate(var_names):
                worksheet.cell(6, 4+i, name)
                worksheet.cell(7, 4+i, f"{popt_y_mm[i]:.6f}")

            # 保存原始数据
            for i in range(len(self.x_data)):
                # 保存时转回mm
                worksheet.cell(i + 2, 1, self.x_data[i] / 1e6)  # 将nm转回mm
                worksheet.cell(i + 2, 2, self.y_data[i])        # 力保持pN单位

            workbook.save(xlsx_file_path)
            
            # 绘制校准曲线
            plt.figure(figsize=(10, 8))
            plt.scatter(np.array(self.x_data) / 1e6, self.y_data, label='Raw Data', color='blue')  # 显示为mm
            
            # 生成拟合曲线
            x_fit = np.linspace(min(self.x_data), max(self.x_data), 100)
            y_fit = force_calibration_func(x_fit, *popt_y)
            plt.plot(x_fit / 1e6, y_fit, 'r-', label='Fitted Curve')  # 显示为mm
            
            plt.xlabel('Magnet Height (mm)')
            plt.ylabel('Force (pN)')
            plt.title(f'{direction.upper()} Direction Force Calibration - {self.current_bead_name}')
            plt.legend()
            plt.grid(True)
            
            # 保存图片
            fig_path = os.path.join(self.Data_Saved_Path, 
                                 f"{self.base_name}_var_{direction}_{self.current_bead_name}.png")
            plt.savefig(fig_path, dpi=300)
            plt.close()

        QMessageBox.information(self, 'Complete', 'Traditional variance method force calibration completed!')
